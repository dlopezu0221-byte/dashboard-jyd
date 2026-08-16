#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
MODELOS.py — Genera dashboards individuales por modelo
Fuente: Cómo vamos Fornax2.xlsx
Salida: dashboard-jyd/modelos/{slug}/index.html
"""

import os, sys, re, base64, json
from datetime import date, datetime, timedelta

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
FORNAX2 = os.path.join(SCRIPT_DIR, '..', 'Centro de Gestión Estratégica Grupo J&D',
                        'COMO VAMOS GRUPO', 'Cómo vamos Fornax2.xlsx')
GE_XLSX  = os.path.join(SCRIPT_DIR, '..', 'Centro de Gestión Estratégica Grupo J&D',
                        'COMO VAMOS GRUPO', 'Cómo vamos Grupo Empresarial.xlsx')
GRUPO_HTML = os.path.join(SCRIPT_DIR, 'grupo579780', 'index.html')
STR_FACTOR = 20.0  # Streamate en Excel = USD → créditos: × 20  (÷ 0.05)

MESES = ['ENERO','FEBRERO','MARZO','ABRIL','MAYO','JUNIO',
         'JULIO','AGOSTO','SEPTIEMBRE','OCTUBRE','NOVIEMBRE','DICIEMBRE']
MESES_CAP = {m: m.capitalize() if m != 'ENERO' else 'Enero' for m in MESES}
MESES_CAP.update({'ENERO':'Enero','FEBRERO':'Febrero','MARZO':'Marzo','ABRIL':'Abril',
                  'MAYO':'Mayo','JUNIO':'Junio','JULIO':'Julio','AGOSTO':'Agosto',
                  'SEPTIEMBRE':'Septiembre','OCTUBRE':'Octubre','NOVIEMBRE':'Noviembre','DICIEMBRE':'Diciembre'})
Q2_END = {'ENERO':'31','FEBRERO':'28','MARZO':'31','ABRIL':'30','MAYO':'31','JUNIO':'30',
          'JULIO':'31','AGOSTO':'31','SEPTIEMBRE':'30','OCTUBRE':'31','NOVIEMBRE':'30','DICIEMBRE':'31'}

BONO_USD = {1:0, 2:8, 3:12, 4:18, 5:22, 6:25}
MAX_BONO_CYCLE = 85  # Jul Q1(0)+Q2(8)+AgoQ1(12)+Q2(18)+SepQ1(22)+Q2(25)

# RE histórico por modelo: {modelo: {mes: re_valor}}
# Cuando el RE de un modelo cambió durante el año, se registra aquí el valor anterior.
# Para meses no listados se usa el RE actual de BASE DE DATOS (profile['re']).
# PINs de acceso por modelo — NO modificar sin actualizar CODIGOS_ACCESO_MODELOS_PRIVADO.csv
MODEL_PINS = {
    'aisha-collin':       ('770487', 'mdl_aisha_collin'),
    'angeline-smith':     ('216739', 'mdl_angeline_smith'),
    'antonella-cooper':   ('126225', 'mdl_antonella_cooper'),
    'arianna-kylee':      ('877572', 'mdl_arianna_kylee'),
    'camila-bustamante':  ('388389', 'mdl_camila_bustamante'),
    'canela-jhonson':     ('356787', 'mdl_canela_jhonson'),
    'cynthia-adams':      ('334053', 'mdl_cynthia_adams'),
    'dahiana-walcott':    ('246316', 'mdl_dahiana_walcott'),
    'danna-diamond':      ('872246', 'mdl_danna_diamond'),
    'dulce-luna':         ('207473', 'mdl_dulce_luna'),
    'emilly-beaumont':    ('809570', 'mdl_emilly_beaumont'),
    'erza-elric':         ('876646', 'mdl_erza_elric'),
    'isa-raven':          ('671858', 'mdl_isa_raven'),
    'isabella-winkler':   ('191161', 'mdl_isabella_winkler'),
    'jack-kum':           ('719176', 'mdl_jack_kum'),
    'juli-saenz':         ('542417', 'mdl_juli_saenz'),
    'katherine-bond':     ('133326', 'mdl_katherine_bond'),
    'katt-souza':         ('131244', 'mdl_katt_souza'),
    'koban-and-damian':   ('198246', 'mdl_koban_and_damian'),
    'leorico':            ('329258', 'mdl_leorico'),
    'liam-terrier':       ('343962', 'mdl_liam_terrier'),
    'lucia-brown':        ('629903', 'mdl_lucia_brown'),
    'maddy-parisi':       ('731262', 'mdl_maddy_parisi'),
    'maximo-marcelo':     ('127824', 'mdl_maximo_marcelo'),
    'melany-oconner':     ('688508', 'mdl_melany_oconner'),
    'mia-monrroe':        ('308496', 'mdl_mia_monrroe'),
    'mioku-doll':         ('850800', 'mdl_mioku_doll'),
    'naty-roxx':          ('781453', 'mdl_naty_roxx'),
    'nicolas-dwayne':     ('835392', 'mdl_nicolas_dwayne'),
    'nicolle-lopez':      ('671412', 'mdl_nicolle_lopez'),
    'sarah-delucca':      ('539898', 'mdl_sarah_delucca'),
    'sofia-blaze':        ('331148', 'mdl_sofia_blaze'),
    'sophi-duval':        ('571029', 'mdl_sophi_duval'),
    'vanesa-foxy':        ('717889', 'mdl_vanesa_foxy'),
    'wheeler-green':      ('391704', 'mdl_wheeler_green'),
    'william-gardener':   ('948749', 'mdl_william_gardener'),
    'zeus-strong':        ('106814', 'mdl_zeus_strong'),
}

PRIMER_SEMESTRE = {'ENERO','FEBRERO','MARZO','ABRIL','MAYO','JUNIO'}
HISTORICAL_RE = {
    # RE anterior al cambio de Q1 Agosto: aplica para H1 2025 + JULIO
    # Nuevo RE (profile['re']) rige desde Q1 AGOSTO en adelante (automático)
    'Emilly Beaumont':  {**{m: 22000 for m in PRIMER_SEMESTRE}, 'JULIO': 22000},
    'Isabella Winkler': {**{m: 25000 for m in PRIMER_SEMESTRE}, 'JULIO': 25000},
    'Antonella Cooper': {**{m: 27000 for m in PRIMER_SEMESTRE}, 'JULIO': 27000},
}

LOGO_B64 = "iVBORw0KGgoAAAANSUhEUgAAAFcAAAAtCAYAAADbcffLAAAI00lEQVR4nO2abYydRRXHf+e5u1u27VJo3da6bS0l1qJgRDGNFgtowkvANkLEpmpisqIBAV9iSBTFaIRIidUoviRgiMakkPKBpCQao7WWFxGhsVEDUm1LKyCCxDZ933vv3w9zZu/c2efuvZveGqj3JE/muTPnnDnznzNnzsxzoUc96lGPetSjk4CsrFJSkbXJzOolfOZ847xmVuu2ka9VKgW3jCRVsqq6manL9pxU1ASupIqZ1SR9GFgLVIE+4HlgU8Iv4DTgPGA2MAJUgBeAT5nZAUnWAz8hSX1e3qpmusfrLeEtJM2WdJakNZIec97hnPf/lfpa1B8CasBhYBD4j4PVJ6kK4DH4FX+eAu6VdCnBq3tEa3ALwjIff8xMCmt9HLzEOwtCDP5FbOuFhNbgdkQJgDVoxOzjtuokoeMCN6cc2CRVg5DOTdmbEx0dyad9lqWPk/F3KtMpdQxuNKJso8oHHfNkBzsNIxVapHBpzuwZS3yvRx2SCjOrZ+FIaZ3rbuqzbDW5jaT6EzvG9Wb1MMnEZTZ0DG4EZKwdYwTA3+cA892gZ81sv9dPGHAKSqIjZh9zgOfNbH+W4sVwNF4n6Q3AgOvcHScq3ysSG19PSCv7gJqZPVWmN5FX3lb2G8IMdUIzJc2TNOJlfOZ6OZiAVpf0Rkl3ARsd3CXAJknrJA37gCtxpiUNS3pE0iZJF7qOZZJuANYTcuwXJX3C+T8u6XFJ90la4pvtWyRdA3wb+DuwQ9IfJH3U24sUBEnLJd0NXAbMIOTqn5X0pKT1kuY7n0n6nqTfS/qdpM1Jn1Fn4b/n+hlhfGWMU5LnfsVz1iNePiFpraSbJY1Ker+kVQ7WlyXNTzo6V9KLLvfeRPelXrdT0hVZfwvVoP2ue5WkRT64IUm3S9rp/Lck/A+57kskDXn7qKSjCc9yrx9wfddLeknSebkX+VijnW/yuivVTM9IGlRwkMKfuW7XEu/DcsU5uIe8XN/OtV3hYkn/cpnNUaekfn/f6m37JC1Qw3MXeF9j+aRkfazw8iZJVUnHFBxgacITJ/kh56lJutfrTvHyh5I+5u/9Dk5s+1oyMaOJ3o2S6gkmX4zyXq6WtDBiAZ2HhWjAQDJT6VPxePMjYJgQlx5U4wIoxqJfA3XgVOCOLEYVhLhXB/ZGvcnEVczskYS3AvQTDjz/dH4DCu/3t85TALNcrurlOuBnzlclbJxHFMLbNYSY/TjwaOKBX3LeAUJMvkXSW81sTNK1wHYz2ysPEVMBN+6cdTNregiW1SS9A7jEDagBD0eZBNwtDnYdWC1pxOvTJWTAQLqLm5m8j4ES2wqgP929/f1AwhNBjfp2RdlE9yLgTmA7sMbMlvvmFveSHcAGwoRVgVOAWyV9BNhtZrudb9yOTsGdjKKO97nxFeAosNPrU3BfIYBXIxyrr/D6/MatVT7bKgct40915hdU/Q5qVWG/uBhYAXzXzC4zs/ucL3qh3Mu/ALxM8N5jwGpgxMx+7jqbMqBugBsHtpLG3a7RnOZFnn3A/qTfd3eh/3Z2NVcG7xpTyFDWETKaPWa2wcy2p+EoWZl1gpe/BHyf5tvBqx34mrJNrBvgRpqejyO+JLniHmAPDa9qmzd3mQoPAVcB24DrgbVm9rRvvIWHu5gqzkhkI3h3Ery3QliB7wRG4wQ0ddZFww9lv9PryZgVLAIWuVHQ5eN3O3KPHQXuBxYAN5nZHl/SVc+v48Z4JnCR2x037IqZvQzcQQA3hrzbJc0OrA3v7Qa4UdlW7yg+1RKeWYRMIcbOJ7rQfycU07NlhIymTtgXtjgYaSyPQN4ALPQ6c6+uSno7wZEeI8TeMeB04Fr33vFY3wrcqVywRMM209ispgFL1Dinx6R6luuOfFsyHZ3SVC+A4jjPpLGci6St4mEgxuMFwChhf4AQTuoKx/mLCRP0SQKwheu7UeGoXo+5ditw+10gT6UmkHdaMbNtwE8JS92A8+NSouENV7q+PuBuM/uLq6lnTzuKuWY7/siT3kXEie0HLjezmpkd87ImaT5wFyGbOehjPCZpJiGj+A4hNf0TcJuP5SgwF/ime+/ErzBqnNCu81NI3csfpO0lcuaxakjh7C1J/5A0LeEZkvSCt21wT4n9jaiZlnp9kfUTDxVXJ7yHPd5FO6LOmxOeX3rdXIVjrxROYYcl/VjSByRd5XZ8y9urCncVSJrj+qJd0dMHJT2b2X52me0pUBVJ35C0S9JeSV+fDNwo5+UMhSPkc5IeVvjGdq6kByT9TdLnfCIsmZSKpBsl7XBjz2gBbpSZKel+t+3Pkk7L2k0hf93oPPckOj7k40rpj5K+Kmm6pHdJOuj1Ox3wC3LQkvJtCsf6XQoOdVvEatKPiArnZiNcxbX9wqDmq7/pwHsIcatCONFsjXpS3kQ+flYaK7vzLelvgLBEW6Z0zlP3zSjGzlOBcwgh499m9kwms4xwzbkP+GvUn9ucjXea66uY2ZF2huenpo5IydJs0d6n/MZoiv3F1dUBz4SlWSaX6iuzrU190aqtnec23axPhZTc6NPYOGvZzMcbfyUytOsz85gJK6BsDCSfihLboLHZpXYUCX9HmczxYPWqoSTWzVb4ipDGe1P55U4q/z/7P0U3T2il5MtmsULu2LRcJV0k6ZxYn8jM9vAxmIPhMXMxIcm/IBtHBXhdrs9/ny3p9MRzY32fpDPkd7HdpBMGbhLbPgh8hvCHkry/EcKJLZf5PDCP8HepAW+LE3IhcL7bnm9kA8Ayf08ncRWwFFgpaV7W10rg09G+bnr2ifTc+EX2N4TT2womHkaeI9yf5jRIwwujTBz0sOvbBszM5ArC97Cc3gw8SDg8DHlcjPY9CvzK7esqnUhw40YznQDIgZJgfxRYk/yO7Q8AlxP+5Fd1ECIYWwir4SzgSa+LcseApxNdMQT8BLgOGAJ2ZZvPMGEFHWTi5L/6SeHo2Kqtv0V9V2/M2qSHZd7+2qEWue1kf76eVEZlX1g7kJmKfT3qUY961KMe9ahHJ5b+C3JMSIByJtokAAAAAElFTkSuQmCC"

# ── helpers ──────────────────────────────────────────────────────────────────

def inject_pin(html, slug):
    """Inyecta protección por PIN al inicio del <body>. Idempotente."""
    if slug not in MODEL_PINS:
        return html
    pin, storage_key = MODEL_PINS[slug]
    if f"sessionStorage.getItem('{storage_key}')" in html:
        return html  # ya tiene PIN
    script = f"""<script>
// ── ACCESO PROTEGIDO — Dashboard Individual ──────────────────────────────────
(function() {{
  var _AUTH = false;
  try {{ _AUTH = sessionStorage.getItem('{storage_key}') === 'ok_{pin}'; }} catch(e) {{}}
  if (_AUTH) return;
  document.body.style.display = 'none';
  var overlay = document.createElement('div');
  overlay.id = '_pin_overlay';
  overlay.style.cssText = 'position:fixed;top:0;left:0;width:100%;height:100%;background:#0F172A;display:flex;align-items:center;justify-content:center;z-index:99999;font-family:Arial,sans-serif';
  overlay.innerHTML = [
    '<div style="background:#1E293B;border:1px solid #334155;border-radius:16px;padding:40px 32px;width:320px;max-width:90vw;text-align:center;box-shadow:0 25px 50px rgba(0,0,0,.6)">',
    '<div style="font-size:32px;margin-bottom:12px">\\uD83D\\uDD10</div>',
    '<div style="color:#F1F5F9;font-size:18px;font-weight:700;margin-bottom:6px">Dashboard Privado</div>',
    '<div style="color:#64748B;font-size:13px;margin-bottom:24px">Ingresa tu c\\u00F3digo de acceso</div>',
    '<input id="_pin_inp" type="password" inputmode="numeric" pattern="[0-9]*" maxlength="6" placeholder="\\u2022 \\u2022 \\u2022 \\u2022 \\u2022 \\u2022" style="width:100%;box-sizing:border-box;background:#0F172A;border:1px solid #334155;border-radius:8px;color:#F1F5F9;font-size:22px;letter-spacing:6px;padding:12px;text-align:center;outline:none;margin-bottom:12px">',
    '<div id="_pin_err" style="display:none;color:#EF4444;font-size:12px;margin-bottom:10px">C\\u00F3digo incorrecto. Intenta de nuevo.</div>',
    '<button id="_pin_btn" style="width:100%;background:#3B82F6;color:#fff;border:none;border-radius:8px;padding:13px;font-size:15px;font-weight:600;cursor:pointer;touch-action:manipulation">Ingresar</button>',
    '<div style="color:#475569;font-size:11px;margin-top:20px">Grupo Empresarial J&D \\u00B7 Acceso restringido</div>',
    '</div>'
  ].join('');
  document.body.parentNode.insertBefore(overlay, document.body);
  function _check() {{
    var val = (document.getElementById('_pin_inp').value || '').trim();
    if (val === '{pin}') {{
      try {{ sessionStorage.setItem('{storage_key}', 'ok_{pin}'); }} catch(e) {{}}
      document.getElementById('_pin_overlay').remove();
      document.body.style.display = '';
    }} else {{
      document.getElementById('_pin_err').style.display = 'block';
      document.getElementById('_pin_inp').value = '';
    }}
  }}
  document.getElementById('_pin_btn').addEventListener('click', _check);
  document.getElementById('_pin_inp').addEventListener('keydown', function(e) {{ if (e.key === 'Enter') _check(); }});
}})();
// ── FIN ACCESO PROTEGIDO ──────────────────────────────────────────────────────
</script>"""
    return html.replace('<body>\n<nav', '<body>\n' + script + '\n<nav', 1)

def fmt(n):
    if n is None: return "0"
    try:
        n = float(n)
        return f"{int(round(n)):,}".replace(",", ".")
    except: return "0"

def pct(a, b):
    if not b: return 0.0
    return round((a / b) * 100, 1)

def slugify(name):
    s = name.lower()
    for a, b in [('á','a'),('é','e'),('í','i'),('ó','o'),('ú','u'),('ñ','n'),('ü','u')]:
        s = s.replace(a, b)
    s = re.sub(r'[^a-z0-9]+', '-', s).strip('-')
    return s

def num(v):
    if v is None: return 0
    try: return float(v)
    except: return 0

# ── Excel reading ─────────────────────────────────────────────────────────────

def load_wb():
    import openpyxl
    path = os.path.normpath(FORNAX2)
    if not os.path.exists(path):
        print(f"  ❌ No se encontró Fornax2: {path}")
        sys.exit(1)
    return openpyxl.load_workbook(path, read_only=True, data_only=True)

def rows(wb, sheet):
    if sheet not in wb.sheetnames: return []
    return list(wb[sheet].iter_rows(values_only=True))

def find_model_row(row_list, model_name, col=0):
    for r in row_list:
        if r and r[col] == model_name:
            return r
    return None

# ── Profile ───────────────────────────────────────────────────────────────────

def read_profile(wb, model_name):
    """Read model profile from Base de datos modelos.
    Columnas reales (índice 0-based):
      0=fuente  1=Estado  2=Nivel  3=Nombre artístico  4=Nombre real
      5=Fuente_ingreso  6=Fecha_ingreso  7=Modalidad  8=%pago
      9=Quien_refiere  10=Monitor  11=RE_asignado  12=None
      13=Fecha_retiro  14=Motivo  15=None  16=None
      17=Estudio/modelo  18=Ejecutivo_Comercial
    """
    for r in rows(wb, 'Base de datos modelos'):
        if not r: continue
        if len(r) >= 12 and r[3] == model_name:
            # Todos los modelos de este dashboard pertenecen a Fornax Studios.
            # Col R contiene el aliado/referidor, NO el estudio del modelo.
            studio = 'Fornax Studios'
            ingreso = ''
            if len(r) > 6 and r[6]:
                try:
                    from datetime import datetime
                    if hasattr(r[6], 'strftime'):
                        ingreso = r[6].strftime('%d/%m/%Y')
                    else:
                        ingreso = str(r[6])
                except Exception:
                    ingreso = str(r[6])
            return {
                'nombre': model_name,
                'nombre_real': str(r[4]).strip() if r[4] else '',
                'nivel': r[2] or 'Órbita',
                'estado': r[1] or 'Activo',
                'modalidad': r[7] or 'Planta',
                'monitor': str(r[10]).strip() if r[10] else '',
                're': int(num(r[11])) if len(r) > 11 and r[11] else 50000,
                'ejecutivo': str(r[18]).strip() if len(r) > 18 and r[18] else '',
                'studio': studio,
                'ingreso': ingreso,
                'plataforma': 'Flirt4Free',  # se actualiza desde monthly data
            }
    return {'nombre': model_name, 'nombre_real': '', 'nivel': 'Órbita', 'estado': 'Activo',
            'modalidad': 'Planta', 'monitor': '', 're': 50000, 'ejecutivo': '',
            'studio': 'Fornax Studios', 'ingreso': '', 'plataforma': ''}

# ── Monthly data ──────────────────────────────────────────────────────────────

def detect_nplats(wb, mes_upper):
    """Detect number of platforms in a monthly sheet (4 or 5)."""
    rr = rows(wb, mes_upper)
    for r in rr:
        if r and r[0] == 'Nombre modelo':
            return 5 if len(r) > 16 else 4
        if r and isinstance(r[0], str) and r[0] not in (None,''):
            # Check if model name repeats at col 13 (4 plats) or 16 (5 plats)
            if len(r) > 16 and r[16] == r[0]: return 5
            if len(r) > 13 and r[13] == r[0]: return 4
    # fallback: check header row
    for r in rr[:5]:
        if r and 'Streamate' in r: return 5
    return 4

def read_monthly(wb, mes_upper, model_name):
    """Read Q1/Q2 totals for a given month from the monthly sheet."""
    np = detect_nplats(wb, mes_upper)
    rr = rows(wb, mes_upper)
    r = find_model_row(rr, model_name)
    if r is None:
        return {'q1':0,'q2':0,'total':0,'q1_plats':{},'q2_plats':{},'n_plats':np}

    if np == 5:
        q1_vals = [num(r[i]) for i in range(1, 6)]
        q2_vals = [num(r[i]) for i in range(6, 11)]
        tot_vals = [num(r[i]) for i in range(11, 16)]
    else:
        q1_vals = [num(r[i]) for i in range(1, 5)]
        q2_vals = [num(r[i]) for i in range(5, 9)]
        tot_vals = [num(r[i]) for i in range(9, 13)]

    plat_keys = ['f4f','sc','cb','cs','str'][:np if np==5 else 4]
    return {
        'q1': sum(q1_vals),
        'q2': sum(q2_vals),
        'total': sum(tot_vals),
        'q1_plats': dict(zip(plat_keys, q1_vals)),
        'q2_plats': dict(zip(plat_keys, q2_vals)),
        'n_plats': np,
    }

def read_quincena_status(wb, mes_upper, model_name, q='q1'):
    """Read RE status from quincena sheet. Returns (credits, re, status_str, monitor)."""
    mc = MESES_CAP[mes_upper]
    end = Q2_END[mes_upper]
    if q == 'q1':
        candidates = [f'{mc} - Periodo 1 al 15', f'{mc} - Periodo 1 al 15 ']
    else:
        candidates = [f'{mc} - Periodo 16 al {end}']

    for sname in candidates:
        rr = rows(wb, sname)
        for r in rr:
            if not r or len(r) < 8: continue
            if r[1] == model_name:
                re_val = int(num(r[0]))
                # detect 4 or 5 platforms: if r[7] is status string → 4 plats (total at r[6])
                # if r[8] is status string → 5 plats (total at r[7])
                if isinstance(r[7], str) and ('cumpl' in r[7].lower() or 'no cumple' in r[7].lower()):
                    credits = num(r[6])
                    status_str = r[7]
                    monitor = r[9] if len(r) > 9 else ''
                elif len(r) > 8 and isinstance(r[8], str) and ('cumpl' in r[8].lower() or 'no cumple' in r[8].lower()):
                    credits = num(r[7])
                    status_str = r[8]
                    monitor = r[10] if len(r) > 10 else ''
                else:
                    credits = num(r[6])
                    status_str = r[7] if len(r) > 7 else ''
                    monitor = r[9] if len(r) > 9 else ''
                return credits, re_val, status_str, monitor
    return 0, 50000, '', ''

def read_top20(wb, mes_upper, model_name):
    """Get Top 20 models for a given month."""
    mc = MESES_CAP[mes_upper]
    rr = rows(wb, f'Cómo vamos {mc}')
    models = []
    in_data = False
    name_col = 0
    total_col = None  # column with pre-calculated TOTAL
    for r in rr:
        if not r: continue
        # Header detection: look for 'Nombre modelo' anywhere in row
        if not in_data and any(v == 'Nombre modelo' for v in r if v):
            in_data = True
            for ci, v in enumerate(r):
                if v == 'Nombre modelo':
                    name_col = ci
                if v == 'TOTAL':
                    total_col = ci
            continue
        if in_data:
            name = r[name_col] if name_col < len(r) else None
            if not name or not isinstance(name, str) or len(name) <= 1:
                continue
            if name.upper() in ('TOTAL', 'SUBTOTAL'):
                continue  # skip aggregate rows
            # Use pre-calculated total column when available, else sum platforms
            if total_col and total_col < len(r) and r[total_col] is not None:
                total = num(r[total_col])
            else:
                total = sum(num(v) for v in r[name_col+1:] if isinstance(v, (int, float)) and v is not None)
            if total > 0:
                models.append({'nombre': name, 'total': total})
    models.sort(key=lambda x: x['total'], reverse=True)
    top = models[:20]
    for i, m in enumerate(top):
        m['rank'] = i + 1
        m['es_modelo'] = (m['nombre'] == model_name)
    return top

PLAT_KEYS_ORDER = ['f4f', 'sc', 'cb', 'cs', 'str']
PLAT_DISPLAY    = {'f4f':'Flirt4Free','sc':'Stripchat','cb':'Chaturbate','cs':'CamSoda','str':'Streamate'}

def read_daily(wb, mes_upper, model_name):
    """Daily production per platform.
    Returns {day: {plat_key: val_credits, 'total': X}}.
    Streamate values in Fornax2 are in USD → converted × STR_FACTOR.
    """
    np = detect_nplats(wb, mes_upper)
    rr = rows(wb, mes_upper)
    r = find_model_row(rr, model_name)
    if r is None: return {}

    name_col = 16 if np == 5 else 13
    cols_per_day = np + 1          # +1 for Horas column
    plat_keys = PLAT_KEYS_ORDER[:np]

    daily = {}
    for day in range(31, 0, -1):
        col_start = name_col + 1 + (31 - day) * cols_per_day
        if col_start >= len(r): continue
        plat_vals = {}
        for p_idx, p_key in enumerate(plat_keys):
            col = col_start + p_idx
            v = num(r[col]) if col < len(r) else 0.0
            if p_key == 'str' and v > 0:
                v = v * STR_FACTOR  # USD → créditos
            plat_vals[p_key] = v
        total = sum(plat_vals.values())
        if total > 0:
            daily[day] = {**plat_vals, 'total': total}
    return daily

def read_studio(wb, mes_upper, model_name):
    """Detect studio from monthly sheet."""
    # Grupo Empresarial Excel has col1=studio; Fornax2 doesn't
    # For Fornax2 models, studio is always Fornax Studios
    return 'Fornax Studios'

def detect_platform(monthly_all):
    """Detect primary platform(s) from monthly data."""
    plat_totals = {'f4f':0,'sc':0,'cb':0,'cs':0,'str':0}
    for _, mdata in monthly_all.items():
        for k in plat_totals:
            plat_totals[k] += mdata.get('q1_plats',{}).get(k,0) + mdata.get('q2_plats',{}).get(k,0)
    active = [k for k,v in plat_totals.items() if v > 0]
    map_names = {'f4f':'Flirt4Free','sc':'Stripchat','cb':'Chaturbate','cs':'CamSoda','str':'Streamate'}
    if not active: return 'Flirt4Free'
    return ', '.join(map_names[k] for k in active)

# ── Top 20 Grupo Empresarial ─────────────────────────────────────────────────

def read_top20_grupo(mes_cap, model_name):
    """Lee el TOP20 oficial del dashboard Grupo (calculado por ACTUALIZAR.py).
    Devuelve lista [{rank, nombre, studio, es_modelo, delta}].
    """
    path = os.path.normpath(GRUPO_HTML)
    if not os.path.exists(path):
        return []
    with open(path, 'r', encoding='utf-8') as f:
        html = f.read()
    m = re.search(r"var TOP20\s*=\s*_b64dec\('([A-Za-z0-9+/=]+)'\)", html)
    if not m:
        return []
    try:
        data = json.loads(base64.b64decode(m.group(1)).decode('utf-8'))
    except Exception:
        return []
    entries = data.get(mes_cap, [])
    result = []
    for i, e in enumerate(entries):
        result.append({
            'rank': i + 1,
            'nombre': e['modelo'],
            'studio': e.get('studio', ''),
            'es_modelo': (e['modelo'] == model_name),
            'delta': 0,
        })
    return result

def load_ge_wb():
    """Carga Cómo vamos Grupo Empresarial.xlsx (read-only). Retorna None si no existe."""
    import openpyxl
    path = os.path.normpath(GE_XLSX)
    if not os.path.exists(path):
        return None
    return openpyxl.load_workbook(path, read_only=True, data_only=True)

def compute_prev_ranking(ge_wb, mes_upper, cutoff_day, top20):
    """Calcula el delta de posición (hoy vs ayer) para cada modelo en top20.
    Usa GE.xlsx: resta la producción del cutoff_day para obtener el ranking anterior.
    STR en GE.xlsx está en USD → × STR_FACTOR para créditos.
    Rellena top20[i]['delta'] con: int (>0 subió, <0 bajó, 0 igual) o 'new'.
    """
    if ge_wb is None or cutoff_day <= 1:
        return top20

    sheet_name = mes_upper  # 'AGOSTO', 'JULIO', etc.
    if sheet_name not in ge_wb.sheetnames:
        return top20

    GE_NAME_COL  = 17
    GE_COLS_DAY  = 5   # F4F, SC, CB, CAM, STR por día
    STR_TOT_COL  = 16  # columna total STR en filas de modelo (0-indexed)
    CURR_COLS    = list(range(12, 17))  # F4F, SC, CB, CAM, STR totales

    model_curr = {}  # {nombre: total_creditos}
    model_day  = {}  # {nombre: producción_del_cutoff_day}

    for row in ge_wb[sheet_name].iter_rows(values_only=True):
        if not row or not row[0] or not isinstance(row[0], str): continue
        r = list(row)
        name = r[0]
        if len(r) < 17: continue

        # Total actual (cols 12-15 en créditos + col 16 STR en USD → × STR_FACTOR)
        curr_credits_no_str = sum(float(r[i] or 0) for i in range(12, 16))
        curr_str_credits     = float(r[16] or 0) * STR_FACTOR if len(r) > 16 else 0.0
        model_curr[name] = curr_credits_no_str + curr_str_credits

        # Producción del cutoff_day
        col_s = GE_NAME_COL + 1 + (31 - cutoff_day) * GE_COLS_DAY
        if col_s + 4 < len(r):
            day_no_str  = sum(float(r[col_s + p] or 0) for p in range(4))
            day_str_cr  = float(r[col_s + 4] or 0) * STR_FACTOR
            model_day[name] = day_no_str + day_str_cr
        else:
            model_day[name] = 0.0

    if not model_curr:
        return top20

    # Ranking actual
    curr_sorted = sorted([(n, v) for n, v in model_curr.items() if v > 0], key=lambda x: -x[1])
    curr_rank   = {n: i + 1 for i, (n, _) in enumerate(curr_sorted)}

    # Ranking ayer (total - producción del cutoff_day)
    prev_totals = {n: max(0, model_curr[n] - model_day.get(n, 0)) for n in model_curr}
    prev_sorted = sorted([(n, v) for n, v in prev_totals.items() if v > 0], key=lambda x: -x[1])
    prev_rank   = {n: i + 1 for i, (n, _) in enumerate(prev_sorted)}

    result = []
    for entry in top20:
        name   = entry['nombre']
        c_rank = curr_rank.get(name, entry['rank'])
        p_rank = prev_rank.get(name)
        if p_rank is None:
            delta = 'new'
        else:
            delta = p_rank - c_rank  # positivo = subió posiciones
        result.append({**entry, 'delta': delta})
    return result


# ── Bono de Racha ─────────────────────────────────────────────────────────────

def get_cycle(mes_upper):
    """Return the bono cycle label and months for a given month."""
    idx = MESES.index(mes_upper) if mes_upper in MESES else 0
    if idx < 3:   return 'Enero–Marzo',   ['ENERO','FEBRERO','MARZO']
    if idx < 6:   return 'Abril–Junio',   ['ABRIL','MAYO','JUNIO']
    if idx < 9:   return 'Julio–Septiembre', ['JULIO','AGOSTO','SEPTIEMBRE']
    return 'Octubre–Diciembre', ['OCTUBRE','NOVIEMBRE','DICIEMBRE']

def compute_bono(quincenas_cycle):
    """
    quincenas_cycle: list of {'label', 'credits', 're', 'status'} 
    status: 'achieved' | 'missed' | 'in_progress' | 'pending'
    Returns enriched list with 'consecutivo' and 'bono_earned'
    """
    consecutivo = 0
    result = []
    for q in quincenas_cycle:
        s = q['status']
        if s == 'achieved':
            consecutivo += 1
            bono = BONO_USD.get(consecutivo, 25)
            result.append({**q, 'consecutivo': consecutivo, 'bono_earned': bono})
        elif s == 'missed':
            consecutivo = 0
            result.append({**q, 'consecutivo': 0, 'bono_earned': 0})
        else:  # in_progress o pending — incrementar consecutivo para proyección optimista
            consecutivo += 1
            pot = BONO_USD.get(consecutivo, 25)
            result.append({**q, 'consecutivo': consecutivo, 'bono_earned': pot, 'potential': True})
    return result


# ── CSS ───────────────────────────────────────────────────────────────────────

CSS = """
:root{--bg:#0B1120;--card:#141E35;--card2:#1A2747;--border:rgba(255,255,255,0.07);
  --gold:#F5B800;--gold-light:#FDD85D;--blue:#3B82F6;--green:#22C55E;--red:#EF4444;
  --orange:#F59E0B;--text:#F1F5F9;--muted:#94A3B8;--nav-h:60px}
*{box-sizing:border-box;margin:0;padding:0}
html{scroll-behavior:smooth}
body{font-family:'Segoe UI',system-ui,-apple-system,sans-serif;background:var(--bg);color:var(--text);line-height:1.6;font-size:15px}
nav{position:sticky;top:0;z-index:100;background:rgba(11,17,32,0.95);backdrop-filter:blur(12px);
  border-bottom:1px solid var(--border);height:var(--nav-h);display:flex;align-items:center;
  justify-content:space-between;padding:0 24px}
.nav-brand{font-size:13px;color:var(--muted);font-weight:500;letter-spacing:.05em;display:flex;align-items:center;gap:6px}
.nav-links{display:flex;gap:6px}
.nav-links a{color:var(--muted);text-decoration:none;font-size:12px;padding:6px 12px;
  border-radius:20px;transition:all .2s;white-space:nowrap}
.nav-links a:hover{background:var(--card2);color:var(--text)}
@media(max-width:680px){.nav-links{display:none}.mobile-nav{display:block}}
.mobile-nav{display:none;position:sticky;top:var(--nav-h);z-index:99;
  background:rgba(11,17,32,.97);border-bottom:1px solid var(--border);
  overflow-x:auto;white-space:nowrap;padding:8px 16px;-webkit-overflow-scrolling:touch;scrollbar-width:none}
.mobile-nav::-webkit-scrollbar{display:none}
.mobile-nav a{display:inline-block;color:var(--muted);text-decoration:none;font-size:12px;
  font-weight:600;padding:6px 14px;border-radius:20px;border:1px solid var(--border);margin-right:6px}
@media(max-width:680px){.mobile-nav{display:block}}
main{max-width:1100px;margin:0 auto;padding:0 20px 80px}
section{padding:56px 0 0}
.hero{padding:48px 0 0;display:grid;grid-template-columns:1fr auto;gap:24px;align-items:start}
@media(max-width:680px){.hero{grid-template-columns:1fr}}
.hero-badge{display:inline-flex;align-items:center;gap:8px;
  background:linear-gradient(135deg,rgba(245,184,0,.15),rgba(245,184,0,.05));
  border:1px solid rgba(245,184,0,.3);color:var(--gold);font-size:12px;font-weight:700;
  padding:5px 14px;border-radius:20px;letter-spacing:.08em;text-transform:uppercase;margin-bottom:16px}
.hero-badge::before{content:'★';font-size:10px}
.hero h1{font-size:clamp(36px,6vw,56px);font-weight:800;line-height:1.1;
  background:linear-gradient(135deg,#fff 30%,var(--gold-light));
  -webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;margin-bottom:8px}
.hero-sub{color:var(--muted);font-size:14px;margin-bottom:20px}
.hero-intro{font-size:16px;color:#CBD5E1;max-width:560px;line-height:1.7;margin-bottom:24px}
.hero-tags{display:flex;flex-wrap:wrap;gap:8px}
.tag{font-size:12px;font-weight:600;padding:4px 12px;border-radius:12px;border:1px solid var(--border)}
.tag-platform{background:rgba(59,130,246,.12);border-color:rgba(59,130,246,.3);color:#93C5FD}
.tag-modality{background:rgba(34,197,94,.1);border-color:rgba(34,197,94,.25);color:#86EFAC}
.tag-status{background:rgba(34,197,94,.1);border-color:rgba(34,197,94,.25);color:#86EFAC}
.hero-aside{background:var(--card);border:1px solid var(--border);border-radius:16px;padding:20px 24px;min-width:220px}
.hero-aside-title{font-size:11px;font-weight:700;color:var(--muted);letter-spacing:.1em;text-transform:uppercase;margin-bottom:14px}
.aside-row{display:flex;justify-content:space-between;align-items:center;padding:8px 0;border-bottom:1px solid var(--border)}
.aside-row:last-child{border-bottom:none}
.aside-label{font-size:12px;color:var(--muted)}
.aside-val{font-size:13px;font-weight:600;color:var(--text);text-align:right}
.section-header{margin-bottom:24px}
.section-header h2{font-size:22px;font-weight:700;color:var(--text);display:flex;align-items:center;gap:10px;margin-bottom:4px}
.section-header p{color:var(--muted);font-size:14px;max-width:600px}
.section-divider{width:40px;height:3px;border-radius:2px;background:linear-gradient(90deg,var(--gold),transparent);margin:10px 0 4px}
.kpi-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:14px}
.kpi-card{background:var(--card);border:1px solid var(--border);border-radius:16px;padding:20px;position:relative;overflow:hidden;transition:transform .2s,border-color .2s}
.kpi-card:hover{transform:translateY(-2px);border-color:rgba(255,255,255,.15)}
.kpi-card::after{content:'';position:absolute;inset:0;border-radius:16px;background:linear-gradient(135deg,var(--accent-color,var(--gold)),transparent 60%);opacity:.04;pointer-events:none}
.kpi-label{font-size:11px;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.08em;margin-bottom:10px}
.kpi-value{font-size:28px;font-weight:800;color:var(--text);line-height:1;margin-bottom:4px}
.kpi-sub{font-size:12px;color:var(--muted)}
.kpi-badge{display:inline-block;font-size:11px;font-weight:700;padding:2px 8px;border-radius:8px;margin-top:8px}
.badge-up{background:rgba(34,197,94,.15);color:var(--green)}
.badge-warn{background:rgba(245,158,11,.15);color:var(--orange)}
.badge-down{background:rgba(239,68,68,.12);color:var(--red)}
.chart-card{background:var(--card);border:1px solid var(--border);border-radius:16px;padding:24px}
.chart-card-title{font-size:14px;font-weight:700;color:var(--text);margin-bottom:4px}
.chart-card-sub{font-size:12px;color:var(--muted);margin-bottom:20px}
.chart-toggle{display:flex;gap:8px;margin-bottom:16px;flex-wrap:wrap}
.evol-radio{display:none;position:absolute}
.toggle-btn{font-size:12px;font-weight:600;padding:5px 14px;border-radius:20px;border:1px solid var(--border);background:transparent;color:var(--muted);cursor:pointer;transition:all .2s;touch-action:manipulation;-webkit-tap-highlight-color:transparent;user-select:none;-webkit-user-select:none}
#rb-monthly:checked~.chart-toggle label[for="rb-monthly"]{background:var(--gold);color:#000;border-color:var(--gold)}
#rb-quincena:checked~.chart-toggle label[for="rb-quincena"]{background:var(--gold);color:#000;border-color:var(--gold)}
#view-quincena{display:none}
#rb-quincena:checked~#view-quincena{display:block}
#rb-quincena:checked~#view-monthly{display:none}
.re-timeline{display:flex;flex-wrap:wrap;gap:8px;margin-top:4px}
.re-block{flex:1 1 90px;min-width:80px;max-width:140px;border-radius:10px;padding:10px 12px;text-align:center}
.re-block.achieved{background:rgba(34,197,94,.12);border:1px solid rgba(34,197,94,.3)}
.re-block.missed{background:rgba(239,68,68,.1);border:1px solid rgba(239,68,68,.25)}
.re-block.in-progress{background:rgba(245,184,0,.1);border:1px dashed rgba(245,184,0,.4)}
.re-block.pending{background:rgba(148,163,184,.05);border:1px dashed rgba(148,163,184,.2)}
.re-block-period{font-size:10px;font-weight:700;color:var(--muted);margin-bottom:4px}
.re-block-credits{font-size:15px;font-weight:800}
.re-block.achieved .re-block-credits{color:var(--green)}
.re-block.missed .re-block-credits{color:var(--red)}
.re-block.in-progress .re-block-credits{color:var(--gold)}
.re-block.pending .re-block-credits{color:var(--muted)}
.re-block-pct{font-size:11px;font-weight:600;margin-top:2px}
.re-block.achieved .re-block-pct{color:rgba(34,197,94,.7)}
.re-block.missed .re-block-pct{color:rgba(239,68,68,.7)}
.re-block.in-progress .re-block-pct{color:rgba(245,184,0,.7)}
.re-block-icon{font-size:16px;display:block;margin-bottom:2px}
.re-stats{display:flex;gap:20px;margin-bottom:20px;flex-wrap:wrap}
.re-stat{text-align:center}
.re-stat-num{font-size:28px;font-weight:800}
.re-stat-label{font-size:11px;color:var(--muted);font-weight:600}
.highlight-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(200px,1fr));gap:12px}
.highlight-card{background:var(--card);border:1px solid var(--border);border-radius:14px;padding:18px;transition:border-color .2s,transform .2s}
.highlight-card:hover{transform:translateY(-2px);border-color:rgba(245,184,0,.3)}
.highlight-emoji{font-size:24px;margin-bottom:10px;display:block}
.highlight-title{font-size:12px;color:var(--muted);font-weight:700;text-transform:uppercase;letter-spacing:.06em;margin-bottom:4px}
.highlight-value{font-size:20px;font-weight:800;color:var(--gold);margin-bottom:4px}
.highlight-desc{font-size:12px;color:var(--muted);line-height:1.5}
.list-card{background:var(--card);border:1px solid var(--border);border-radius:16px;padding:24px}
.list-item{display:flex;gap:14px;padding:14px 0;border-bottom:1px solid var(--border)}
.list-item:last-child{border-bottom:none}
.list-num{width:28px;height:28px;border-radius:50%;flex-shrink:0;display:flex;align-items:center;justify-content:center;font-size:13px;font-weight:800}
.num-green{background:rgba(34,197,94,.15);color:var(--green)}
.num-gold{background:rgba(245,184,0,.15);color:var(--gold)}
.num-blue{background:rgba(59,130,246,.15);color:var(--blue)}
.list-content-title{font-size:14px;font-weight:700;color:var(--text);margin-bottom:4px}
.list-content-desc{font-size:13px;color:var(--muted);line-height:1.6}
.goal-card{background:linear-gradient(135deg,rgba(245,184,0,.1),rgba(245,184,0,.03));border:1px solid rgba(245,184,0,.25);border-radius:20px;padding:32px;display:grid;grid-template-columns:1fr auto;gap:24px;align-items:center}
@media(max-width:680px){.goal-card{grid-template-columns:1fr}}
.goal-label{font-size:11px;font-weight:700;color:var(--gold);letter-spacing:.1em;text-transform:uppercase;margin-bottom:8px}
.goal-title{font-size:26px;font-weight:800;color:var(--text);margin-bottom:12px}
.goal-desc{font-size:14px;color:var(--muted);line-height:1.7}
.goal-number{text-align:center;background:rgba(245,184,0,.1);border:2px solid rgba(245,184,0,.3);border-radius:16px;padding:20px 28px}
.goal-number-val{font-size:36px;font-weight:900;color:var(--gold);display:block}
.goal-number-label{font-size:11px;color:var(--muted);font-weight:600}
.action-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(260px,1fr));gap:14px}
.action-card{background:var(--card);border:1px solid var(--border);border-radius:16px;padding:20px;transition:transform .2s}
.action-card:hover{transform:translateY(-2px)}
.action-num{font-size:11px;font-weight:800;letter-spacing:.1em;color:var(--gold);margin-bottom:10px}
.action-title{font-size:15px;font-weight:700;color:var(--text);margin-bottom:8px}
.action-desc{font-size:13px;color:var(--muted);line-height:1.6}
.closing{background:linear-gradient(135deg,var(--card2),var(--card));border:1px solid var(--border);border-radius:20px;padding:40px;text-align:center;margin-top:56px}
.closing-icon{font-size:36px;margin-bottom:16px;display:block}
.closing h2{font-size:24px;font-weight:800;color:var(--text);margin-bottom:14px}
.closing p{font-size:16px;color:var(--muted);max-width:560px;margin:0 auto;line-height:1.7}
.tri-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(200px,1fr));gap:14px}
.tri-card{background:var(--card);border:1px solid var(--border);border-radius:16px;padding:20px}
.tri-label{font-size:11px;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.08em;margin-bottom:12px}
.tri-credits{font-size:24px;font-weight:800;color:var(--text);margin-bottom:4px}
.tri-usd{font-size:14px;color:var(--muted);margin-bottom:8px}
.tri-re{font-size:12px}
.process-timeline{position:relative;padding-left:24px}
.process-timeline::before{content:'';position:absolute;left:0;top:8px;bottom:8px;width:2px;background:linear-gradient(to bottom,var(--gold),rgba(245,184,0,.1));border-radius:1px}
.timeline-item{position:relative;margin-bottom:28px}
.timeline-dot{position:absolute;left:-29px;top:4px;width:12px;height:12px;border-radius:50%;background:var(--gold);border:2px solid var(--bg);box-shadow:0 0 8px rgba(245,184,0,.4)}
.timeline-period{font-size:11px;font-weight:700;color:var(--gold);margin-bottom:4px;text-transform:uppercase;letter-spacing:.08em}
.timeline-text{font-size:13px;color:var(--muted);line-height:1.6}
.timeline-type{font-size:10px;font-weight:700;padding:2px 8px;border-radius:8px;display:inline-block;margin-bottom:6px}
.type-dato{background:rgba(59,130,246,.15);color:#93C5FD}
.type-logro{background:rgba(34,197,94,.12);color:#86EFAC}
.two-col{display:grid;grid-template-columns:1fr 1fr;gap:14px}
@media(max-width:680px){.two-col{grid-template-columns:1fr}}
footer{text-align:center;padding:32px 20px;font-size:11px;color:var(--muted);border-top:1px solid var(--border);margin-top:40px}
nav .nav-brand{display:flex;align-items:center;gap:6px}
.timestamp-badge{background:rgba(255,255,255,.05);border:1px solid rgba(255,255,255,.1);border-radius:8px;padding:4px 10px}
"""


# ── HTML generators ──────────────────────────────────────────────────────────

def h_bar(val, max_val, color='#F5B800', height=8):
    w = min(100, round((val / max_val * 100) if max_val else 0))
    return (f'<div style="height:{height}px;background:rgba(255,255,255,.08);border-radius:{height//2}px;overflow:hidden">'
            f'<div style="height:{height}px;width:{w}%;background:{color};border-radius:{height//2}px"></div></div>')

def h_re_block(period_label, credits, re, status):
    icon = {'achieved':'✅','missed':'❌','in-progress':'⏳','pending':'—'}[status]
    p = pct(credits, re)
    cr_str = fmt(credits) if credits else ('—' if status == 'pending' else '0')
    pct_str = f'{p:.1f}%' if credits else ''
    return (f'<div class="re-block {status}">'
            f'<span class="re-block-icon">{icon}</span>'
            f'<div class="re-block-period">{period_label}</div>'
            f'<div class="re-block-credits">{cr_str}</div>'
            f'<div class="re-block-pct">{pct_str}</div>'
            f'</div>')

def h_monthly_bar_chart(monthly_all, current_mes):
    """HTML table-based bar chart for monthly production."""
    MESES_SHOW = [m for m in MESES if m in monthly_all]
    if not MESES_SHOW: return ''
    max_val = max((monthly_all[m]['total'] for m in MESES_SHOW), default=1) or 1
    mes_label = {'ENERO':'Ene','FEBRERO':'Feb','MARZO':'Mar','ABRIL':'Abr','MAYO':'May',
                 'JUNIO':'Jun','JULIO':'Jul','AGOSTO':'Ago','SEPTIEMBRE':'Sep','OCTUBRE':'Oct',
                 'NOVIEMBRE':'Nov','DICIEMBRE':'Dic'}
    rows_html = ''
    for m in MESES_SHOW:
        d = monthly_all[m]
        tot = d['total']
        is_cur = (m == current_mes)
        bar_color = '#3B82F6' if is_cur else '#F5B800'
        w = round((tot / max_val) * 100) if max_val else 0
        label = mes_label.get(m, m[:3])
        cur_tag = ' <span style="background:#3B82F6;color:#fff;font-size:9px;padding:1px 5px;border-radius:6px">En curso</span>' if is_cur else ''
        rows_html += (
            f'<tr style="background:{"rgba(59,130,246,.04)" if is_cur else "transparent"}">'
            f'<td style="padding:6px 10px;font-size:12px;color:#CBD5E1;font-weight:600;width:40px">{label}</td>'
            f'<td style="padding:6px 8px;min-width:200px"><div style="height:10px;background:rgba(255,255,255,.06);border-radius:5px;overflow:hidden">'
            f'<div style="height:10px;width:{w}%;background:{bar_color};border-radius:5px"></div></div></td>'
            f'<td style="padding:6px 10px;font-size:12px;font-weight:700;color:{bar_color};text-align:right;white-space:nowrap">{fmt(tot)} cr{cur_tag}</td>'
            f'</tr>'
        )
    return f'<div style="overflow-x:auto"><table style="width:100%;border-collapse:collapse">{rows_html}</table></div>'

def h_quincena_bar_chart(monthly_all, current_mes, re):
    """Table-based bar chart by quincena."""
    max_val = max((max(d['q1'],d['q2']) for d in monthly_all.values() if d['q1'] or d['q2']), default=1) or 1
    mes_label = {'ENERO':'Ene','FEBRERO':'Feb','MARZO':'Mar','ABRIL':'Abr','MAYO':'May',
                 'JUNIO':'Jun','JULIO':'Jul','AGOSTO':'Ago'}
    rows_html = ''
    for m in [x for x in MESES if x in monthly_all]:
        d = monthly_all[m]
        mc = MESES_CAP[m]
        for q, val in [('Q1', d['q1']), ('Q2', d['q2'])]:
            if val == 0 and q == 'Q2' and m == current_mes: continue
            ok = val >= re
            is_ip = (val > 0 and val < re and q == 'Q1' and m == current_mes)
            color = '#22C55E' if ok else ('#3B82F6' if is_ip else '#EF4444')
            if val == 0 and m != current_mes: color = '#475569'
            w = round((val / max_val)*100) if max_val else 0
            lbl = f'{mes_label.get(m,m[:3])} {q}'
            rows_html += (
                f'<tr><td style="padding:5px 10px;font-size:11px;color:#CBD5E1;width:60px">{lbl}</td>'
                f'<td style="padding:5px 8px;min-width:180px"><div style="height:8px;background:rgba(255,255,255,.06);border-radius:4px;overflow:hidden">'
                f'<div style="height:8px;width:{w}%;background:{color};border-radius:4px"></div></div></td>'
                f'<td style="padding:5px 10px;font-size:11px;font-weight:700;color:{color};text-align:right">{fmt(val)} cr</td>'
                f'</tr>'
            )
    return f'<div style="overflow-x:auto"><table style="width:100%;border-collapse:collapse">{rows_html}</table></div>'

def _delta_badge(d):
    """HTML badge for position change: int (>0=subió) | 'new' | 0."""
    if d == 'new':
        return '<span style="background:#3B82F6;color:#fff;font-size:9px;padding:1px 6px;border-radius:8px;margin-left:6px;font-weight:700">NUEVO</span>'
    try:
        d = int(d)
    except (TypeError, ValueError):
        return ''
    if d > 0:
        return f'<span style="color:#22C55E;font-size:11px;font-weight:700;margin-left:8px">↑ +{d}</span>'
    if d < 0:
        return f'<span style="color:#EF4444;font-size:11px;font-weight:700;margin-left:8px">↓ {d}</span>'
    return '<span style="color:#64748B;font-size:11px;margin-left:8px">—</span>'

def h_top20_table(top20, model_name):
    """Top 20 Grupo Empresarial — sin cifras de facturación, con indicador ↑↓."""
    if not top20:
        return '<p style="color:var(--muted)">Sin datos disponibles</p>'

    medals = ['🥇','🥈','🥉']
    pod_colors = [
        ('rgba(245,184,0,.25)', 'rgba(245,184,0,.08)', '#F5B800'),
        ('rgba(203,213,225,.2)','rgba(203,213,225,.05)','#94A3B8'),
        ('rgba(205,127,50,.2)', 'rgba(205,127,50,.05)', '#B87333'),
    ]
    podio_html = '<div style="display:flex;gap:14px;margin-bottom:24px;flex-wrap:wrap">'
    for i, m in enumerate(top20[:3]):
        bc, bg, cc = pod_colors[i]
        is_me = m['es_modelo']
        border_c = '#3B82F6' if is_me else cc
        name_extra = ('<span style="background:#3B82F6;color:#fff;font-size:8px;padding:1px 5px;'
                      'border-radius:8px;margin-left:6px;font-weight:700">TÚ</span>') if is_me else ''
        podio_html += (
            f'<div style="flex:1;min-width:140px;background:linear-gradient(135deg,{bc},{bg});'
            f'border:2px solid {border_c};border-radius:16px;padding:20px 14px;text-align:center">'
            f'<div style="font-size:36px;margin-bottom:8px">{medals[i]}</div>'
            f'<div style="font-size:11px;font-weight:800;color:{cc};letter-spacing:.04em;margin-bottom:2px">#{m["rank"]}</div>'
            f'<div style="font-size:14px;font-weight:700;color:#F1F5F9;margin-bottom:6px;line-height:1.3">{m["nombre"]}{name_extra}</div>'
            f'<div style="font-size:11px;color:#64748B;margin-bottom:6px">{m.get("studio","")}</div>'
            f'{_delta_badge(m.get("delta", 0))}'
            f'</div>'
        )
    podio_html += '</div>'

    # Tabla posiciones 4–20
    tabla_html = '<table style="width:100%;border-collapse:collapse;min-width:300px">'
    for m in top20[3:]:
        is_me = m['es_modelo']
        bg     = 'rgba(59,130,246,.1)' if is_me else ('rgba(255,255,255,.04)' if m['rank']%2==0 else 'transparent')
        border = 'border:1px solid #3B82F6' if is_me else 'border:1px solid transparent'
        name_extra = ('<span style="background:#3B82F6;color:#fff;font-size:8px;padding:1px 5px;'
                      'border-radius:8px;font-weight:700;margin-left:6px">TÚ</span>') if is_me else ''
        tabla_html += (
            f'<tr style="background:{bg};{border};border-radius:6px">'
            f'<td style="padding:8px 10px;font-size:13px;font-weight:700;color:#94A3B8;width:36px">#{m["rank"]}</td>'
            f'<td style="padding:8px 6px;font-size:13px;color:#F1F5F9">{m["nombre"]}{name_extra}</td>'
            f'<td style="padding:8px 6px;font-size:11px;color:#64748B">{m.get("studio","")}</td>'
            f'<td style="padding:8px 10px;text-align:right">{_delta_badge(m.get("delta", 0))}</td>'
            f'</tr>'
        )
    tabla_html += '</table>'

    return podio_html + f'<div style="overflow-x:auto">{tabla_html}</div>'

def h_daily_table(daily_plats, current_mes, re, days_in_period, active_plats=None):
    """Tabla diaria por plataforma. daily_plats = {day: {plat_key: val_cr, 'total': X}}.
    active_plats: lista de claves activas ['f4f','sc',...]. Si None, infiere del dict.
    Solo muestra columnas de plataformas con producción real.
    """
    mc_label  = MESES_CAP.get(current_mes, current_mes.capitalize())[:3].lower()
    max_day   = max(days_in_period, 15)

    # Determinar plataformas activas (solo las que tienen datos)
    if active_plats is None:
        active_plats = set()
        for ddata in daily_plats.values():
            for k in PLAT_KEYS_ORDER:
                if ddata.get(k, 0) > 0:
                    active_plats.add(k)
    active = [p for p in PLAT_KEYS_ORDER if p in active_plats]
    if not active: active = [PLAT_KEYS_ORDER[0]]  # al menos F4F

    total_so_far = sum(d.get('total', 0) for d in daily_plats.values())

    # Encabezado: solo columnas activas
    th = 'style="padding:8px 10px;font-size:11px;color:#94A3B8;font-weight:600'
    header = f'<tr style="background:rgba(255,255,255,.06)"><th {th};text-align:left">Fecha</th>'
    for p in active:
        header += f'<th {th};text-align:right">{PLAT_DISPLAY[p]}</th>'
    header += f'<th {th};text-align:right">Total</th></tr>'

    rows_html = ''
    for d in range(1, max_day + 1):
        day_data = daily_plats.get(d)
        bg = 'rgba(255,255,255,.04)' if d % 2 == 0 else 'transparent'
        row = f'<tr style="background:{bg}"><td style="padding:7px 10px;font-size:12px;color:#CBD5E1;font-weight:600">{d} {mc_label}</td>'
        if day_data:
            total = day_data.get('total', 0)
            for p in active:
                v = day_data.get(p, 0)
                col = '#F5B800' if v > 0 else '#475569'
                row += f'<td style="padding:7px 10px;font-size:12px;text-align:right;color:{col}">{fmt(v) if v > 0 else "—"}</td>'
            row += f'<td style="padding:7px 10px;font-size:13px;font-weight:700;text-align:right;color:#F5B800">{fmt(total)}</td>'
        else:
            for _ in active:
                row += '<td style="padding:7px 10px;font-size:12px;text-align:right;color:#475569">—</td>'
            row += '<td style="padding:7px 10px;font-size:12px;text-align:right;color:#475569">—</td>'
        rows_html += row + '</tr>'

    # Fila TOTAL
    rows_html += '<tr style="background:rgba(245,184,0,.08);border-top:2px solid rgba(245,184,0,.3)">'
    rows_html += '<td style="padding:8px 10px;font-size:12px;font-weight:700;color:#F5B800">TOTAL</td>'
    for p in active:
        t = sum(d.get(p, 0) for d in daily_plats.values())
        rows_html += f'<td style="padding:8px 10px;font-size:12px;font-weight:700;text-align:right;color:#F5B800">{fmt(t) if t else "—"}</td>'
    rows_html += f'<td style="padding:8px 10px;font-size:14px;font-weight:900;text-align:right;color:#F5B800">{fmt(total_so_far)}</td></tr>'

    return f'<div style="overflow-x:auto"><table style="width:100%;border-collapse:collapse;min-width:380px">{header}{rows_html}</table></div>'


# ── Main HTML builder ─────────────────────────────────────────────────────────

def _fmt_fecha_es(d):
    """Formatea una fecha como '10 de agosto de 2026' en español."""
    s = d.strftime('%-d de %B de %Y') if hasattr(d, 'strftime') else str(d)
    for en, es in [('January','enero'),('February','febrero'),('March','marzo'),('April','abril'),
                   ('May','mayo'),('June','junio'),('July','julio'),('August','agosto'),
                   ('September','septiembre'),('October','octubre'),('November','noviembre'),('December','diciembre')]:
        s = s.replace(en, es)
    return s

def build_html(profile, monthly_all, quincenas, top20, daily, today, current_mes, prev_mes, data_cutoff=None, gen_dt=None):
    model = profile['nombre']
    nombre_real = profile['nombre_real']
    nivel = profile['nivel']
    modalidad = profile['modalidad']
    monitor = profile['monitor']
    re = profile['re']
    studio = profile['studio']
    plataforma = profile['plataforma']
    slug_upper = model.upper()

    # Plataformas activas del modelo (solo las con producción real)
    plat_totals = {p: 0 for p in PLAT_KEYS_ORDER}
    for _, mdata in monthly_all.items():
        for k in plat_totals:
            plat_totals[k] += mdata.get('q1_plats', {}).get(k, 0) + mdata.get('q2_plats', {}).get(k, 0)
    active_plats = {k for k, v in plat_totals.items() if v > 0}
    if not active_plats: active_plats = {'f4f'}

    # ── Fechas ──────────────────────────────────────────────────────────────
    if data_cutoff is None:
        data_cutoff = today - timedelta(days=1)
    if gen_dt is None:
        gen_dt = datetime.now()
    today_str       = _fmt_fecha_es(today)       # fecha de generación del informe
    data_cutoff_str = _fmt_fecha_es(data_cutoff) # fecha hasta la que llegan los datos (ayer)
    gen_time_str    = gen_dt.strftime('%d/%m/%Y – %H:%M')  # hora real de generación

    closed_meses = [m for m in MESES if m in monthly_all and m != current_mes]
    total_closed = sum(monthly_all[m]['total'] for m in closed_meses)
    cur_q1 = monthly_all.get(current_mes, {}).get('q1', 0)
    cur_q2 = monthly_all.get(current_mes, {}).get('q2', 0)
    cur_total = cur_q1 + cur_q2

    # Best quincena and month among closed
    best_q = 0; best_q_label = ''
    best_m = 0; best_m_label = ''
    for m in closed_meses:
        d = monthly_all[m]
        if d['q1'] > best_q:
            best_q = d['q1']; best_q_label = f'{MESES_CAP[m]} — Quincena 1'
        if d['q2'] > best_q:
            best_q = d['q2']; best_q_label = f'{MESES_CAP[m]} — Quincena 2'
        if d['total'] > best_m:
            best_m = d['total']; best_m_label = MESES_CAP[m]

    # RE stats
    re_met = sum(1 for q in quincenas if q.get('status') == 'achieved')
    re_total_closed = len([q for q in quincenas if q.get('status') in ('achieved','missed')])
    re_pct = pct(re_met, re_total_closed) if re_total_closed else 0

    # Average monthly (closed)
    avg_monthly = round(total_closed / len(closed_meses)) if closed_meses else 0
    avg_monthly_pct = pct(avg_monthly, re * 2)

    # Current period info — usa data_cutoff.day (no today.day)
    mc_label = MESES_CAP.get(current_mes, current_mes.capitalize())
    day_num = data_cutoff.day if hasattr(data_cutoff, 'day') else (today.day - 1)
    is_q1 = day_num <= 15
    q1_complete = is_q1 and day_num == 15  # día 15 = último día de Q1 → período finalizado
    days_elapsed = day_num if is_q1 else day_num - 15
    days_total = 15 if is_q1 else 16
    days_remaining = 0 if q1_complete else max(0, days_total - days_elapsed)
    cur_credits = cur_q1 if is_q1 else cur_q2
    cur_pct = pct(cur_credits, re)
    ritmo = round(cur_credits / days_elapsed) if days_elapsed else 0
    proyeccion = ritmo * days_total
    needed_per_day = round((re - cur_credits) / days_remaining) if days_remaining > 0 and cur_credits < re else 0
    q_label = 'Quincena 1' if is_q1 else 'Quincena 2'
    q_label_short = 'Q1' if is_q1 else 'Q2'
    period_range = f'1–15 {mc_label[:3].lower()}' if is_q1 else f'16–{31 if current_mes not in ("FEBRERO","ABRIL","JUNIO","SEPTIEMBRE","NOVIEMBRE") else ("28" if current_mes=="FEBRERO" else "30")} {mc_label[:3].lower()}'

    # Consecutive streak at end of prev closed quincena
    consecutive_at_close = 0
    for q in reversed(quincenas):
        if q.get('status') == 'achieved': consecutive_at_close += 1
        elif q.get('status') in ('missed',): consecutive_at_close = 0; break
        else: break

    # Bono cycle info
    cycle_label, cycle_months = get_cycle(current_mes)
    cycle_q = []
    for cm in cycle_months:
        for qn in ['q1','q2']:
            lbl = f'{MESES_CAP[cm]} — Quincena {"1" if qn=="q1" else "2"}'
            d = monthly_all.get(cm, {})
            cr = d.get('q1', 0) if qn=='q1' else d.get('q2', 0)
            # Determine status
            if cm not in monthly_all:
                st = 'pending'
            elif cm == current_mes and qn == q_label_short.lower():
                if q1_complete:
                    st = 'achieved' if cr >= re else 'missed'
                else:
                    st = 'in_progress'
            elif cm == current_mes and qn == 'q2' and not is_q1:
                st = 'in_progress'
            elif cm == current_mes and qn == 'q2' and is_q1:
                st = 'pending'
            elif cr == 0 and cm == current_mes:
                st = 'pending'
            elif cr == 0:
                st = 'pending'
            else:
                # Find in quincenas list
                matched = [q for q in quincenas if q.get('mes')==cm and q.get('q')==qn]
                if matched:
                    st = matched[0]['status']
                else:
                    st = 'achieved' if cr >= re else 'missed'
            cycle_q.append({'label': lbl, 'credits': cr, 're': re, 'status': st, 'mes': cm, 'q': qn})

    bono_rows = compute_bono(cycle_q)
    bono_earned = sum(r['bono_earned'] for r in bono_rows if r.get('status')=='achieved' and not r.get('potential'))
    next_bono = next((r['bono_earned'] for r in bono_rows if r.get('status') in ('in_progress','pending') and r.get('potential')), 0)

    # Highlights
    best_racha = max(consecutive_at_close, 1)  # at least current streak

    # Nav links for current month section
    nav_mes_label = mc_label

    # ── Trimestres ────────────────────────────────────────────────────────────
    tri_data = [
        ('Primer Trimestre', ['ENERO','FEBRERO','MARZO']),
        ('Segundo Trimestre', ['ABRIL','MAYO','JUNIO']),
        ('Tercer Trimestre', ['JULIO','AGOSTO','SEPTIEMBRE']),
    ]

    # ── Build HTML ─────────────────────────────────────────────────────────────
    nav_slug = slug_upper.replace(' ', '-')

    html = f'''<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1.0"/>
<title>Dashboard — {model} | {studio}</title>
<style>{CSS}</style>
</head>
<body>
<nav>
  <div class="nav-brand">
    <img src="data:image/png;base64,{LOGO_B64}" alt="{studio}" style="height:48px;vertical-align:middle;margin-right:8px">
    <span style="color:#94A3B8;font-size:10px">&nbsp;·&nbsp; {slug_upper}</span>
  </div>
  <div style="display:flex;align-items:center;gap:14px">
    <div class="timestamp-badge"><div style="font-size:10px;color:#64748B;text-align:right;line-height:1.6">
      <span style="display:block;font-weight:600;color:#94A3B8">Actualizado: {gen_time_str}</span>
      <span style="display:block;color:#64748B">Datos al: {data_cutoff_str}</span></div></div>
    <div class="nav-links" style="display:none" id="nav-links-desktop">
      <a href="#resumen">Resumen</a>
      <a href="#evolucion">Evolución</a>
      <a href="#reto-estrella">Reto Estrella</a>
      <a href="#top20">Top 20</a>
      <a href="#evolucion-diaria">{nav_mes_label}</a>
      <a href="#bono-racha">Bono de Racha</a>
      <a href="#mi-proceso">Mi Proceso</a>
      <a href="#fortalezas">Fortalezas</a>
      <a href="#meta">Mi Meta</a>
    </div>
  </div>
</nav>
<div class="mobile-nav">
  <a href="#resumen">⚡ Resumen</a>
  <a href="#evolucion">📈 Evolución</a>
  <a href="#reto-estrella">⭐ RE</a>
  <a href="#top20">🏆 Top 20</a>
  <a href="#evolucion-diaria">📅 {mc_label[:3]}</a>
  <a href="#bono-racha">💚 Bono</a>
  <a href="#mi-proceso">🗂 Proceso</a>
  <a href="#fortalezas">💪 Fortalezas</a>
  <a href="#meta">🎯 Meta</a>
</div>
<main>

  <!-- HERO -->
  <section class="hero" id="inicio">
    <div>
      <div class="hero-badge">⭐ Nivel {nivel} · {studio}</div>
      <h1>{model}</h1>
      <p class="hero-sub">{nombre_real} &nbsp;·&nbsp; {modalidad} &nbsp;·&nbsp; {plataforma}</p>
      <p class="hero-intro">Este espacio reúne tu evolución, tus logros y los resultados que has construido dentro de {studio}. Informe generado el <strong>{today_str}</strong> · Datos con corte al <strong>{data_cutoff_str}</strong>. {MESES_CAP.get(prev_mes, '')} completamente cerrado · {mc_label} en progreso.</p>
      <div class="hero-tags">
        <span class="tag tag-platform">{plataforma}</span>
        <span class="tag tag-modality">Modalidad {modalidad}</span>
        <span class="tag tag-status">● Activo</span>
      </div>
    </div>
    <div class="hero-aside">
      <div class="hero-aside-title">Tu perfil · Corte {data_cutoff_str}</div>
      <div class="aside-row"><span class="aside-label">Monitor</span><span class="aside-val" style="font-size:11px">{monitor}</span></div>
      <div class="aside-row"><span class="aside-label">Nivel Fornax Nova</span><span class="aside-val">{nivel}</span></div>
      <div class="aside-row"><span class="aside-label">Plataforma</span><span class="aside-val">{plataforma}</span></div>
      <div class="aside-row"><span class="aside-label">RE asignado</span><span class="aside-val">{fmt(re)} cr/quincena</span></div>
      <div class="aside-row"><span class="aside-label">Bono de Racha</span><span class="aside-val" style="color:#22C55E;font-size:11px">{"🏆 USD " + str(bono_earned) + " ganado" if bono_earned else "En construcción"}</span></div>
      <div class="aside-row"><span class="aside-label">{mc_label} ({1}-{day_num} {mc_label[:3].lower()})</span><span class="aside-val" style="font-size:12px;color:#3B82F6">{fmt(cur_credits)} cr ({cur_pct:.1f}%)</span></div>
    </div>
  </section>

  <!-- RESUMEN -->
  <section id="resumen">
    <div class="section-header">
      <h2><span class="icon">⚡</span> Tu Proceso en un Vistazo</h2>
      <div class="section-divider"></div>
      <p>Indicadores clave al {today_str}. {MESES_CAP.get(prev_mes,'')} cerrado · {mc_label} en progreso.</p>
    </div>
    <div class="kpi-grid">
      <div class="kpi-card" style="--accent-color:var(--gold)">
        <div class="kpi-label">Mejor Quincena</div>
        <div class="kpi-value">{fmt(best_q)}</div>
        <div class="kpi-sub">créditos · {best_q_label}</div>
      </div>
      <div class="kpi-card" style="--accent-color:var(--green)">
        <div class="kpi-label">Reto Estrella{(' (' + ', '.join(m[:3].capitalize() for m in closed_meses[:3]) + '–' + MESES_CAP.get(closed_meses[-1],'')[:3] + ')') if closed_meses else ''}</div>
        <div class="kpi-value">{re_met}/{re_total_closed}</div>
        <div class="kpi-sub">quincenas cumplidas</div>
        <span class="kpi-badge {"badge-up" if re_pct >= 80 else "badge-warn"}">{re_pct:.1f}% de cumplimiento</span>
      </div>
      <div class="kpi-card" style="--accent-color:var(--blue)">
        <div class="kpi-label">{mc_label} — {"Q1 Finalizado" if q1_complete else "En Curso"}</div>
        <div class="kpi-value">{fmt(cur_credits)}</div>
        <div class="kpi-sub">créditos · {q_label} ({"período cerrado" if q1_complete else f"{days_elapsed} de {days_total} días"})</div>
        <span class="kpi-badge {"badge-up" if q1_complete and cur_credits >= re else "badge-warn"}">{cur_pct:.1f}% del RE{"  ✅" if q1_complete and cur_credits >= re else "  ❌" if q1_complete else ""}</span>
      </div>
      <div class="kpi-card" style="--accent-color:var(--muted)">
        <div class="kpi-label">Promedio Mensual</div>
        <div class="kpi-value">{fmt(avg_monthly)}</div>
        <div class="kpi-sub">créditos por mes ({len(closed_meses)} meses cerrados)</div>
      </div>
      <div class="kpi-card" style="--accent-color:var(--green)">
        <div class="kpi-label">Bono de Racha — {MESES_CAP.get(prev_mes,'')}</div>
        <div class="kpi-value">{"USD " + str(bono_earned) if bono_earned else "—"}</div>
        <div class="kpi-sub">{"✅ confirmado" if bono_earned else "En construcción"}</div>
      </div>
      <div class="kpi-card" style="--accent-color:var(--orange)">
        <div class="kpi-label">Total {(MESES_CAP.get(closed_meses[0],"")[:3] + "–") if closed_meses else ""}{MESES_CAP.get(prev_mes,"")[:3]} {today.year if hasattr(today,"year") else 2026}</div>
        <div class="kpi-value">{fmt(total_closed)}</div>
        <div class="kpi-sub">créditos en {len(closed_meses)} meses cerrados</div>
      </div>
    </div>
  </section>

  <!-- EVOLUCIÓN -->
  <section id="evolucion">
    <div class="section-header">
      <h2><span class="icon">📈</span> Tu Evolución</h2>
      <div class="section-divider"></div>
    </div>
    <div class="chart-card">
      <input type="radio" id="rb-monthly" name="evol" checked class="evol-radio">
      <input type="radio" id="rb-quincena" name="evol" class="evol-radio">
      <div class="chart-toggle">
        <label for="rb-monthly" class="toggle-btn">Por Mes</label>
        <label for="rb-quincena" class="toggle-btn">Por Quincena</label>
      </div>
      <div id="view-monthly">
        <div class="chart-card-title">Producción Mensual — {(MESES_CAP.get(closed_meses[0],"") + " a ") if closed_meses else ""}{MESES_CAP.get(prev_mes, mc_label)} {today.year if hasattr(today,"year") else 2026} + {mc_label} en curso</div>
        <div class="chart-card-sub">Total de créditos producidos por mes</div>
        {h_monthly_bar_chart(monthly_all, current_mes)}
      </div>
      <div id="view-quincena">
        <div class="chart-card-title">Producción por Quincena</div>
        <div class="chart-card-sub">Créditos por quincena · RE = {fmt(re)} cr</div>
        {h_quincena_bar_chart(monthly_all, current_mes, re)}
      </div>
    </div>'''

    # Tendencia text cards
    last_closed_data = monthly_all.get(prev_mes, {})
    cur_closed_data = monthly_all.get(closed_meses[-2], {}) if len(closed_meses) >= 2 else {}
    delta_month = last_closed_data.get('total',0) - cur_closed_data.get('total',0)
    delta_pct = pct(abs(delta_month), cur_closed_data.get('total',1)) if cur_closed_data.get('total') else 0
    delta_sign = '+' if delta_month >= 0 else '-'

    html += f'''
    <div style="margin-top:20px;display:grid;grid-template-columns:1fr 1fr;gap:14px">
      <div class="chart-card" style="padding:16px">
        <div class="chart-card-title" style="margin-bottom:8px">📊 {MESES_CAP.get(prev_mes,'')} — Cerrado</div>
        <div style="font-size:13px;color:#CBD5E1;line-height:1.7">
          {MESES_CAP.get(prev_mes,'')} cerró con {fmt(last_closed_data.get("total",0))} cr
          (Q1: {fmt(last_closed_data.get("q1",0))} cr {"✅" if last_closed_data.get("q1",0) >= re else "❌"} · 
          Q2: {fmt(last_closed_data.get("q2",0))} cr {"✅" if last_closed_data.get("q2",0) >= re else "❌"}).
          {"Bono de Racha: 🏆 USD " + str(bono_earned) + " ✅" if bono_earned else ""}.
        </div>
      </div>
      <div class="chart-card" style="padding:16px">
        <div class="chart-card-title" style="margin-bottom:8px">📅 {mc_label} — {"Período Finalizado ✅" if q1_complete and cur_credits >= re else "Período Finalizado ❌" if q1_complete else "En Progreso"}</div>
        <div style="font-size:13px;color:#CBD5E1;line-height:1.7">
          {mc_label} {q_label_short} acumula {fmt(cur_credits)} cr en {days_elapsed} días ({cur_pct:.1f}% del RE de {fmt(re)} cr).
          Ritmo diario: {fmt(ritmo)} cr/día. Proyección al día {days_total}: {fmt(proyeccion)} cr.
        </div>
      </div>
    </div>'''

    # Trimestres
    html += '''
    <div class="tri-grid" style="margin-top:20px">'''
    for tri_name, tri_months in tri_data:
        tri_total = sum(monthly_all.get(m,{}).get('total',0) for m in tri_months)
        tri_q1_met = sum(1 for m in tri_months if monthly_all.get(m,{}).get('q1',0) >= re and m in monthly_all)
        tri_q2_met = sum(1 for m in tri_months if monthly_all.get(m,{}).get('q2',0) >= re and m in monthly_all and m != current_mes)
        tri_has_data = any(m in monthly_all for m in tri_months)
        tri_is_current = current_mes in tri_months
        tri_label_sub = 'En progreso' if tri_is_current else ('Cerrado' if tri_has_data else 'Pendiente')
        tri_re_str = f'{tri_q1_met+tri_q2_met}/{sum(2 for m in tri_months if m in closed_meses)} RE cumplidos'
        html += f'''
      <div class="tri-card">
        <div class="tri-label">{tri_name} {"⏳" if tri_is_current else ("✅" if tri_has_data else "—")}</div>
        <div class="tri-credits">{fmt(tri_total) if tri_total else "—"}</div>
        <div class="tri-usd" style="color:#94A3B8;font-size:11px">{tri_label_sub}</div>
        <div class="tri-re" style="font-size:11px;color:#94A3B8">{tri_re_str if tri_has_data else "Sin datos"}</div>
      </div>'''
    html += '</div>'

    # ── Reto Estrella ──────────────────────────────────────────────────────────
    n_quinc = len([q for q in quincenas if q.get('status') in ('achieved','missed')])
    html += f'''

  <!-- RETO ESTRELLA -->
  <section id="reto-estrella">
    <div class="section-header">
      <h2><span class="icon">⭐</span> Tu Historia con el Reto Estrella</h2>
      <div class="section-divider"></div>
      <p>{n_quinc} quincenas evaluadas con RE histórico {fmt(re)} cr. {mc_label} {q_label} en progreso.</p>
    </div>
    <div class="re-stats">
      <div class="re-stat">
        <div class="re-stat-num" style="color:var(--green)">{re_met}</div>
        <div class="re-stat-label">RE CUMPLIDOS</div>
      </div>
      <div class="re-stat">
        <div class="re-stat-num" style="color:var(--red)">{re_total_closed - re_met}</div>
        <div class="re-stat-label">NO CUMPLIDOS</div>
      </div>
      <div class="re-stat">
        <div class="re-stat-num" style="color:var(--gold)">{re_pct:.1f}%</div>
        <div class="re-stat-label">CUMPLIMIENTO</div>
      </div>
      <div class="re-stat">
        <div class="re-stat-num" style="color:var(--blue)">{consecutive_at_close}</div>
        <div class="re-stat-label">RACHA ACTUAL</div>
      </div>
    </div>
    <div class="re-timeline">'''

    for q in quincenas:
        html += h_re_block(q['label'], q['credits'], q.get('re', re), q['status'])

    # Add current in-progress quincena
    html += h_re_block(f'{mc_label} Q{q_label_short[1]}', cur_credits, re, 'in-progress')

    html += f'''
    </div>
  </section>

  <!-- TOP 20 -->
  <section id="top20">
    <div class="section-header">
      <h2><span class="icon">🏆</span> Top 20 — Grupo Empresarial J&amp;D</h2>
      <div class="section-divider"></div>
      <p>Ranking oficial Grupo Empresarial J&amp;D · {mc_label} {today.year if hasattr(today,"year") else 2026} · Datos al {data_cutoff_str} · ↑↓ variación vs día anterior</p>
    </div>
    {h_top20_table(top20, model)}
  </section>

  <!-- EVOLUCIÓN DIARIA -->
  <section id="evolucion-diaria">
    <div class="section-header">
      <h2><span class="icon">📅</span> Mi Evolución Diaria — {mc_label} {today.year if hasattr(today,"year") else 2026}</h2>
      <div class="section-divider"></div>
      <p>Producción por día · Período: 1 al {day_num} de {mc_label.lower()} {today.year if hasattr(today,"year") else 2026} · Reto Estrella {mc_label}: {fmt(re)} cr</p>
    </div>
    <div style="display:flex;gap:12px;margin-bottom:20px;flex-wrap:wrap">
      <div style="background:rgba(245,184,0,.08);border:1px solid rgba(245,184,0,.3);border-radius:10px;padding:14px 18px;min-width:130px;text-align:center">
        <div style="font-size:10px;color:#94A3B8;margin-bottom:4px">Total acumulado</div>
        <div style="font-size:22px;font-weight:900;color:#F5B800">{fmt(cur_credits)}</div>
        <div style="font-size:10px;color:#94A3B8">créditos · {days_elapsed} días</div>
      </div>
      <div style="background:rgba(59,130,246,.08);border:1px solid rgba(59,130,246,.3);border-radius:10px;padding:14px 18px;min-width:130px;text-align:center">
        <div style="font-size:10px;color:#94A3B8;margin-bottom:4px">Promedio diario</div>
        <div style="font-size:22px;font-weight:900;color:#3B82F6">{fmt(ritmo)}</div>
        <div style="font-size:10px;color:#94A3B8">cr/día (ritmo actual)</div>
      </div>
      <div style="background:rgba(34,197,94,.08);border:1px solid rgba(34,197,94,.3);border-radius:10px;padding:14px 18px;min-width:130px;text-align:center">
        <div style="font-size:10px;color:#94A3B8;margin-bottom:4px">Proyección {q_label}</div>
        <div style="font-size:22px;font-weight:900;color:#F59E0B">{fmt(proyeccion)}</div>
        <div style="font-size:10px;color:#94A3B8">cr (a este ritmo)</div>
      </div>
      <div style="background:rgba(239,68,68,.08);border:1px solid rgba(239,68,68,.3);border-radius:10px;padding:14px 18px;min-width:130px;text-align:center">
        <div style="font-size:10px;color:#94A3B8;margin-bottom:4px">RE {mc_label} {q_label_short}</div>
        <div style="font-size:22px;font-weight:900;color:#EF4444">{fmt(re)}</div>
        <div style="font-size:10px;color:#94A3B8">créditos · meta</div>
      </div>
    </div>
    <div style="margin-bottom:16px">
      <div style="font-size:11px;color:#94A3B8;margin-bottom:6px">Progreso del RE — {mc_label} {q_label} ({cur_pct:.1f}%)</div>
      <div style="height:12px;background:rgba(255,255,255,.08);border-radius:6px;overflow:hidden">
        <div style="height:12px;width:{min(100,cur_pct):.0f}%;background:#F5B800;border-radius:6px;transition:width .5s"></div>
      </div>
      <div style="font-size:10px;color:#94A3B8;margin-top:4px">{fmt(cur_credits)} de {fmt(re)} cr · {fmt(max(0,re-cur_credits))} cr restantes para cumplir el RE</div>
    </div>
    <div class="chart-card">
      <div class="chart-card-title">Detalle por día y plataforma</div>
      <div class="chart-card-sub">Créditos producidos cada día · Datos disponibles del 1 al {day_num} de {mc_label.lower()} {today.year if hasattr(today,"year") else 2026}</div>
      <div style="margin-top:12px">
        {h_daily_table(daily, current_mes, re, 15 if is_q1 else 16, active_plats)}
      </div>
    </div>
    <div style="margin-top:10px;font-size:11px;color:#475569;text-align:center">
      Datos actualizados al {today.strftime('%d/%m/%Y') if hasattr(today,'strftime') else ''} – 23:59 · Los días pendientes se completarán con cada actualización del sistema
    </div>
  </section>'''

    # ── Bono de Racha ──────────────────────────────────────────────────────────
    bono_title = f"🏆 GANADOR BONO DE RACHA — {MESES_CAP.get(prev_mes,'').upper()}" if bono_earned else f"💪 BONO DE RACHA EN CONSTRUCCIÓN"
    bono_subtitle = (f"Cumpliste el RE en <strong style='color:#F5B800'>ambas quincenas de {MESES_CAP.get(prev_mes,'')}</strong>. USD {bono_earned} confirmados."
                     if bono_earned else
                     f"Construye quincenas consecutivas con RE cumplido para desbloquear el bono.")
    bono_banner_style = ('background:linear-gradient(135deg,rgba(245,184,0,.2),rgba(245,184,0,.05));border:2px solid #F5B800'
                         if bono_earned else
                         'background:rgba(59,130,246,.08);border:1px solid rgba(59,130,246,.3)')

    html += f'''

  <!-- BONO DE RACHA -->
  <section id="bono-racha">
    <div class="section-header">
      <h2><span class="icon">💚</span> Tu Bono de Racha — Ciclo {cycle_label} {today.year if hasattr(today,"year") else 2026}</h2>
      <div class="section-divider"></div>
      <p>Cada quincena sin cumplir el RE reinicia el consecutivo desde cero.</p>
    </div>
    <div style="margin-bottom:20px;{bono_banner_style};border-radius:14px;padding:18px 24px;text-align:center">
      <div style="font-size:32px;margin-bottom:8px">{"🏆" if bono_earned else "⏳"}</div>
      <div style="font-size:20px;font-weight:900;color:#F5B800;letter-spacing:.04em">{bono_title}</div>
      <div style="font-size:13px;color:#CBD5E1;margin-top:6px">{bono_subtitle}</div>
    </div>
    <div style="display:flex;gap:12px;margin-bottom:20px;flex-wrap:wrap">
      <div style="background:rgba(34,197,94,.1);border:1px solid rgba(34,197,94,.3);border-radius:10px;padding:12px 20px;text-align:center;min-width:120px">
        <div style="font-size:11px;color:#94A3B8;margin-bottom:4px">Consecutivas cerradas</div>
        <div style="font-size:28px;font-weight:900;color:#22C55E;line-height:1">{consecutive_at_close}</div>
        <div style="font-size:10px;color:#22C55E">{consecutive_at_close} quincena(s) consecutiva(s)</div>
      </div>
      <div style="background:rgba(34,197,94,.1);border:1px solid rgba(34,197,94,.3);border-radius:10px;padding:12px 20px;text-align:center;min-width:120px">
        <div style="font-size:11px;color:#94A3B8;margin-bottom:4px">Bono ganado</div>
        <div style="font-size:28px;font-weight:900;color:#22C55E;line-height:1">{"USD " + str(bono_earned) if bono_earned else "USD 0"}</div>
        <div style="font-size:10px;color:#22C55E">{"✅ confirmado" if bono_earned else "En construcción"}</div>
      </div>
      <div style="background:rgba(59,130,246,.08);border:1px solid rgba(59,130,246,.3);border-radius:10px;padding:12px 20px;text-align:center;min-width:120px">
        <div style="font-size:11px;color:#94A3B8;margin-bottom:4px">{mc_label} — en curso</div>
        <div style="font-size:28px;font-weight:900;color:#3B82F6;line-height:1">{fmt(cur_credits)}</div>
        <div style="font-size:10px;color:#3B82F6">cr · {days_elapsed} de {days_total} días</div>
      </div>
      <div style="background:rgba(245,184,0,.08);border:1px solid rgba(245,184,0,.3);border-radius:10px;padding:12px 20px;text-align:center;min-width:120px">
        <div style="font-size:11px;color:#94A3B8;margin-bottom:4px">Potencial ciclo completo</div>
        <div style="font-size:28px;font-weight:900;color:#F5B800;line-height:1">USD {MAX_BONO_CYCLE}</div>
        <div style="font-size:10px;color:#F5B800">máximo del ciclo</div>
      </div>
    </div>
    <div class="chart-card">
      <div class="chart-card-title">Ciclo {cycle_label} {today.year if hasattr(today,"year") else 2026} · Detalle de resultados</div>
      <div class="chart-card-sub">RE: {fmt(re)} cr · Si no cumple el RE en una quincena, el consecutivo se reinicia desde cero</div>
      <div style="margin-top:12px;overflow-x:auto">
        <table style="width:100%;border-collapse:collapse;min-width:480px">
          <tr style="background:rgba(255,255,255,.06)">
            <th style="padding:8px 10px;font-size:11px;color:#94A3B8;font-weight:600;text-align:left">Período</th>
            <th style="padding:8px 10px;font-size:11px;color:#94A3B8;font-weight:600;text-align:left">Resultado</th>
            <th style="padding:8px 10px;font-size:11px;color:#94A3B8;font-weight:600;text-align:left">Racha</th>
            <th style="padding:8px 10px;font-size:11px;color:#94A3B8;font-weight:600;text-align:left">Bono</th>
          </tr>'''

    for br in bono_rows:
        st = br['status']
        bg = 'rgba(34,197,94,.05)' if st=='achieved' else ('rgba(59,130,246,.05)' if st in ('in_progress','pending') else 'transparent')
        cr_str = fmt(br['credits']) if br['credits'] else '—'
        re_pct_str = f"({pct(br['credits'],br['re']):.1f}%)" if br['credits'] else ''
        icon = '✅ Cumplió · ' if st=='achieved' else ('En curso · ' if st=='in_progress' else ('Pendiente' if st=='pending' else '❌ No cumplió · '))
        col = '#22C55E' if st=='achieved' else ('#3B82F6' if st=='in_progress' else ('#475569' if st=='pending' else '#EF4444'))
        bono_val = f"USD {br['bono_earned']} {'✅ GANADO' if st=='achieved' and not br.get('potential') else '(si cumple RE)'}" if br['bono_earned'] else 'USD 0 (base)'
        bono_col = '#22C55E' if (st=='achieved' and not br.get('potential')) else ('#3B82F6' if br.get('potential') else '#94A3B8')
        racha_text = f"Consecutivo {br['consecutivo']} → USD {br['bono_earned']}" if br.get('potential') else (f"Consecutivo {br['consecutivo']}" if st=='achieved' else 'Racha reiniciada' if st=='missed' else '—')

        html += f'''
          <tr style="background:{bg}">
            <td style="padding:7px 10px;font-size:12px;color:#CBD5E1;border-bottom:1px solid rgba(255,255,255,.06)">{br["label"]}</td>
            <td style="padding:7px 10px;font-size:11px;color:{col};border-bottom:1px solid rgba(255,255,255,.06)">{icon}{cr_str} {re_pct_str}</td>
            <td style="padding:7px 10px;font-size:11px;color:#94A3B8;border-bottom:1px solid rgba(255,255,255,.06)">{racha_text}</td>
            <td style="padding:7px 10px;font-size:12px;font-weight:700;color:{bono_col};border-bottom:1px solid rgba(255,255,255,.06)">{bono_val}</td>
          </tr>'''

    html += '''
        </table>
      </div>
    </div>
  </section>'''

    # ── Mejores Momentos ───────────────────────────────────────────────────────
    html += f'''

  <!-- MEJORES MOMENTOS -->
  <section>
    <div class="section-header">
      <h2><span class="icon">🏆</span> Tus Mejores Momentos</h2>
      <div class="section-divider"></div>
      <p>Los hitos más importantes — {MESES_CAP.get(closed_meses[0],"Agosto") if closed_meses else mc_label} a {MESES_CAP.get(prev_mes, mc_label)}{" " + str(today.year if hasattr(today,"year") else 2026)}.</p>
    </div>
    <div class="highlight-grid">
      <div class="highlight-card">
        <span class="highlight-emoji">🚀</span>
        <div class="highlight-title">Récord Personal</div>
        <div class="highlight-value">{fmt(best_q)} cr</div>
        <div class="highlight-desc">{best_q_label} · {pct(best_q,re):.1f}% del RE.</div>
      </div>
      <div class="highlight-card">
        <span class="highlight-emoji">📅</span>
        <div class="highlight-title">Mejor Mes{(' (' + MESES_CAP.get(closed_meses[0],'Ene') + '–' + MESES_CAP.get(prev_mes,'') + ')') if closed_meses else ''}</div>
        <div class="highlight-value">{fmt(best_m)} cr</div>
        <div class="highlight-desc">{best_m_label} {today.year if hasattr(today,"year") else 2026} — Mes de mayor producción.</div>
      </div>
      <div class="highlight-card">
        <span class="highlight-emoji">🔥</span>
        <div class="highlight-title">Racha Activa</div>
        <div class="highlight-value">{consecutive_at_close} quincenas</div>
        <div class="highlight-desc">Quincenas consecutivas con RE cumplido al cierre de {MESES_CAP.get(prev_mes,"julio")}.</div>
      </div>
      <div class="highlight-card">
        <span class="highlight-emoji">💎</span>
        <div class="highlight-title">Total {(MESES_CAP.get(closed_meses[0],'Ene') + '–') if closed_meses else ''}{MESES_CAP.get(prev_mes,"")}</div>
        <div class="highlight-value">{fmt(total_closed)} cr</div>
        <div class="highlight-desc">Producción total de {MESES_CAP.get(closed_meses[0],"agosto").lower() if closed_meses else mc_label.lower()} a {MESES_CAP.get(prev_mes, mc_label).lower()} {today.year if hasattr(today,"year") else 2026}.</div>
      </div>
      <div class="highlight-card">
        <span class="highlight-emoji">⭐</span>
        <div class="highlight-title">Nivel Alcanzado</div>
        <div class="highlight-value">{nivel.upper()}</div>
        <div class="highlight-desc">Nivel {nivel} del programa Fornax Nova.</div>
      </div>
      <div class="highlight-card">
        <span class="highlight-emoji">🎯</span>
        <div class="highlight-title">RE Cumplimiento</div>
        <div class="highlight-value">{re_met}/{re_total_closed}</div>
        <div class="highlight-desc">{re_pct:.1f}% de quincenas con Reto Estrella cumplido.</div>
      </div>
    </div>
  </section>

  <!-- MI PROCESO -->
  <section id="mi-proceso">
    <div class="section-header">
      <h2><span class="icon">🗂</span> Lo que Hemos Venido Trabajando</h2>
      <div class="section-divider"></div>
      <p>Evolución documentada — corte {today_str}.</p>
    </div>
    <div class="list-card">
      <div class="process-timeline">'''

    for tri_name, tri_months in tri_data:
        tri_total = sum(monthly_all.get(m,{}).get('total',0) for m in tri_months)
        tri_has_data = any(m in monthly_all and m != current_mes for m in tri_months)
        tri_is_current = current_mes in tri_months
        if not tri_has_data and not tri_is_current: continue
        tri_q_met = sum(1 for m in tri_months for qk in ('q1','q2')
                        if monthly_all.get(m,{}).get(qk,0) >= re and m in monthly_all and m != current_mes)
        tri_q_total = sum(2 for m in tri_months if m in closed_meses)
        dot_color = '#3B82F6' if tri_is_current else '#F5B800'
        tag_class = 'type-logro' if (not tri_is_current and tri_q_met == tri_q_total) else 'type-dato'
        tag_text = f'{tri_q_met}/{tri_q_total} RE' if not tri_is_current else 'En curso'
        status_str = f'Cerrado con {fmt(tri_total)} cr y {tri_q_met}/{tri_q_total} quincenas cumplidas.' if not tri_is_current else f'Acumula {fmt(tri_total)} cr hasta la fecha.'
        html += f'''
        <div class="timeline-item">
          <div class="timeline-dot" style="background:{dot_color}"></div>
          <div class="timeline-period">{tri_name} {today.year if hasattr(today,"year") else 2026}{" · CERRADO" if not tri_is_current else ""}</div>
          <span class="timeline-type {tag_class}">{tag_text}</span>
          <div class="timeline-text">{status_str}</div>
        </div>'''

    # Add current month item
    html += f'''
        <div class="timeline-item">
          <div class="timeline-dot" style="background:#3B82F6"></div>
          <div class="timeline-period">{mc_label} {today.year if hasattr(today,"year") else 2026} · {"Q1 Finalizado" if q1_complete else f"En Progreso (1–{day_num} de {mc_label.lower()})"}</div>
          <span class="timeline-type" style="background:{"rgba(34,197,94,.15);color:#22C55E" if q1_complete and cur_credits >= re else "rgba(239,68,68,.15);color:#EF4444" if q1_complete else "rgba(59,130,246,.15);color:#3B82F6"}">{"✅ Cumplió RE" if q1_complete and cur_credits >= re else "❌ No cumplió RE" if q1_complete else "En curso"}</span>
          <div class="timeline-text">{mc_label} {q_label_short} {"cerró con" if q1_complete else "acumula"} <strong style="color:{"#22C55E" if q1_complete and cur_credits >= re else "#EF4444" if q1_complete else "#3B82F6"}">{fmt(cur_credits)} créditos</strong> en {days_elapsed} días ({cur_pct:.1f}% del RE de {fmt(re)} cr). {"RE " + ("cumplido ✅" if cur_credits >= re else f"no cumplido ❌ — faltaron {fmt(re-cur_credits)} cr") if q1_complete else "Restan " + fmt(max(0,re-cur_credits)) + " cr para cumplir el RE del período." + (" Bono de Racha: si cumple " + mc_label + " Q1, consecutivo " + str(consecutive_at_close+1) + " → USD " + str(BONO_USD.get(consecutive_at_close+1,25)) + "." if consecutive_at_close > 0 else "")}</div>
        </div>
      </div>
    </div>
  </section>

  <!-- FORTALEZAS -->
  <section id="fortalezas">
    <div class="section-header">
      <h2><span class="icon">💪</span> Lo que Estás Haciendo Bien</h2>
      <div class="section-divider"></div>
      <p>Fortalezas identificadas a partir de los resultados reales.</p>
    </div>
    <div class="two-col">
      <div class="list-card">
        <div class="list-item">
          <div class="list-num num-green">1</div>
          <div><div class="list-content-title">Producción sostenida</div>
          <div class="list-content-desc">Total de {fmt(total_closed)} cr en {len(closed_meses)} meses. Promedio mensual de {fmt(avg_monthly)} cr.</div></div>
        </div>
        <div class="list-item">
          <div class="list-num num-green">2</div>
          <div><div class="list-content-title">Mejor mes: {best_m_label}</div>
          <div class="list-content-desc">Producción de {fmt(best_m)} cr. Capacidad de rendimiento alto sostenido.</div></div>
        </div>
        <div class="list-item">
          <div class="list-num num-green">3</div>
          <div><div class="list-content-title">Récord de quincena: {fmt(best_q)} cr</div>
          <div class="list-content-desc">{best_q_label} — {pct(best_q,re):.1f}% del RE de {fmt(re)} cr.</div></div>
        </div>
      </div>
      <div class="list-card">
        <div class="list-item">
          <div class="list-num num-gold">4</div>
          <div><div class="list-content-title">Cumplimiento del Reto Estrella</div>
          <div class="list-content-desc">{re_met} de {re_total_closed} quincenas evaluadas con RE cumplido ({re_pct:.1f}%).</div></div>
        </div>
        <div class="list-item">
          <div class="list-num num-gold">5</div>
          <div><div class="list-content-title">Racha activa: {consecutive_at_close} quincenas</div>
          <div class="list-content-desc">Quincenas consecutivas con RE cumplido al cierre de {MESES_CAP.get(prev_mes,"julio")}.</div></div>
        </div>
        <div class="list-item">
          <div class="list-num num-gold">6</div>
          <div><div class="list-content-title">Inicio de {mc_label}</div>
          <div class="list-content-desc">{fmt(cur_credits)} cr en los primeros {days_elapsed} días ({cur_pct:.1f}% del RE de {mc_label}).</div></div>
        </div>
      </div>
    </div>
  </section>

  <!-- DÓNDE CRECER -->
  <section>
    <div class="section-header">
      <h2><span class="icon">🌱</span> Dónde Podemos Seguir Creciendo</h2>
      <div class="section-divider"></div>
      <p>Áreas con potencial real de mejora.</p>
    </div>
    <div class="list-card">'''

    missed_quincs = [q for q in quincenas if q.get('status') == 'missed']
    if missed_quincs:
        min_q = min(missed_quincs, key=lambda x: x['credits'])
        html += f'''
      <div class="list-item"><div class="list-num num-blue">1</div>
        <div><div class="list-content-title">Quincenas por debajo del RE</div>
        <div class="list-content-desc"><strong>Observado:</strong> {len(missed_quincs)} quincena(s) no alcanzaron el RE de {fmt(re)} cr. La más baja: {min_q["label"]} con {fmt(min_q["credits"])} cr.<br/>
        <strong>Oportunidad:</strong> Identificar los factores de esos períodos y activar recuperaciones tempranas.</div></div>
      </div>'''

    html += f'''
      <div class="list-item"><div class="list-num num-blue">{len(missed_quincs)+1 if missed_quincs else 1}</div>
        <div><div class="list-content-title">Cumplir el RE de {mc_label} (RE: {fmt(re)} cr)</div>
        <div class="list-content-desc"><strong>Observado:</strong> {mc_label} {q_label_short} lleva {fmt(cur_credits)} cr en {days_elapsed} días — restan {fmt(max(0,re-cur_credits))} cr para cumplir el RE.<br/>
        <strong>Oportunidad:</strong> {"Continuar el ritmo actual de " + fmt(ritmo) + " cr/día." if proyeccion >= re else "Se necesitan " + fmt(needed_per_day) + " cr/día en los " + str(days_remaining) + " días restantes para cumplir el RE."}</div></div>
      </div>
      <div class="list-item"><div class="list-num num-blue">{len(missed_quincs)+2 if missed_quincs else 2}</div>
        <div><div class="list-content-title">Construir y sostener la racha del Bono</div>
        <div class="list-content-desc"><strong>Observado:</strong> Con {consecutive_at_close} quincena(s) consecutiva(s) al cierre de {MESES_CAP.get(prev_mes,"")}, {mc_label} {q_label_short} puede ser el consecutivo {consecutive_at_close+1} (USD {BONO_USD.get(consecutive_at_close+1,25)}).<br/>
        <strong>Oportunidad:</strong> El Bono de Racha acumula hasta USD {MAX_BONO_CYCLE} en el ciclo. Cada quincena consecutiva importa.</div></div>
      </div>
    </div>
  </section>

  <!-- META -->
  <section id="meta">
    <div class="section-header">
      <h2><span class="icon">🎯</span> Tu Próximo Objetivo — {mc_label} {today.year if hasattr(today,"year") else 2026}</h2>
      <div class="section-divider"></div>
    </div>
    <div class="goal-card">
      <div>
        <div class="goal-label">{mc_label} — {q_label} · {"Período Finalizado" if q1_complete else f"En Progreso ({period_range} {today.year if hasattr(today,'year') else 2026})"}</div>
        <div class="goal-title">{"Mantener el ritmo y cumplir el RE" if proyeccion >= re else "Acelerar para cerrar el RE"}</div>
        <div class="goal-desc">{mc_label} {q_label_short} lleva <strong style="color:var(--gold)">{fmt(cur_credits)} créditos</strong> en {days_elapsed} días ({cur_pct:.1f}% del RE). {"Proyección al día " + str(days_total) + ": " + fmt(proyeccion) + " cr — " + ("✅ en camino a cumplir el RE." if proyeccion >= re else "⚠️ por debajo del RE, se necesitan " + fmt(needed_per_day) + " cr/día.") + "<br/><br/>" if days_remaining > 0 else ""}{"Con la racha de " + str(consecutive_at_close) + " quincenas activa desde " + MESES_CAP.get(prev_mes,"") + ", cumplir " + mc_label + " " + q_label_short + " activaría el consecutivo " + str(consecutive_at_close+1) + " del Bono (USD " + str(BONO_USD.get(consecutive_at_close+1,25)) + ")." if consecutive_at_close > 0 else ""}</div>
      </div>
      <div class="goal-number">
        <span class="goal-number-val">{fmt(re)}</span>
        <div class="goal-number-label">RE {mc_label} · créditos<br/>por quincena</div>
      </div>
    </div>
  </section>

  <!-- ACCIONES -->
  <section>
    <div class="section-header">
      <h2><span class="icon">📋</span> ¿En qué nos vamos a Enfocar?</h2>
      <div class="section-divider"></div>
    </div>
    <div class="action-grid">
      <div class="action-card">
        <div class="action-num">01 — PRIORIDAD PRINCIPAL</div>
        <div class="action-title">Cerrar {mc_label} {q_label_short} con {fmt(re)} cr o más</div>
        <div class="action-desc">{"Faltan " + fmt(max(0,re-cur_credits)) + " cr para cumplir el RE de " + mc_label + " " + q_label_short + ". Con " + str(days_remaining) + " días restantes, el ritmo necesario es " + fmt(needed_per_day) + " cr/día." if cur_credits < re else "¡Vas por buen camino! Mantén el ritmo para asegurar el RE."}</div>
      </div>
      <div class="action-card">
        <div class="action-num">02 — BONO DE RACHA</div>
        <div class="action-title">Extender la racha del Bono</div>
        <div class="action-desc">{MESES_CAP.get(prev_mes,"Julio")} dejó {consecutive_at_close} quincena(s) consecutiva(s). Cumplir {mc_label} {q_label_short} lleva el Bono a USD {BONO_USD.get(consecutive_at_close+1,25)}.</div>
      </div>
      <div class="action-card">
        <div class="action-num">03 — SEGUIMIENTO</div>
        <div class="action-title">Monitoreo activo y metas diarias</div>
        <div class="action-desc">El monitor realiza seguimiento continuo para detectar caídas antes del día 10 de cada período. Las metas diarias concretas son clave para no depender de recuperaciones tardías.</div>
      </div>
    </div>
  </section>

  <!-- CLOSING -->
  <div class="closing">
    <span class="closing-icon">🏆</span>
    <h2>{model} — {"Bono de Racha confirmado, " + mc_label + " suma más" if bono_earned else mc_label + " en construcción"}</h2>
    <p>{"Con " + str(consecutive_at_close) + " quincena(s) consecutiva(s) cerradas, " + mc_label + " " + q_label_short + " es la clave para el siguiente nivel del Bono. Quincena a quincena es como se construye el máximo." if consecutive_at_close > 0 else mc_label + " " + q_label_short + " lleva " + fmt(cur_credits) + " cr. Cada día cuenta para cerrar el RE."}</p>
  </div>
</main>

<footer>
  {studio} · Grupo Empresarial J&amp;D · Documento de uso interno — Confidencial<br/>
  Actualizado al {today_str} · Última actualización: {today.strftime('%d/%m/%Y') if hasattr(today,'strftime') else ''} – 23:59 · Fuente: Cómo Vamos Fornax Studios · {model}
</footer>

<script>
// Nav desktop
var nl=document.getElementById('nav-links-desktop');
if(nl && window.innerWidth>=900) nl.style.display='flex';
</script>
</body>
</html>'''

    return html


# ── Main ──────────────────────────────────────────────────────────────────────

def main(model_targets=None):
    today  = date.today()
    gen_dt = datetime.now()  # hora exacta de generación del informe
    print(f"\n{'='*60}")
    print(f"  MODELOS.py — Dashboards individuales")
    print(f"  Fecha: {today.strftime('%d/%m/%Y')} · Fuente: Fornax2.xlsx")
    print(f"{'='*60}\n")

    data_cutoff = date(2026, 8, 15)  # Corte explícito: 15 de agosto
    print(f"  📅 Fecha informe: {today.strftime('%d/%m/%Y')} · Corte datos: {data_cutoff.strftime('%d/%m/%Y')}")

    print("  📂 Cargando Cómo vamos Fornax2.xlsx…")
    wb = load_wb()

    print("  📂 Cargando Cómo vamos Grupo Empresarial.xlsx…")
    ge_wb = load_ge_wb()
    if ge_wb is None:
        print("  ⚠  GE.xlsx no encontrado — se usará Fornax2 para Top 20 y sin delta de posición.")

    # Detect current month: last sheet that has ANY numeric data in col 1
    current_mes = None
    for m in reversed(MESES):
        if m not in wb.sheetnames: continue
        current_mes = m  # fallback: último mes que exista
        rr = rows(wb, m)
        for r in rr:
            if r and len(r) > 2 and isinstance(r[0], str) and r[0] not in ('', 'Nombre modelo') and isinstance(r[1], (int, float)) and (r[1] or 0) > 0:
                break  # mes tiene datos numéricos
        else:
            continue
        break  # este mes tiene datos, usarlo

    prev_idx = MESES.index(current_mes) - 1 if current_mes in MESES else 0
    prev_mes = MESES[prev_idx] if prev_idx >= 0 else MESES[0]

    print(f"  📅 Mes actual detectado: {MESES_CAP.get(current_mes, current_mes)}")
    print(f"  📅 Mes previo: {MESES_CAP.get(prev_mes, prev_mes)}")

    # Get all active models if not specified
    if model_targets is None:
        # Read from Base de datos modelos
        model_targets = []
        for r in rows(wb, 'Base de datos modelos'):
            if r and len(r) >= 3 and r[3] and isinstance(r[3], str) and r[1] == 'Activo':
                model_targets.append(r[3])
        model_targets = model_targets or ['Isa Raven']
        print(f"  📋 Modelos activos encontrados: {len(model_targets)}")

    for model_name in model_targets:
        print(f"\n  🔄 Generando dashboard: {model_name}")

        # Profile
        profile = read_profile(wb, model_name)
        profile['nombre'] = model_name

        # Monthly data for all available months
        monthly_all = {}
        for m in MESES:
            if m not in wb.sheetnames: continue
            mdata = read_monthly(wb, m, model_name)
            if mdata['q1'] > 0 or mdata['q2'] > 0 or mdata['total'] > 0 or m == current_mes:
                monthly_all[m] = mdata

        if not monthly_all:
            print(f"     ⚠  Sin datos para {model_name}, omitiendo.")
            continue

        # Detect primary platform
        profile['plataforma'] = detect_platform(monthly_all)
        if not profile['plataforma']:
            profile['plataforma'] = 'Flirt4Free'

        # Quincena RE status for closed months (not current)
        quincenas = []
        closed_meses = [m for m in MESES if m in monthly_all and m != current_mes]

        for m in closed_meses:
            mc = MESES_CAP[m]
            np = monthly_all[m]['n_plats']
            for qn, q_num in [('q1','1'),('q2','2')]:
                credits, _re_v, status_str, mon = read_quincena_status(wb, m, model_name, qn)
                # RE siempre desde BASE DE DATOS — nunca sobrescribir con hoja de quincena
                if mon and not profile['monitor']: profile['monitor'] = mon
                cred = monthly_all[m][qn]
                if credits == 0 and cred > 0: credits = cred
                # Usar RE histórico si el modelo tenía un RE diferente en ese mes
                re_val = HISTORICAL_RE.get(model_name, {}).get(m, profile['re'])
                # Si el mes está en HISTORICAL_RE para este modelo, recalcular SIEMPRE
                # desde créditos: la hoja de quincena puede haberse generado con RE incorrecto
                if m in HISTORICAL_RE.get(model_name, {}):
                    status = 'achieved' if cred >= re_val else 'missed'
                elif not status_str:
                    status = 'achieved' if cred >= re_val else 'missed'
                else:
                    status = 'achieved' if 'cumpl' in status_str.lower() and 'no' not in status_str.lower() else 'missed'
                if cred == 0 and m not in monthly_all:
                    status = 'pending'
                quincenas.append({
                    'mes': m, 'q': qn,
                    'label': f'{mc} — Quincena {q_num}',
                    'credits': cred, 're': re_val, 'status': status
                })

        # Si el corte es día 15, Q1 del mes actual quedó cerrado → agregarlo a quincenas
        if data_cutoff.day == 15 and current_mes in monthly_all:
            mc_cur = MESES_CAP[current_mes]
            cred_q1 = monthly_all[current_mes].get('q1', 0)
            re_cur = HISTORICAL_RE.get(model_name, {}).get(current_mes, profile['re'])
            st_q1 = 'achieved' if cred_q1 >= re_cur else 'missed'
            quincenas.append({
                'mes': current_mes, 'q': 'q1',
                'label': f'{mc_cur} — Quincena 1',
                'credits': cred_q1, 're': re_cur, 'status': st_q1
            })

        # Top 20 Grupo Empresarial (fuente oficial: grupo579780/index.html)
        mc_label_cap = MESES_CAP.get(current_mes, current_mes.capitalize())
        top20 = read_top20_grupo(mc_label_cap, model_name)
        if not top20:
            # Fallback: Fornax2 si no se puede leer el grupo HTML
            top20 = read_top20(wb, current_mes, model_name)

        # Delta de posición (hoy vs ayer)
        if ge_wb is not None:
            top20 = compute_prev_ranking(ge_wb, current_mes, data_cutoff.day, top20)

        # Datos diarios por plataforma (con conversión Streamate)
        daily = read_daily(wb, current_mes, model_name)

        # Generate HTML
        html = build_html(profile, monthly_all, quincenas, top20, daily, today, current_mes, prev_mes, data_cutoff, gen_dt)

        # Save
        slug = slugify(model_name)
        out_dir = os.path.join(SCRIPT_DIR, 'modelos', slug)
        os.makedirs(out_dir, exist_ok=True)
        out_path = os.path.join(out_dir, 'index.html')
        html = inject_pin(html, slug)  # protección PIN — siempre al final
        with open(out_path, 'w', encoding='utf-8') as f:
            f.write(html)

        size_kb = os.path.getsize(out_path) / 1024
        print(f"     ✅ Guardado: modelos/{slug}/index.html ({size_kb:.1f} KB)")
        print(f"     📊 Meses: {len(monthly_all)} | RE cumplidos: {sum(1 for q in quincenas if q['status']=='achieved')}/{len(quincenas)}")

    print(f"\n{'='*60}")
    print(f"  ✅ MODELOS.py completado.")
    print(f"{'='*60}\n")


if __name__ == '__main__':
    import sys
    if len(sys.argv) > 1:
        targets = sys.argv[1:]
        main(targets)
    else:
        # Default: todos los modelos activos de BASE DE DATOS
        main()
