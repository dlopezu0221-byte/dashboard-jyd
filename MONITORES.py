#!/usr/bin/env python3
"""
MONITORES.py v2 — Dashboard de Monitores Fornax Studios
Fuente Fornax: Cómo vamos Fornax2.xlsx
Fuente Grupo:  Cómo vamos Grupo Empresarial.xlsx (ALIADOS del dashboard grupo)
Salida: dashboard-jyd/monitores/index.html
"""

import json, os, re, base64
from datetime import datetime, timedelta, date
from openpyxl import load_workbook

# ── RUTAS ──────────────────────────────────────────────────────────────────────
BASE  = os.path.dirname(os.path.abspath(__file__))
CVMG  = os.path.join(os.path.dirname(BASE),
    "Centro de Gestión Estratégica Grupo J&D", "COMO VAMOS GRUPO")
EXCEL_FORNAX = os.path.join(CVMG, "Cómo vamos Fornax2.xlsx")
GRUPO_HTML   = os.path.join(BASE, "grupo579780", "index.html")
OUT_DIR  = os.path.join(BASE, "monitores")
OUT_HTML = os.path.join(OUT_DIR, "index.html")
os.makedirs(OUT_DIR, exist_ok=True)

# ── FECHAS ─────────────────────────────────────────────────────────────────────
HOY    = date(2026, 8, 15)
CORTE  = date(2026, 8, 15)                 # Corte explícito: 15 de agosto
GEN_DT = datetime.now()
HOY_STR   = GEN_DT.strftime("%d/%m/%Y — %H:%M")  # fecha + hora real de generación
CORTE_STR = CORTE.strftime("%d/%m/%Y")            # fecha de corte de datos

# ── QUINCENA SHEETS (nombres exactos en Excel) ────────────────────────────────
QUINCENAS = {
    "Enero":   [("Enero - Periodo 1 al 15","q1"),   ("Enero - Periodo 16 al 31","q2")],
    "Febrero": [("Febrero - Periodo 1 al 15","q1"), ("Febrero - Periodo 16 al 28","q2")],
    "Marzo":   [("Marzo - Periodo 1 al 15","q1"),   ("Marzo - Periodo 16 al 31","q2")],
    "Abril":   [("Abril - Periodo 1 al 15","q1"),   ("Abril - Periodo 16 al 31","q2")],
    "Mayo":    [("Mayo - Periodo 1 al 15 ","q1"),   ("Mayo - Periodo 16 al 31","q2")],
    "Junio":   [("Junio - Periodo 1 al 15","q1"),   ("Junio - Periodo 16 al 30","q2")],
    "Julio":   [("Julio - Periodo 1 al 15","q1"),   ("Julio - Periodo 16 al 31","q2")],
    "Agosto":  [("Agosto - Periodo 1 al 15","q1"),  ("Agosto - Periodo 16 al 31","q2")],
}

# ── HELPERS ───────────────────────────────────────────────────────────────────
def v(x):
    if x is None: return 0.0
    if isinstance(x, (int, float)): return float(x)
    try: return float(x)
    except: return 0.0

def str_cell(x):
    return "" if x is None else str(x).strip()

def fmt_date(d):
    if d is None: return ""
    if isinstance(d, (datetime, date)): return d.strftime("%d/%m/%Y")
    return str(d)

# ── BASE DE DATOS MODELOS ─────────────────────────────────────────────────────
def read_bd_modelos(wb):
    ws = wb["Base de datos modelos"]
    out = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[3] or str(row[3]).strip() == "":
            continue
        out.append({
            "modelo":        str(row[3]).strip(),
            "nombre_real":   str_cell(row[4]),
            "estado":        str_cell(row[1]),
            "nivel":         str_cell(row[2]),
            "monitor":       str_cell(row[10]) or "Sin monitor",
            "reto_estrella": v(row[11]),
            "pct_pago":      v(row[8]),
            "modalidad":     str_cell(row[7]),
            "ejecutivo":     str_cell(row[18]),
        })
    return out

# ── LEER QUINCENA ─────────────────────────────────────────────────────────────
def read_quincena(wb, sheet_name, mes_label, qid):
    if sheet_name not in wb.sheetnames:
        return None
    ws   = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    N    = len(rows)

    def cell(r, c):
        if r >= N or c >= len(rows[r]): return None
        return rows[r][c]

    # ── Fechas / días ──────────────────────────────────────────────────────────
    fecha_ini      = cell(13, 2)
    fecha_fin      = cell(13, 3)
    dias_periodo   = v(cell(13, 5))
    dias_facturados= v(cell(14, 3)) or v(cell(14, 2))

    # ── Daily totals ───────────────────────────────────────────────────────────
    daily_hdr = 17
    for ri in range(10, 26):
        r = rows[ri] if ri < N else []
        if any(c is not None and 'Día 1' in str(c) for c in r):
            daily_hdr = ri; break

    daily_labels, daily_totals = [], []
    if daily_hdr + 1 < N:
        hdr  = rows[daily_hdr]
        vals = rows[daily_hdr + 1]
        for ci, c in enumerate(hdr):
            if c is not None and 'Día' in str(c):
                try:   dn = int(str(c).replace('Día','').strip())
                except: continue
                daily_labels.append(dn)
                daily_totals.append(round(v(vals[ci]) if ci < len(vals) else 0))

    # ── Meta / cumplimiento / deberíamos ir ───────────────────────────────────
    meta_periodo = prod_total = pct_cumpl = deberia_ir = hacer_dia = 0

    for ri, row in enumerate(rows):
        if any(c is not None and 'Meta PERIODO' in str(c) for c in row):
            nrow = rows[ri+1] if ri+1 < N else []
            nums = [(ci, x) for ci, x in enumerate(nrow) if isinstance(x, (int, float))]
            if nums:
                by_idx = {ci: x for ci, x in nums}
                meta_periodo = by_idx.get(2, 0) or by_idx.get(3, 0) or (nums[0][1] if nums else 0)
                # producción = segundo número numérico
                if len(nums) >= 2: prod_total = nums[1][1]
                # pct cumplimiento
                for ci, x in nums:
                    if 0 < x < 2 and ci > 3: pct_cumpl = x; break
            # Deberíamos ir: buscar texto 'Deberiamos' en misma fila
            for ci2, x in enumerate(nrow):
                if x is not None and 'Deberia' in str(x):
                    if ci2+1 < len(nrow) and isinstance(nrow[ci2+1], (int,float)):
                        deberia_ir = nrow[ci2+1]
                    break
            break

    for ri, row in enumerate(rows):
        if any(c is not None and 'Hacer por día' in str(c) for c in row):
            nums = [x for x in row if isinstance(x, (int,float)) and x > 0]
            if nums: hacer_dia = min(nums)   # el menor es el real (no la meta total)
            break

    # Calcular deberíamos ir si no se encontró en la hoja
    if deberia_ir == 0 and dias_periodo > 0:
        deberia_ir = dias_facturados / dias_periodo

    # Normalizar: si viene como fracción (0.67) → convertir a porcentaje
    if 0 < deberia_ir <= 1.0:
        deberia_ir_pct = round(deberia_ir * 100, 1)
    else:
        deberia_ir_pct = round(float(deberia_ir), 1)

    if 0 < pct_cumpl <= 1.0:
        pct_pct = round(pct_cumpl * 100, 1)
    else:
        pct_pct = round(float(pct_cumpl), 1)

    # ── Plataformas (sección por página) ──────────────────────────────────────
    PLAT_MAP = {"Flirt4free":"F4F","Stripchat":"SC","Chaturbate":"CB","CamSoda":"CAM","Streamate":"STR"}
    plats = {"F4F":0,"SC":0,"CB":0,"CAM":0,"STR":0}
    for row in rows:
        c2 = str_cell(row[2]) if len(row)>2 else ""
        if c2 in PLAT_MAP:
            plats[PLAT_MAP[c2]] += v(row[3]) if len(row)>3 else 0

    # ── Modelos ────────────────────────────────────────────────────────────────
    nombre_row = None
    for ri, row in enumerate(rows):
        if len(row)>2 and row[2] is not None and 'Nombre modelo' in str(row[2]):
            nombre_row = ri

    modelos = []
    mon_com  = {}

    if nombre_row is not None:
        # Detectar estructura: 4 plataformas (Ene-Jun) vs 5 con Streamate (Jul-Ago)
        plat_hdr = rows[nombre_row+1] if nombre_row+1 < N else []
        has_streamate = any(c is not None and 'Streamate' in str(c) for c in plat_hdr)
        if has_streamate:
            # 5 plataformas: F4F=3 SC=4 CB=5 CAM=6 STR=7 TOT=8 RETO=9 COM=11 MON=12
            _ci = dict(f4f=3,sc=4,cb=5,cam=6,str_=7,tot=8,reto=9,com=11,mon=12)
        else:
            # 4 plataformas: F4F=3 SC=4 CB=5 CAM=6 TOT=7 RETO=8 COM=10 MON=11
            _ci = dict(f4f=3,sc=4,cb=5,cam=6,str_=None,tot=7,reto=8,com=10,mon=11)

        def _cv(row, key):
            """Retorna valor numérico de columna según layout detectado."""
            idx = _ci[key]
            if idx is None or idx >= len(row): return 0
            return row[idx]  # puede ser None, int, float, str

        for ri in range(nombre_row+2, N):
            row = rows[ri]
            if len(row) < 3: continue
            nombre = str_cell(row[2]) if len(row)>2 else ""
            if nombre == "TOTAL": break
            # Solo filas reales de modelo: col 1 debe ser numérico (reto target)
            if not isinstance(row[1] if len(row)>1 else None, (int,float)):
                continue
            if not nombre: continue
            reto_meta  = v(row[1])
            f4f        = v(_cv(row,'f4f'))
            sc         = v(_cv(row,'sc'))
            cb         = v(_cv(row,'cb'))
            cam        = v(_cv(row,'cam'))
            str_val    = v(_cv(row,'str_'))
            total      = v(_cv(row,'tot'))
            reto_texto = str_cell(_cv(row,'reto'))
            # comision en USD ($0, $5, $15 por modelo)
            com_usd    = round(v(_cv(row,'com')))
            monitor_nm = str_cell(_cv(row,'mon'))
            modelos.append({
                "nombre":      nombre,
                "reto_meta":   round(reto_meta),
                "f4f":         round(f4f),
                "sc":          round(sc),
                "cb":          round(cb),
                "cam":         round(cam),
                "str":         round(str_val),
                "total":       round(total),
                "reto_cumple": 'Cumplió' in reto_texto,
                "com_usd":     com_usd,
                "monitor":     monitor_nm,
            })

        # Comisiones monitor (USD)
        for ri in range(nombre_row+1, N):
            row = rows[ri]
            cells = [str_cell(c) for c in row if c is not None]
            if 'MONITOR' in cells:
                for rj in range(ri+1, min(ri+8, N)):
                    mr = rows[rj]
                    nm = ""
                    nums = []
                    for c in mr:
                        if c is None: continue
                        if isinstance(c, str) and c.strip() and not nm: nm = c.strip()
                        elif isinstance(c, (int, float)): nums.append(c)
                    if nm and len(nums) >= 4:
                        mon_com[nm] = {
                            "meta_base": round(nums[0]),
                            "reto_est":  round(nums[1]),
                            "x_modelo":  round(nums[2]),
                            "total_usd": round(nums[3]),
                        }
                    elif nm and nums:
                        mon_com[nm] = {"meta_base":0,"reto_est":0,"x_modelo":round(nums[-1]),"total_usd":round(nums[-1])}
                break

    # Recalcular producción si no se encontró en meta row
    if prod_total == 0 and modelos:
        prod_total = sum(m["total"] for m in modelos)

    return {
        "sheet":        sheet_name,
        "mes":          mes_label,
        "qid":          qid,
        "label":        "Quincena 1 (1-15)" if qid=="q1" else "Quincena 2 (16-31)",
        "fecha_ini":    fmt_date(fecha_ini),
        "fecha_fin":    fmt_date(fecha_fin),
        "dias_periodo": int(dias_periodo) if dias_periodo else (15 if qid=="q1" else 16),
        "dias_fact":    int(dias_facturados),
        "meta":         round(meta_periodo),
        "produccion":   round(prod_total),
        "pct":          pct_pct,
        "deberia_ir":   deberia_ir_pct,
        "hacer_dia":    round(hacer_dia),
        "daily_labels": daily_labels,
        "daily_totals": daily_totals,
        "plats":        plats,
        "modelos":      modelos,
        "mon_com":      mon_com,
    }

# ── TOP 20 GRUPO — desde ALIADOS del dashboard grupo ─────────────────────────
def build_top20_grupo():
    if not os.path.exists(GRUPO_HTML):
        print("   ⚠️  grupo dashboard no encontrado — TOP 20 Grupo omitido")
        return {}
    with open(GRUPO_HTML, encoding="utf-8") as f:
        html = f.read()
    m = re.search(r'(?:const|var)\s+ALIADOS\s*=\s*_b64dec\(["\']([A-Za-z0-9+/=]+)["\']\)', html)
    if not m:
        print("   ⚠️  ALIADOS no encontrado en grupo dashboard")
        return {}
    aliados = json.loads(base64.b64decode(m.group(1)).decode('utf-8'))

    def model_total_for_period(day_data, qid):
        total = 0.0
        for dk, plats in day_data.items():
            try: dn = int(dk)
            except: continue
            if qid == "q1" and dn > 15: continue
            if qid == "q2" and dn <= 15: continue
            total += sum(vv for vv in plats.values() if vv is not None)
        return total

    result = {}  # {mes: {qid: [{modelo, estudio, produccion}]}}
    meses = list(QUINCENAS.keys())

    for mes in meses:
        result[mes] = {}
        for _, qid in QUINCENAS[mes]:
            rows_list = []
            for studio_key, studio in aliados.items():
                mes_data = studio['data'].get(mes, {})
                for mod_nm, day_data in mes_data.get('modelos', {}).items():
                    if not day_data: continue
                    prod = model_total_for_period(day_data, qid)
                    if prod > 0:
                        rows_list.append({
                            "modelo":    mod_nm,
                            "estudio":   studio_key,
                            "produccion":round(prod),
                        })
            rows_list.sort(key=lambda x: -x['produccion'])
            result[mes][qid] = rows_list[:20]

    return result

# ── TOP 20 FORNAX — desde datos de quincena ────────────────────────────────────
def build_top20_fornax(periodos):
    result = {}
    for mes, qs in periodos.items():
        result[mes] = {}
        for qid, p in qs.items():
            if not p: result[mes][qid] = []; continue
            rows_list = sorted(
                [{"modelo": m["nombre"], "estudio": "Fornax Studios", "produccion": m["total"]}
                 for m in p["modelos"] if m["total"] > 0],
                key=lambda x: -x["produccion"]
            )
            result[mes][qid] = rows_list[:20]
    return result

# ── BUILD DATA ────────────────────────────────────────────────────────────────
def build_data():
    print(f"📖 Cargando Fornax2.xlsx...")
    wb = load_workbook(EXCEL_FORNAX, data_only=True)

    bd = read_bd_modelos(wb)
    print(f"✅ BD Modelos: {len(bd)} total")

    periodos = {}
    for mes, quincenas in QUINCENAS.items():
        periodos[mes] = {}
        for sname, qid in quincenas:
            p = read_quincena(wb, sname, mes, qid)
            if p:
                tag = "✅" if p["produccion"] > 0 else "⬜"
                print(f"   {tag} {sname.strip()}: {len(p['modelos'])} mod | prod={p['produccion']:,.0f} | días={p['dias_fact']} | DebIr={p['deberia_ir']}%")
                periodos[mes][qid] = p
            else:
                print(f"   ⚠️  '{sname}': no encontrada")
                periodos[mes][qid] = None

    # ── Corrección histórica de monitor ──────────────────────────────────────────
    # Hasta el 15 de julio (inclusive) la monitora era María Isabel Serna Pareja.
    # A partir del 16 de julio es Juliana Lara Navarro.
    # El Excel registra "Juliana Lara Navarro" en todos los períodos, por lo que
    # corregimos en post-proceso los períodos anteriores a Julio Q2.
    _JULIANA = "Juliana Lara Navarro"
    _MARIA   = "María Isabel Serna Pareja"
    _PERIODOS_MARIA = {
        "Enero":   ["q1","q2"],
        "Febrero": ["q1","q2"],
        "Marzo":   ["q1","q2"],
        "Abril":   ["q1","q2"],
        "Mayo":    ["q1","q2"],
        "Junio":   ["q1","q2"],
        "Julio":   ["q1"],          # Solo Q1 (1-15); Q2 (16-31) ya es Juliana
    }
    corr_count = 0
    for mes, qids in _PERIODOS_MARIA.items():
        for qid in qids:
            p = periodos.get(mes, {}).get(qid)
            if not p:
                continue
            for m in p.get("modelos", []):
                if m.get("monitor") == _JULIANA:
                    m["monitor"] = _MARIA
                    corr_count += 1
            # Corregir también la clave en mon_com si existe
            if _JULIANA in p.get("mon_com", {}):
                p["mon_com"][_MARIA] = p["mon_com"].pop(_JULIANA)
    if corr_count:
        print(f"   ✅ Corrección histórica: {corr_count} asignaciones "
              f"'{_JULIANA}' → '{_MARIA}' (Ene–Jul Q1)")

    print("\n📊 Construyendo TOP 20 Grupo desde ALIADOS...")
    t20_grupo = build_top20_grupo()
    for mes, qs in t20_grupo.items():
        for qid, lst in qs.items():
            print(f"   {mes} {qid}: {len(lst)} en top20 grupo")

    print("\n🏆 Construyendo TOP 20 Fornax...")
    t20_fornax = build_top20_fornax(periodos)

    monitores = sorted(set(m["monitor"] for m in bd if m["monitor"] and m["monitor"] != "Sin monitor"))
    niveles   = {m["modelo"]: m.get("nivel","") for m in bd}

    return {
        "generado":    HOY_STR,
        "corte":       CORTE_STR,
        "bd_modelos":  bd,
        "monitores_bd":monitores,
        "periodos":    periodos,
        "niveles":     niveles,
        "top20_grupo": t20_grupo,
        "top20_fornax":t20_fornax,
    }

# ── HTML TEMPLATE ─────────────────────────────────────────────────────────────
HTML = r"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Monitores · Fornax Studios</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
<style>
:root{
  --bg:#0f1117;--bg2:#161b27;--bg3:#1e2535;--bg4:#252d40;
  --border:#2a3452;--text:#e8eaf0;--text2:#9ba3bf;--text3:#5a6480;
  --accent:#6366f1;--green:#22c55e;--red:#ef4444;
  --yellow:#f59e0b;--blue:#3b82f6;--purple:#a855f7;--cyan:#06b6d4;--pink:#ec4899;
}
*{box-sizing:border-box;margin:0;padding:0;-webkit-tap-highlight-color:transparent}
html{font-size:15px}
body{background:var(--bg);color:var(--text);font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',system-ui,sans-serif;min-height:100vh;overflow-x:hidden}

/* ── HEADER ── */
.hdr{background:linear-gradient(135deg,#1a1f35,#0f1117);border-bottom:1px solid var(--border);padding:12px 16px;position:sticky;top:0;z-index:200}
.hdr-in{max-width:1380px;margin:0 auto;display:flex;align-items:center;gap:12px;flex-wrap:wrap}
.logo-box{width:36px;height:36px;background:linear-gradient(135deg,var(--accent),var(--purple));border-radius:9px;display:grid;place-items:center;font-size:17px;flex-shrink:0}
.logo h1{font-size:14px;font-weight:700;color:#fff;line-height:1.2}
.logo p{font-size:10px;color:var(--text3)}
.hdr-meta{margin-left:auto;font-size:11px;color:var(--text3);text-align:right;line-height:1.6}
.hdr-meta strong{color:var(--text2)}

/* ── CONTROLES ── */
.ctrl-bar{background:var(--bg2);border-bottom:1px solid var(--border);padding:10px 16px;position:sticky;top:61px;z-index:199;overflow-x:auto;-webkit-overflow-scrolling:touch}
.ctrl-in{max-width:1380px;margin:0 auto;display:flex;gap:8px;align-items:center;flex-wrap:nowrap;min-width:max-content}
.ctrl-grp{display:flex;align-items:center;gap:6px}
.ctrl-lbl{font-size:11px;color:var(--text3);white-space:nowrap}
select{background:var(--bg3);border:1px solid var(--border);border-radius:8px;color:var(--text);padding:8px 12px;font-size:13px;outline:none;cursor:pointer;-webkit-appearance:none;appearance:none;min-width:100px}
select:focus{border-color:var(--accent)}
.sep{width:1px;height:22px;background:var(--border);flex-shrink:0}
.btn-sm{background:var(--bg3);border:1px solid var(--border);border-radius:8px;color:var(--text);padding:7px 14px;font-size:12px;cursor:pointer;white-space:nowrap}
.btn-sm:hover{background:var(--bg4);border-color:var(--accent)}

/* ── MAIN ── */
.main{max-width:1380px;margin:0 auto;padding:16px}

/* ── KPIs ── */
.kpi-row{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-bottom:16px}
.kpi{background:var(--bg2);border:1px solid var(--border);border-radius:11px;padding:14px 12px;position:relative;overflow:hidden}
.kpi::after{content:'';position:absolute;top:0;left:0;right:0;height:3px;background:var(--c,var(--blue))}
.kpi-lbl{font-size:10px;color:var(--text2);text-transform:uppercase;letter-spacing:.5px;margin-bottom:4px}
.kpi-val{font-size:20px;font-weight:700;color:#fff;line-height:1.2}
.kpi-sub{font-size:10px;color:var(--text3);margin-top:3px}
.pbar{background:var(--bg3);border-radius:99px;height:5px;margin-top:7px;overflow:hidden}
.pfill{height:100%;border-radius:99px;transition:width .5s}

/* Ritmo indicator */
.ritmo{display:inline-flex;align-items:center;gap:5px;font-size:11px;font-weight:600;padding:3px 8px;border-radius:99px;margin-top:5px}
.ritmo.arriba{background:rgba(34,197,94,.15);color:var(--green)}
.ritmo.igual{background:rgba(245,158,11,.12);color:var(--yellow)}
.ritmo.abajo{background:rgba(239,68,68,.1);color:var(--red)}

/* ── TABS ── */
.tabs{display:flex;gap:2px;background:var(--bg3);border-radius:9px;padding:3px;width:fit-content;margin-bottom:14px;overflow-x:auto;-webkit-overflow-scrolling:touch;max-width:100%}
.tab{padding:8px 14px;border-radius:7px;cursor:pointer;font-size:13px;color:var(--text2);transition:.15s;white-space:nowrap;flex-shrink:0}
.tab.on{background:var(--bg2);color:#fff}

/* ── CHART ── */
.chart-card{background:var(--bg2);border:1px solid var(--border);border-radius:11px;padding:16px;margin-bottom:16px}
.chart-card h3{font-size:11px;font-weight:600;color:var(--text2);margin-bottom:12px;text-transform:uppercase;letter-spacing:.5px}
.chart-wrap{height:190px;position:relative}

/* ── MONITOR GRID ── */
.mon-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(320px,1fr));gap:14px}
.mon-card{background:var(--bg2);border:1px solid var(--border);border-radius:11px;overflow:hidden}
.mon-hdr{padding:12px 14px;display:flex;align-items:center;gap:10px;border-bottom:1px solid var(--border)}
.mon-av{width:34px;height:34px;border-radius:50%;display:grid;place-items:center;font-size:13px;font-weight:700;flex-shrink:0}
.mon-nm{font-size:13px;font-weight:600;color:#fff}
.mon-meta{font-size:10px;color:var(--text3)}
.mon-com{margin-left:auto;text-align:right}
.mon-com-val{font-size:15px;font-weight:700;color:var(--yellow)}
.mon-com-lbl{font-size:10px;color:var(--text3)}
.mon-body{padding:10px 14px;display:flex;flex-direction:column;gap:6px}

/* ── MODELO ROW ── */
.mod-row{background:var(--bg3);border-radius:8px;padding:9px 11px}
.mod-top{display:flex;align-items:center;gap:5px;flex-wrap:wrap;margin-bottom:5px}
.mod-nm{font-size:12px;font-weight:600;color:#fff;flex:1 1 auto;min-width:100px}
.badge{font-size:10px;padding:2px 6px;border-radius:99px;font-weight:600;flex-shrink:0;white-space:nowrap}
.b-ok{background:rgba(34,197,94,.15);color:var(--green);border:1px solid rgba(34,197,94,.25)}
.b-no{background:rgba(239,68,68,.1);color:var(--red);border:1px solid rgba(239,68,68,.2)}
.b-c2{background:rgba(245,158,11,.15);color:var(--yellow);border:1px solid rgba(245,158,11,.3)}
.b-c1{background:rgba(59,130,246,.12);color:var(--blue);border:1px solid rgba(59,130,246,.25)}
.b-c0{background:rgba(90,100,128,.1);color:var(--text3);border:1px solid rgba(90,100,128,.2)}
.b-orb{background:rgba(99,102,241,.12);color:var(--accent);border:1px solid rgba(99,102,241,.25)}
.b-lla{background:rgba(248,113,113,.12);color:#f87171;border:1px solid rgba(248,113,113,.25)}
.b-nuc{background:rgba(168,85,247,.12);color:var(--purple);border:1px solid rgba(168,85,247,.3)}
.mod-bar{display:flex;align-items:center;gap:7px;margin-bottom:4px}
.mod-bar-t{font-size:10px;color:var(--text3);white-space:nowrap}
.mod-prog{flex:1;background:var(--bg4);border-radius:99px;height:5px;overflow:hidden}
.mod-fill{height:100%;border-radius:99px}
.mod-pct{font-size:10px;color:var(--text3);min-width:30px;text-align:right}
.plats{display:flex;gap:4px;flex-wrap:wrap}
.pc{font-size:10px;color:var(--text3);background:var(--bg4);border-radius:4px;padding:1px 5px}
.pc b{color:var(--text)}

/* ── TABLA ── */
.tbl-wrap{overflow-x:auto;border-radius:11px;border:1px solid var(--border);-webkit-overflow-scrolling:touch}
table{width:100%;border-collapse:collapse}
th{background:var(--bg3);padding:9px 10px;text-align:left;font-size:10px;font-weight:600;color:var(--text2);text-transform:uppercase;letter-spacing:.5px;white-space:nowrap;border-bottom:1px solid var(--border)}
td{padding:8px 10px;border-bottom:1px solid rgba(42,52,82,.4);font-size:12px;white-space:nowrap}
tr:last-child td{border-bottom:none}
tr:hover td{background:rgba(255,255,255,.015)}
.tr{text-align:right;font-family:monospace}
.tw{font-weight:700;color:#fff}
.tfoot{background:var(--bg3)}

/* ── PLATAFORMAS ── */
.plat-grid{display:grid;grid-template-columns:repeat(5,1fr);gap:10px;margin-bottom:16px}
.plat-c{background:var(--bg2);border:1px solid var(--border);border-radius:10px;padding:12px;text-align:center}
.plat-nm{font-size:10px;font-weight:600;text-transform:uppercase;margin-bottom:5px}
.plat-vl{font-size:18px;font-weight:700;color:#fff}
.plat-pt{font-size:10px;color:var(--text3);margin-top:2px}

/* ── TOP 20 ── */
.top20-ctrl{display:flex;gap:8px;margin-bottom:14px;flex-wrap:wrap}
.sub-btn{background:var(--bg3);border:1px solid var(--border);border-radius:8px;color:var(--text2);padding:8px 16px;font-size:13px;cursor:pointer}
.sub-btn.on{background:var(--accent);border-color:var(--accent);color:#fff}
.top20-list{display:flex;flex-direction:column;gap:6px}
.t20-row{background:var(--bg2);border:1px solid var(--border);border-radius:9px;padding:10px 14px;display:flex;align-items:center;gap:12px}
.t20-pos{font-size:13px;font-weight:700;color:var(--text3);min-width:28px;text-align:center}
.t20-pos.gold{color:var(--yellow)}
.t20-pos.silver{color:#9ba3bf}
.t20-pos.bronze{color:#cd7f32}
.t20-name{font-size:13px;font-weight:600;color:#fff;flex:1;min-width:0;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.t20-studio{font-size:10px;color:var(--text3)}
.t20-prod{font-size:13px;font-weight:700;color:var(--cyan);margin-left:auto;font-family:monospace;white-space:nowrap}

/* ── EMPTY ── */
.empty{text-align:center;padding:32px;color:var(--text3);font-size:13px}

/* ── RESPONSIVE ── */
@media(max-width:900px){
  .kpi-row{grid-template-columns:repeat(2,1fr)}
  .plat-grid{grid-template-columns:repeat(3,1fr)}
  .mon-grid{grid-template-columns:1fr}
  .hdr-meta{display:none}
}
@media(max-width:600px){
  html{font-size:14px}
  .kpi-row{grid-template-columns:repeat(2,1fr)}
  .kpi{padding:12px 10px}
  .kpi-val{font-size:18px}
  .plat-grid{grid-template-columns:repeat(2,1fr)}
  .main{padding:12px 10px}
  .chart-wrap{height:160px}
  .hdr{padding:10px 12px}
  .ctrl-bar{padding:8px 12px;top:57px}
  select{padding:7px 10px;font-size:12px}
  .tab{padding:7px 12px;font-size:12px}
  .t20-row{padding:9px 12px;gap:8px}
  .t20-name{font-size:12px}
  .t20-prod{font-size:12px}
  .mon-hdr{padding:10px 12px}
  .mon-body{padding:8px 12px}
  table{font-size:11px}
  th,td{padding:7px 8px}
}
@media(max-width:380px){
  .kpi-val{font-size:16px}
}
</style>
</head>
<body>

<!-- HEADER -->
<div class="hdr">
  <div class="hdr-in">
    <div class="logo-box">🌟</div>
    <div class="logo">
      <h1>Monitores · Fornax Studios</h1>
      <p>Seguimiento por monitor y modelo</p>
    </div>
    <div class="hdr-meta">
      <div>Actualizado: <strong id="h-gen"></strong></div>
      <div>Corte: <strong id="h-cor"></strong></div>
    </div>
  </div>
</div>

<!-- CONTROLES -->
<div class="ctrl-bar">
  <div class="ctrl-in">
    <div class="ctrl-grp">
      <span class="ctrl-lbl">Mes</span>
      <select id="s-mes"></select>
    </div>
    <div class="ctrl-grp">
      <span class="ctrl-lbl">Período</span>
      <select id="s-qid">
        <option value="q1">Q1 (1–15)</option>
        <option value="q2">Q2 (16–31)</option>
      </select>
    </div>
    <div class="sep"></div>
    <div class="ctrl-grp">
      <span class="ctrl-lbl">Monitor</span>
      <select id="s-mon"><option value="">Todos</option></select>
    </div>
    <div class="ctrl-grp">
      <span class="ctrl-lbl">Estado</span>
      <select id="s-est">
        <option value="">Todos</option>
        <option value="Activo">Activo</option>
        <option value="Inactivo">Inactivo</option>
        <option value="Retirado">Retirado</option>
      </select>
    </div>
    <div class="sep"></div>
    <button class="btn-sm" onclick="exportCSV()">⬇ CSV</button>
  </div>
</div>

<!-- MAIN -->
<div class="main">
  <div class="kpi-row" id="kpis"></div>

  <div class="tabs" id="main-tabs">
    <div class="tab on" onclick="setTab('mon',this)">Por Monitor</div>
    <div class="tab" onclick="setTab('tbl',this)">Tabla</div>
    <div class="tab" onclick="setTab('plt',this)">Plataformas</div>
    <div class="tab" onclick="setTab('top',this)">Top 20</div>
  </div>

  <!-- Vista: Por Monitor -->
  <div id="v-mon">
    <div class="chart-card">
      <h3>📈 Producción diaria del estudio</h3>
      <div class="chart-wrap"><canvas id="c-daily"></canvas></div>
    </div>
    <div class="mon-grid" id="mon-grid"></div>
  </div>

  <!-- Vista: Tabla -->
  <div id="v-tbl" style="display:none">
    <div class="tbl-wrap"><table id="t-global"></table></div>
  </div>

  <!-- Vista: Plataformas -->
  <div id="v-plt" style="display:none">
    <div class="plat-grid" id="plat-grid"></div>
    <div class="chart-card">
      <h3>🥧 Distribución por plataforma</h3>
      <div class="chart-wrap"><canvas id="c-plat"></canvas></div>
    </div>
  </div>

  <!-- Vista: Top 20 -->
  <div id="v-top" style="display:none">
    <div class="top20-ctrl">
      <button class="sub-btn on" id="btn-t20-fx" onclick="setTop('fornax',this)">🏠 Fornax</button>
      <button class="sub-btn" id="btn-t20-gp" onclick="setTop('grupo',this)">🌐 Grupo Empresarial</button>
    </div>
    <div class="top20-list" id="top20-list"></div>
  </div>
</div>

<script>
// ── DETECCIÓN DE SMARTPHONE ───────────────────────────────────────────────────
// Basada en User Agent (no en ancho de pantalla) para evitar falsos positivos
// con ventanas pequeñas de escritorio.
// Detecta: iPhone, iPod, Android teléfonos (Android+Mobile), Windows Phone, etc.
// NO detecta: iPad, tablets Android (sin "Mobile"), laptops, desktops.
var _PHONE = (function() {
  var ua = navigator.userAgent;
  return /Android.*Mobile|iPhone|iPod|Windows Phone|BlackBerry|IEMobile|Opera Mobi/i.test(ua);
})();

if (_PHONE) {
  document.body.innerHTML =
    '<div style="min-height:100vh;background:#0f1117;display:flex;flex-direction:column;' +
    'align-items:center;justify-content:center;padding:40px 24px;text-align:center;' +
    'font-family:-apple-system,BlinkMacSystemFont,\'Segoe UI\',sans-serif">' +
    '<div style="font-size:64px;margin-bottom:28px">💻</div>' +
    '<div style="color:#e8eaf0;font-size:20px;font-weight:700;margin-bottom:14px;line-height:1.3">' +
    'Esta información solo se puede<br>visualizar desde un computador.' +
    '</div>' +
    '<div style="color:#9ba3bf;font-size:14px;line-height:1.7;max-width:300px">' +
    'Por favor, ingresa desde un computador o laptop para consultar el Dashboard de Monitores.' +
    '</div>' +
    '<div style="margin-top:32px;padding:12px 24px;background:#1e2535;border:1px solid #2a3452;' +
    'border-radius:10px;color:#5a6480;font-size:12px">' +
    'Fornax Studios · Dashboard de Monitores' +
    '</div>' +
    '</div>';
}

// ── VERIFICACIÓN DE ACCESO (PIN) ─────────────────────────────────────────────
var _AUTH = (function() {
  try { return sessionStorage.getItem('mon_auth') === 'ok_2017'; } catch(e) { return false; }
})();

if (!_PHONE && !_AUTH) {
  document.body.innerHTML =
    '<div style="min-height:100vh;background:#0f1117;display:flex;flex-direction:column;' +
    'align-items:center;justify-content:center;font-family:-apple-system,BlinkMacSystemFont,\'Segoe UI\',sans-serif">' +
    '<div style="background:#161b27;border:1px solid #2a3452;border-radius:16px;padding:40px 36px;' +
    'width:100%;max-width:360px;text-align:center;box-shadow:0 20px 60px rgba(0,0,0,.5)">' +
    '<div style="width:52px;height:52px;background:linear-gradient(135deg,#6366f1,#a855f7);border-radius:13px;' +
    'display:flex;align-items:center;justify-content:center;font-size:22px;margin:0 auto 20px">🌟</div>' +
    '<div style="color:#e8eaf0;font-size:17px;font-weight:700;margin-bottom:4px">Dashboard de Monitores</div>' +
    '<div style="color:#5a6480;font-size:12px;margin-bottom:28px">Fornax Studios · Acceso restringido</div>' +
    '<input id="pin-inp" type="password" maxlength="10" placeholder="Ingresa el PIN" ' +
    'style="width:100%;background:#1e2535;border:1px solid #2a3452;border-radius:9px;color:#e8eaf0;' +
    'padding:13px 16px;font-size:16px;outline:none;text-align:center;letter-spacing:6px;' +
    'font-family:monospace;box-sizing:border-box;margin-bottom:12px">' +
    '<button id="pin-btn" ' +
    'style="width:100%;background:linear-gradient(135deg,#6366f1,#4f46e5);border:none;border-radius:9px;' +
    'color:#fff;padding:13px;font-size:14px;font-weight:600;cursor:pointer;letter-spacing:.3px">' +
    'Ingresar</button>' +
    '<div id="pin-err" style="display:none;margin-top:14px;color:#ef4444;font-size:12px;' +
    'background:rgba(239,68,68,.1);border:1px solid rgba(239,68,68,.2);border-radius:7px;padding:8px 12px">' +
    '❌ PIN incorrecto. Intenta de nuevo.</div>' +
    '</div></div>';

  document.getElementById('pin-inp').focus();

  function _checkPin() {
    var val = document.getElementById('pin-inp').value.trim();
    if (val === '2017') {
      try { sessionStorage.setItem('mon_auth', 'ok_2017'); } catch(e) {}
      location.reload();
    } else {
      document.getElementById('pin-err').style.display = 'block';
      document.getElementById('pin-inp').value = '';
      document.getElementById('pin-inp').focus();
    }
  }
  document.getElementById('pin-btn').addEventListener('click', _checkPin);
  document.getElementById('pin-inp').addEventListener('keydown', function(e) {
    if (e.key === 'Enter') _checkPin();
    if (document.getElementById('pin-err').style.display !== 'none') {
      document.getElementById('pin-err').style.display = 'none';
    }
  });
}

const D = DATA_PLACEHOLDER;
const PLAT = {f4f:'Flirt4Free',sc:'Stripchat',cb:'Chaturbate',cam:'CamSoda',str:'Streamate'};
const PC   = {f4f:'#6366f1',sc:'#22c55e',cb:'#f59e0b',cam:'#3b82f6',str:'#ec4899'};
const MC   = ['#6366f1','#22c55e','#f59e0b','#ec4899','#06b6d4','#a855f7'];

if (!_PHONE && _AUTH) {  // ── INICIO BLOQUE DESKTOP AUTENTICADO ────────────────

let curTab = 'mon', curTop = 'fornax', cDay=null, cPlt=null;

document.getElementById('h-gen').textContent = D.generado;
document.getElementById('h-cor').textContent = D.corte;

// Mes selector
const sMes = document.getElementById('s-mes');
Object.keys(D.periodos).forEach(m => {
  const o = document.createElement('option');
  o.value = m; o.textContent = m; sMes.appendChild(o);
});
const lastMes = [...Object.keys(D.periodos)].reverse()
  .find(m => D.periodos[m]?.q1 || D.periodos[m]?.q2) || Object.keys(D.periodos)[0];
sMes.value = lastMes;

// Monitor selector
const sMon = document.getElementById('s-mon');
D.monitores_bd.forEach(m => {
  const o = document.createElement('option');
  o.value = m;
  // Abbreviate long names on mobile
  o.textContent = m.split(' ').slice(0,3).join(' ');
  sMon.appendChild(o);
});

const sQid = document.getElementById('s-qid');

function bestQid(mes) {
  const p = D.periodos[mes];
  if (!p) return 'q1';
  if (p.q2?.produccion > 0) return 'q2';
  if (p.q1?.produccion > 0) return 'q1';
  return 'q1';
}
sQid.value = bestQid(lastMes);

sMes.addEventListener('change', () => { sQid.value = bestQid(sMes.value); render(); });
sQid.addEventListener('change', render);
sMon.addEventListener('change', render);
document.getElementById('s-est').addEventListener('change', render);

function getPeriod()  { return D.periodos[sMes.value]?.[sQid.value] || null; }
function bdOf(nombre) { return D.bd_modelos.find(m => m.modelo === nombre) || null; }
function fmt(n)       { return Math.round(n).toLocaleString('es-CO'); }
function fmtPct(n)    { return (+n).toFixed(1)+'%'; }
function fmtUSD(n)    { return '$'+Math.round(n)+' USD'; }

function filteredMods(p) {
  const mon = sMon.value;
  const est = document.getElementById('s-est').value;
  return (p?.modelos||[]).filter(m => {
    if (mon && m.monitor !== mon) return false;
    if (est) { const bd = bdOf(m.nombre); if(bd && bd.estado !== est) return false; }
    return true;
  });
}

// ── KPIs ──────────────────────────────────────────────────────────────────────
function renderKPIs(p) {
  if (!p) { document.getElementById('kpis').innerHTML='<div class="empty">Sin datos</div>'; return; }
  const mods = filteredMods(p);
  const prod = mods.reduce((s,m)=>s+m.total,0);
  const activos = mods.filter(m=>m.total>0).length;
  const cumplen = mods.filter(m=>m.reto_cumple).length;
  const pct = p.meta > 0 ? prod/p.meta*100 : p.pct;
  const debIr = p.deberia_ir || 0;
  const promDia = p.dias_fact > 0 ? prod/p.dias_fact : 0;
  const pctColor = pct>=100?'var(--green)':pct>=70?'var(--yellow)':'var(--red)';

  // Ritmo
  const diff = pct - debIr;
  let ritmoClass, ritmoTxt;
  if (diff >= -2) { ritmoClass='arriba'; ritmoTxt='🟢 Por encima'; }
  else if (diff >= -8) { ritmoClass='igual'; ritmoTxt='🟡 En ritmo'; }
  else { ritmoClass='abajo'; ritmoTxt='🔴 Por debajo'; }

  // Monitor commissions
  const totalUSD = Object.values(p.mon_com||{}).reduce((s,c)=>s+(c.total_usd||0),0);

  document.getElementById('kpis').innerHTML = `
    <div class="kpi" style="--c:var(--blue)">
      <div class="kpi-lbl">Producción</div>
      <div class="kpi-val">${fmt(prod)}</div>
      <div class="kpi-sub">Meta: ${fmt(p.meta)}</div>
      <div class="pbar"><div class="pfill" style="width:${Math.min(pct,100)}%;background:${pctColor}"></div></div>
    </div>
    <div class="kpi" style="--c:var(--green)">
      <div class="kpi-lbl">Cumplimiento</div>
      <div class="kpi-val" style="color:${pctColor}">${fmtPct(pct)}</div>
      <div class="kpi-sub">${p.dias_fact} / ${p.dias_periodo} días</div>
    </div>
    <div class="kpi" style="--c:var(--cyan)">
      <div class="kpi-lbl">Deberíamos ir</div>
      <div class="kpi-val" style="color:var(--cyan)">${fmtPct(debIr)}</div>
      <div class="ritmo ${ritmoClass}">${ritmoTxt}</div>
    </div>
    <div class="kpi" style="--c:var(--yellow)">
      <div class="kpi-lbl">Promedio / día</div>
      <div class="kpi-val">${fmt(promDia)}</div>
      <div class="kpi-sub">${p.hacer_dia>0?'Falta: '+fmt(p.hacer_dia)+'/día':''}</div>
    </div>
    <div class="kpi" style="--c:var(--purple)">
      <div class="kpi-lbl">Reto Estrella ✅</div>
      <div class="kpi-val">${cumplen}<span style="font-size:13px;color:var(--text3)"> / ${mods.length}</span></div>
      <div class="kpi-sub">modelos cumplen</div>
    </div>
    <div class="kpi" style="--c:var(--pink)">
      <div class="kpi-lbl">Comisiones</div>
      <div class="kpi-val" style="color:var(--yellow)">${fmtUSD(totalUSD)}</div>
      <div class="kpi-sub">total monitores</div>
    </div>
    <div class="kpi" style="--c:var(--orange,#f97316)">
      <div class="kpi-lbl">Modelos activos</div>
      <div class="kpi-val">${activos}</div>
      <div class="kpi-sub">con producción</div>
    </div>
    <div class="kpi" style="--c:var(--accent)">
      <div class="kpi-lbl">Período</div>
      <div class="kpi-val" style="font-size:13px;font-weight:600">${p.label}</div>
      <div class="kpi-sub">${p.fecha_ini} → ${p.fecha_fin}</div>
    </div>`;
}

// ── CHART DAILY ───────────────────────────────────────────────────────────────
function _makeChartBar(ctx, labels, data) {
  return new Chart(ctx, {
    type:'bar',
    data:{labels,datasets:[{label:'Tokens',data,backgroundColor:'rgba(99,102,241,.7)',borderColor:'#6366f1',borderWidth:1,borderRadius:4}]},
    options:{
      responsive:true,maintainAspectRatio:false,
      plugins:{legend:{display:false},tooltip:{callbacks:{label:c=>' '+fmt(c.raw)+' tok'}}},
      scales:{
        x:{grid:{color:'rgba(255,255,255,.04)'},ticks:{color:'#9ba3bf',font:{size:10}}},
        y:{grid:{color:'rgba(255,255,255,.04)'},ticks:{color:'#9ba3bf',font:{size:10},callback:vv=>fmt(vv)}}
      }
    }
  });
}
function renderChartDaily(p) {
  const ctx = document.getElementById('c-daily');
  if (!ctx) return;
  if (cDay) { try { cDay.destroy(); } catch(e){} cDay = null; }
  const labels = (p?.daily_labels||[]).map(d=>'D'+d);
  const data   = p?.daily_totals||[];
  try {
    cDay = _makeChartBar(ctx, labels, data);
  } catch(e) {
    setTimeout(function() {
      try {
        if (cDay) { try { cDay.destroy(); } catch(ex){} cDay = null; }
        cDay = _makeChartBar(ctx, labels, data);
      } catch(e2) {}
    }, 250);
  }
}

// ── VISTA MONITORES ───────────────────────────────────────────────────────────
function renderMon(p) {
  if (!p) { document.getElementById('mon-grid').innerHTML='<div class="empty">Sin datos</div>'; return; }
  const monFil = sMon.value;
  const estFil = document.getElementById('s-est').value;
  const byMon  = {};
  (p.modelos||[]).forEach(m => {
    if (monFil && m.monitor !== monFil) return;
    if (estFil) { const bd=bdOf(m.nombre); if(bd&&bd.estado!==estFil) return; }
    (byMon[m.monitor||'Sin monitor'] = byMon[m.monitor||'Sin monitor']||[]).push(m);
  });
  if (!Object.keys(byMon).length) {
    document.getElementById('mon-grid').innerHTML='<div class="empty">Sin datos para el filtro</div>'; return;
  }
  let html='';
  Object.keys(byMon).sort().forEach((mon,mi)=>{
    const color = MC[mi%MC.length];
    const mods  = byMon[mon].sort((a,b)=>b.total-a.total);
    const totMon= mods.reduce((s,m)=>s+m.total,0);
    const cum   = mods.filter(m=>m.reto_cumple).length;
    const ci    = p.mon_com?.[mon];
    const comUSD= ci?.total_usd ?? mods.reduce((s,m)=>s+m.com_usd,0);
    const ini   = mon.split(' ').filter(w=>w.length>2).slice(0,2).map(w=>w[0]).join('');
    html+=`<div class="mon-card">
      <div class="mon-hdr">
        <div class="mon-av" style="background:${color}22;color:${color}">${ini}</div>
        <div style="min-width:0">
          <div class="mon-nm">${mon}</div>
          <div class="mon-meta">${mods.length} modelos · ${cum} ✅ · ${fmt(totMon)} tok</div>
        </div>
        <div class="mon-com" style="flex-shrink:0;margin-left:10px">
          <div class="mon-com-val">${fmtUSD(comUSD)}</div>
          <div class="mon-com-lbl">comisión</div>
        </div>
      </div>
      <div class="mon-body">`;
    mods.forEach(m => {
      const pct2 = m.reto_meta>0?Math.min(m.total/m.reto_meta*100,120):0;
      const fc   = m.reto_cumple?'var(--green)':pct2>=70?'var(--yellow)':'var(--red)';
      const nivel= D.niveles?.[m.nombre]||'';
      const nMap = {'Órbita':'b-orb','Llama':'b-lla','Núcleo':'b-nuc'};
      const nBdg = nivel?`<span class="badge ${nMap[nivel]||''}">${nivel}</span>`:'';
      const cBdg = m.com_usd>=15?'b-c2':m.com_usd>=5?'b-c1':'b-c0';
      const plH  = Object.keys(PLAT).filter(k=>m[k]>0)
        .map(k=>`<span class="pc" style="border-left:2px solid ${PC[k]}">${PLAT[k].slice(0,3)}: <b>${fmt(m[k])}</b></span>`).join('');
      html+=`<div class="mod-row">
        <div class="mod-top">
          <div class="mod-nm">${m.nombre}</div>
          ${nBdg}
          <span class="badge ${m.reto_cumple?'b-ok':'b-no'}">${m.reto_cumple?'✅':'❌'}</span>
          <span class="badge ${cBdg}">${m.com_usd>0?'$'+m.com_usd:'-'}</span>
        </div>
        <div class="mod-bar">
          <div class="mod-bar-t">${fmt(m.total)}</div>
          <div class="mod-prog"><div class="mod-fill" style="width:${pct2}%;background:${fc}"></div></div>
          <div class="mod-pct">${pct2.toFixed(0)}%</div>
        </div>
        ${plH?`<div class="plats">${plH}</div>`:''}
      </div>`;
    });
    html+=`</div></div>`;
  });
  document.getElementById('mon-grid').innerHTML=html;
  // Chart se renderiza AL FINAL para que el grid esté en DOM primero
  renderChartDaily(p);
}

// ── VISTA TABLA ───────────────────────────────────────────────────────────────
function renderTbl(p) {
  const mods=filteredMods(p).sort((a,b)=>b.total-a.total);
  const t=document.getElementById('t-global');
  if(!mods.length){t.innerHTML='<tr><td colspan="11" class="empty">Sin datos</td></tr>';return;}
  const tot=mods.reduce((s,m)=>({f4f:s.f4f+m.f4f,sc:s.sc+m.sc,cb:s.cb+m.cb,cam:s.cam+m.cam,str:s.str+m.str,total:s.total+m.total,com_usd:s.com_usd+m.com_usd}),{f4f:0,sc:0,cb:0,cam:0,str:0,total:0,com_usd:0});
  let h=`<thead><tr><th>#</th><th>Modelo</th><th>Monitor</th><th class="tr">F4F</th><th class="tr">SC</th><th class="tr">CB</th><th class="tr">CAM</th><th class="tr">STR</th><th class="tr">Total</th><th>Reto</th><th class="tr">Com.</th></tr></thead><tbody>`;
  mods.forEach((m,i)=>{
    const rp=m.reto_meta>0?(m.total/m.reto_meta*100).toFixed(0)+'%':'-';
    const cb=m.com_usd>=15?'b-c2':m.com_usd>=5?'b-c1':'b-c0';
    h+=`<tr>
      <td style="color:var(--text3)">${i+1}</td>
      <td style="color:#fff;font-weight:500">${m.nombre}</td>
      <td style="color:var(--text2);max-width:80px;overflow:hidden;text-overflow:ellipsis">${(m.monitor||'').split(' ')[0]}</td>
      <td class="tr" style="color:#6366f1">${m.f4f?fmt(m.f4f):'-'}</td>
      <td class="tr" style="color:#22c55e">${m.sc?fmt(m.sc):'-'}</td>
      <td class="tr" style="color:#f59e0b">${m.cb?fmt(m.cb):'-'}</td>
      <td class="tr" style="color:#3b82f6">${m.cam?fmt(m.cam):'-'}</td>
      <td class="tr" style="color:#ec4899">${m.str?fmt(m.str):'-'}</td>
      <td class="tr tw">${fmt(m.total)}</td>
      <td><span class="badge ${m.reto_cumple?'b-ok':'b-no'}">${m.reto_cumple?'✅':'❌'} ${rp}</span></td>
      <td class="tr"><span class="badge ${cb}">${m.com_usd>0?'$'+m.com_usd:'$0'}</span></td>
    </tr>`;
  });
  h+=`<tr class="tfoot"><td></td><td class="tw">TOTAL</td><td></td>
    <td class="tr tw">${fmt(tot.f4f)}</td><td class="tr tw">${fmt(tot.sc)}</td>
    <td class="tr tw">${fmt(tot.cb)}</td><td class="tr tw">${fmt(tot.cam)}</td>
    <td class="tr tw">${fmt(tot.str)}</td><td class="tr tw">${fmt(tot.total)}</td>
    <td></td><td class="tr tw" style="color:var(--yellow)">${fmtUSD(tot.com_usd)}</td>
  </tr></tbody>`;
  t.innerHTML=h;
}

// ── VISTA PLATAFORMAS ─────────────────────────────────────────────────────────
function renderPlt(p) {
  const mods=filteredMods(p);
  const tots={f4f:0,sc:0,cb:0,cam:0,str:0};
  mods.forEach(m=>{tots.f4f+=m.f4f;tots.sc+=m.sc;tots.cb+=m.cb;tots.cam+=m.cam;tots.str+=m.str});
  const grand=Object.values(tots).reduce((a,b)=>a+b,0);
  document.getElementById('plat-grid').innerHTML=Object.entries(PLAT).map(([k,lbl])=>{
    const vv=tots[k],pct=grand>0?vv/grand*100:0;
    return `<div class="plat-c">
      <div class="plat-nm" style="color:${PC[k]}">${lbl}</div>
      <div class="plat-vl">${fmt(vv)}</div>
      <div class="plat-pt">${fmtPct(pct)}</div>
      <div class="pbar" style="margin-top:8px"><div class="pfill" style="width:${pct}%;background:${PC[k]}"></div></div>
    </div>`;
  }).join('');
  const ctx=document.getElementById('c-plat');
  if(ctx){
    if(cPlt){try{cPlt.destroy();}catch(e){} cPlt=null;}
    try{cPlt=new Chart(ctx,{type:'doughnut',
      data:{labels:Object.values(PLAT),datasets:[{data:Object.keys(PLAT).map(k=>tots[k]),backgroundColor:Object.values(PC),borderWidth:2,borderColor:'#161b27'}]},
      options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{position:'right',labels:{color:'#9ba3bf',font:{size:11},padding:12}},tooltip:{callbacks:{label:c=>' '+fmt(c.raw)+' ('+fmtPct(grand?c.raw/grand*100:0)+')'}}}}});}
    catch(e){setTimeout(function(){
      try{if(cPlt){try{cPlt.destroy();}catch(ex){} cPlt=null;}
        cPlt=new Chart(ctx,{type:'doughnut',
          data:{labels:Object.values(PLAT),datasets:[{data:Object.keys(PLAT).map(k=>tots[k]),backgroundColor:Object.values(PC),borderWidth:2,borderColor:'#161b27'}]},
          options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{position:'right',labels:{color:'#9ba3bf',font:{size:11},padding:12}},tooltip:{callbacks:{label:c=>' '+fmt(c.raw)+' ('+fmtPct(grand?c.raw/grand*100:0)+')'}}}}});
      }catch(e2){}
    },250);}
  }
}

// ── VISTA TOP 20 ──────────────────────────────────────────────────────────────
function setTop(src, el) {
  curTop = src;
  document.querySelectorAll('.sub-btn').forEach(b=>b.classList.remove('on'));
  el.classList.add('on');
  renderTop20();
}

function renderTop20() {
  const mes = sMes.value;
  const qid = sQid.value;
  const src  = curTop === 'fornax' ? D.top20_fornax : D.top20_grupo;
  const list = src?.[mes]?.[qid] || [];
  const div  = document.getElementById('top20-list');
  const srcLabel = curTop === 'fornax' ? 'Cómo vamos Fornax 2' : 'Cómo vamos Grupo Empresarial';

  if (!list.length) {
    div.innerHTML=`<div class="empty">Sin datos para ${mes} ${qid === 'q1' ? 'Q1' : 'Q2'} en ${srcLabel}</div>`; return;
  }
  div.innerHTML = list.slice(0,20).map((r,i)=>{
    const posClass = i===0?'gold':i===1?'silver':i===2?'bronze':'';
    const medal    = i===0?'🥇':i===1?'🥈':i===2?'🥉':'';
    return `<div class="t20-row">
      <div class="t20-pos ${posClass}">${medal||('#'+(i+1))}</div>
      <div style="flex:1;min-width:0">
        <div class="t20-name">${r.modelo}</div>
        <div class="t20-studio">${r.estudio||''}</div>
      </div>
      <div class="t20-prod">${fmt(r.produccion)}</div>
    </div>`;
  }).join('');
}

// ── TABS ──────────────────────────────────────────────────────────────────────
function setTab(tab,el){
  curTab=tab;
  document.querySelectorAll('#main-tabs .tab').forEach(t=>t.classList.remove('on'));
  el.classList.add('on');
  ['mon','tbl','plt','top'].forEach(id=>{
    document.getElementById('v-'+id).style.display = id===tab?'':'none';
  });
  render();
}

// ── CSV ───────────────────────────────────────────────────────────────────────
function exportCSV(){
  const p=getPeriod(); if(!p) return;
  const lines=['Modelo,Monitor,F4F,SC,CB,CAM,STR,Total,Reto Meta,Cumple,Com USD'];
  filteredMods(p).forEach(m=>lines.push(`"${m.nombre}","${m.monitor}",${m.f4f},${m.sc},${m.cb},${m.cam},${m.str},${m.total},${m.reto_meta},${m.reto_cumple?'Sí':'No'},${m.com_usd}`));
  const b=new Blob([lines.join('\n')],{type:'text/csv'});
  const a=document.createElement('a');a.href=URL.createObjectURL(b);
  a.download=`monitores_${sMes.value}_${sQid.value}.csv`;a.click();
}

// ── RENDER ────────────────────────────────────────────────────────────────────
function render(){
  const p=getPeriod();
  renderKPIs(p);
  if(curTab==='mon')      renderMon(p);
  else if(curTab==='tbl') renderTbl(p);
  else if(curTab==='plt') renderPlt(p);
  else if(curTab==='top') renderTop20();
}

render();

window.addEventListener('resize', function() {
  if (cDay) { try { cDay.resize(); } catch(e) {} }
  if (cPlt) { try { cPlt.resize(); } catch(e) {} }
});

} // ── FIN BLOQUE DESKTOP AUTENTICADO ──────────────────────────────────────────
</script>
</body>
</html>"""

def inject_data(html, data):
    return html.replace('DATA_PLACEHOLDER', json.dumps(data, ensure_ascii=False, separators=(',',':')))

if __name__ == "__main__":
    print("="*60)
    print("MONITORES.py v2")
    print(f"Hoy: {HOY_STR} | Corte: {CORTE_STR}")
    print("="*60)
    data = build_data()
    out  = inject_data(HTML, data)
    with open(OUT_HTML, "w", encoding="utf-8") as f:
        f.write(out)
    n_meses = sum(1 for m in data["periodos"].values()
                  if any(q for q in m.values() if q and q.get("produccion",0)>0))
    print(f"\n✅ Dashboard: {OUT_HTML}")
    print(f"   Meses con datos: {n_meses}")
    print(f"   Tamaño: {os.path.getsize(OUT_HTML)/1024:.1f} KB")
