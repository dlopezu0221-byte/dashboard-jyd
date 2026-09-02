#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ELITE.py — Programa J&D ELITE / Modelo de la Semana
====================================================
Fuente de verdad : Cómo vamos Grupo Empresarial.xlsx (Desktop LOCAL)
Salida           : dashboard-jyd/elite/index.html                  (podio del grupo)
                   dashboard-jyd/elite/modelos/<slug>/index.html   (dashboard individual)
                   dashboard-jyd/elite/data/semana_<fecha>.json    (resultado completo)
                   dashboard-jyd/elite/historial.csv               (una fila por semana)

Qué hace
  1. Lee las hojas mensuales del Excel (datos diarios por modelo / estudio / plataforma).
  2. Agrupa por semana lunes–domingo: facturación y días activos.
  3. Calcula el Índice de Desempeño Semanal (IDS) de cada modelo.
  4. Define los 5 títulos con la regla de no acumulación.
  5. Genera las páginas HTML y guarda el histórico.

FÓRMULA VERIFICADA: col_dia(d) = 19 + (nd - d) * 5   [5 plataformas por día]
                    orden de plataformas: Flirt4free, Stripchat, Chaturbate, CamSoda, Streamate
                    Streamate viene en USD  →  se multiplica por STR_FACTOR (igual que MODELOS.py)

Uso
    python ELITE.py                      # última semana completa
    python ELITE.py --semana 2026-08-24  # una semana concreta (lunes)
    python ELITE.py --solo-datos         # no genera HTML, solo JSON + CSV
    python ELITE.py --dry-run            # no escribe nada, imprime el podio

NO hacer commit/push hasta validar el resultado con Gerencia.
"""

import os, re, sys, json, csv, ast, argparse, calendar, unicodedata
from datetime import date, timedelta

try:
    import openpyxl
except ImportError:
    sys.exit("Falta openpyxl.  Instalar con:  pip install openpyxl")

# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURACIÓN  —  todo lo editable está aquí
# ══════════════════════════════════════════════════════════════════════════════

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
GE_XLSX = os.path.join(SCRIPT_DIR, '..', 'Centro de Gestión Estratégica Grupo J&D',
                       'COMO VAMOS GRUPO', 'Cómo vamos Grupo Empresarial.xlsx')
OUT_DIR = os.path.join(SCRIPT_DIR, 'elite')
MODELOS_PY = os.path.join(SCRIPT_DIR, 'MODELOS.py')      # de aquí se leen los PIN

STR_FACTOR = 20.0        # Streamate: USD → créditos (÷ 0,05).  Igual que MODELOS.py
ANIO = 2026

# — Reglas de elegibilidad ————————————————————————————————————————————————
PISO_CREDITOS   = 3000   # facturación mínima de la semana para competir
DIAS_MIN        = 3      # días activos mínimos en la semana
SEMANAS_MIN     = 2      # semanas con actividad de las 4 anteriores
PROMEDIO_MIN    = 500    # promedio histórico mínimo

# — Pesos del índice (deben sumar 100) ————————————————————————————————————
P_RESULTADO     = 35
P_CRECIMIENTO   = 30
P_DIAS          = 12     # parte de Constancia
P_SEMANAS       = 8      # parte de Constancia
P_EVOLUCION     = 15

TOPE_CRECIMIENTO = 2.0   # veces el propio promedio para el puntaje completo
TOPE_DIAS        = 6     # días para el puntaje completo (el 7º no penaliza el descanso)
VENTANA_RECORD   = 12    # semanas hacia atrás para la "mejor semana"
UMBRAL_BREAK     = 1.20  # hay que superar el récord en 20% para el título Breakthrough

MESES = ['ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO',
         'JULIO', 'AGOSTO', 'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE']
MESES_TXT = ['enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio',
             'julio', 'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre']

# ══════════════════════════════════════════════════════════════════════════════
# UTILIDADES
# ══════════════════════════════════════════════════════════════════════════════

def slugify(name):
    """Mismo slug que MODELOS.py, para que las rutas coincidan."""
    s = name.lower()
    for a, b in [('á', 'a'), ('é', 'e'), ('í', 'i'), ('ó', 'o'),
                 ('ú', 'u'), ('ñ', 'n'), ('ü', 'u')]:
        s = s.replace(a, b)
    s = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode('ascii')
    return re.sub(r'[^a-z0-9]+', '-', s).strip('-')


def cargar_pins():
    """Lee el diccionario MODEL_PINS de MODELOS.py sin ejecutar el módulo."""
    if not os.path.exists(MODELOS_PY):
        return {}
    src = open(MODELOS_PY, encoding='utf-8').read()
    m = re.search(r'MODEL_PINS\s*=\s*(\{.*?\n\})', src, re.S)
    if not m:
        return {}
    try:
        return ast.literal_eval(m.group(1))
    except Exception:
        return {}


def lunes_de(d):
    return d - timedelta(days=d.weekday())


def fmt(n, dec=0):
    """Formato colombiano: 39.184 · 2,20"""
    s = f"{n:,.{dec}f}"
    return s.replace(',', '@').replace('.', ',').replace('@', '.')


def esc(t):
    return (str(t).replace('&', '&amp;').replace('<', '&lt;')
            .replace('>', '&gt;').replace('"', '&quot;'))


def rango_txt(lunes):
    dom = lunes + timedelta(days=6)
    if lunes.month == dom.month:
        return f"{lunes.day} al {dom.day} de {MESES_TXT[dom.month-1]} de {dom.year}"
    return (f"{lunes.day} de {MESES_TXT[lunes.month-1]} al "
            f"{dom.day} de {MESES_TXT[dom.month-1]} de {dom.year}")


# ══════════════════════════════════════════════════════════════════════════════
# LECTURA DEL EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def mapa_dias(ws):
    """
    Lee el encabezado de la hoja y devuelve {día: [(col0, factor), ...]}.

    El layout cambió a mitad de año: hasta JUNIO hay 4 plataformas por día y los
    bloques arrancan en la columna P; desde JULIO hay 5 (se agregó Streamate) y
    arrancan en la S. Por eso NO se calcula la columna con una fórmula fija: se
    lee la fila 3 (número de día) y la fila 4 (nombre de plataforma).
    """
    fila3 = [c.value for c in next(ws.iter_rows(min_row=3, max_row=3))]
    fila4 = [c.value for c in next(ws.iter_rows(min_row=4, max_row=4))]

    dias = []   # (día, col0)
    for j, v in enumerate(fila3):
        if isinstance(v, (int, float)) and not isinstance(v, bool) and 1 <= int(v) <= 31:
            dias.append((int(v), j))
    if not dias:
        return {}
    dias.sort(key=lambda t: t[1])

    fin = len(fila4)
    mapa = {}
    for k, (d, ini) in enumerate(dias):
        tope = dias[k + 1][1] if k + 1 < len(dias) else fin
        cols = []
        for j in range(ini, min(tope, fin)):
            nombre = str(fila4[j]).strip().lower() if fila4[j] else ''
            if not nombre:
                continue
            cols.append((j, STR_FACTOR if nombre.startswith('streamate') else 1.0))
        if cols:
            mapa[d] = cols
    return mapa


def leer_diario(path, verbose=True):
    """Devuelve {(modelo, estudio): {date: créditos}} leyendo las hojas mensuales."""
    if not os.path.exists(path):
        sys.exit(f"No se encuentra el archivo maestro:\n  {path}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    datos = {}
    hojas = 0
    for i, mes in enumerate(MESES, start=1):
        if mes not in wb.sheetnames:
            continue
        ws = wb[mes]
        mapa = mapa_dias(ws)
        nd = calendar.monthrange(ANIO, i)[1]
        mapa = {d: c for d, c in mapa.items() if d <= nd}
        if not mapa:
            if verbose:
                print(f"  ! {mes}: no se reconoció el bloque diario, se omite")
            continue
        hojas += 1
        anchos = {len(c) for c in mapa.values()}
        if verbose:
            print(f"  · {mes}: {len(mapa)} días, {max(anchos)} plataformas por día")
        for fila in ws.iter_rows(min_row=5, values_only=True):
            nombre, estudio = fila[0], fila[1]
            if not nombre:
                continue
            nombre = str(nombre).strip()
            if nombre.lower().startswith('gran total'):
                continue
            estudio = str(estudio).strip() if estudio else 'SIN ESTUDIO'
            # filas consolidadas: el "modelo" es el nombre del estudio
            if nombre.lower() == estudio.lower():
                continue
            dd = datos.setdefault((nombre, estudio), {})
            for d, cols in mapa.items():
                total = 0.0
                for j, factor in cols:
                    v = fila[j] if j < len(fila) else None
                    if isinstance(v, (int, float)) and not isinstance(v, bool):
                        total += v * factor
                if total:
                    f = date(ANIO, i, d)
                    dd[f] = dd.get(f, 0.0) + total
    wb.close()
    if not hojas:
        sys.exit("El archivo maestro no tiene ninguna hoja mensual reconocible.")
    return datos


def agrupar_semanas(datos):
    """{(modelo, estudio): {lunes: (créditos, días_activos)}}"""
    sem = {}
    for clave, dias in datos.items():
        acc = {}
        for f, cr in dias.items():
            L = lunes_de(f)
            c, n = acc.get(L, (0.0, 0))
            acc[L] = (c + cr, n + 1)
        sem[clave] = acc
    return sem


def semanas_disponibles(sem, hoy=None):
    """Lunes de todas las semanas COMPLETAS (domingo ya pasó)."""
    hoy = hoy or date.today()
    todas = sorted({L for a in sem.values() for L in a})
    return [L for L in todas if L + timedelta(days=6) < hoy]


# ══════════════════════════════════════════════════════════════════════════════
# CÁLCULO DEL ÍNDICE
# ══════════════════════════════════════════════════════════════════════════════

def evaluar(sem, lunes):
    """Devuelve (compiten, todos) — listas de dicts ordenadas por IDS."""
    base = [lunes - timedelta(weeks=k) for k in range(4, 0, -1)]
    hist = [lunes - timedelta(weeks=k) for k in range(VENTANA_RECORD, 0, -1)]

    filas = []
    for (modelo, estudio), acc in sem.items():
        cr, dias = acc.get(lunes, (0.0, 0))
        base_vals = [acc.get(b, (0.0, 0))[0] for b in base]
        prom4 = sum(base_vals) / 4.0
        sem_act = sum(1 for v in base_vals if v > 0)
        mejor = max([acc.get(h, (0.0, 0))[0] for h in hist] or [0.0])
        if cr <= 0 and prom4 <= 0:
            continue
        filas.append({
            'modelo': modelo, 'estudio': estudio, 'slug': slugify(modelo),
            'cr': round(cr, 2), 'dias': dias, 'prom4': round(prom4, 2),
            'sem_act': sem_act, 'mejor12': round(mejor, 2),
            'base': [round(v, 2) for v in base_vals],
            'crecimiento': round(cr / prom4, 4) if prom4 > 0 else 0.0,
            'pct_record': round(cr / mejor, 4) if mejor > 0 else 1.0,
            'cr_dia': round(cr / dias, 0) if dias else 0.0,
            'nuevos': round(cr - prom4, 2),
        })

    for f in filas:
        f['compite'] = (f['cr'] >= PISO_CREDITOS and f['dias'] >= DIAS_MIN
                        and f['sem_act'] >= SEMANAS_MIN and f['prom4'] >= PROMEDIO_MIN)

    compiten = [f for f in filas if f['compite']]
    n = len(compiten)
    for f in compiten:
        menores = sum(1 for g in compiten if g['cr'] <= f['cr'])   # incluye a sí misma
        f['p_resultado'] = round(P_RESULTADO * (menores - 1) / (n - 1), 2) if n > 1 else P_RESULTADO
        f['p_crecimiento'] = round(P_CRECIMIENTO * min(f['crecimiento'], TOPE_CRECIMIENTO) / TOPE_CRECIMIENTO, 2)
        f['p_constancia'] = round(P_DIAS * min(f['dias'], TOPE_DIAS) / TOPE_DIAS
                                  + P_SEMANAS * f['sem_act'] / 4.0, 2)
        f['p_evolucion'] = round(P_EVOLUCION * min(f['pct_record'], 1.0), 2)
        f['ids'] = round(f['p_resultado'] + f['p_crecimiento']
                         + f['p_constancia'] + f['p_evolucion'], 1)

    compiten.sort(key=lambda f: (-f['ids'], -f['nuevos']))
    for i, f in enumerate(compiten, start=1):
        f['puesto'] = i
    return compiten, filas


def definir_titulos(compiten):
    """Los 5 títulos con la regla de no acumulación."""
    titulos, tomados = [], set()

    def elegir(nombre, criterio, candidatos, dato):
        libres = [f for f in candidatos if f['modelo'] not in tomados]
        if not libres:
            titulos.append({'titulo': nombre, 'modelo': None})
            return
        g = max(libres, key=criterio)
        tomados.add(g['modelo'])
        titulos.append({'titulo': nombre, 'modelo': g['modelo'], 'estudio': g['estudio'],
                        'slug': g['slug'], 'ids': g['ids'], 'dato': dato(g)})

    elegir('MODELO DE LA SEMANA', lambda f: (f['ids'], f['nuevos']), compiten,
           lambda g: f"{fmt(g['ids'],1)} puntos · {fmt(g['cr'])} créditos · "
                     f"{fmt(g['crecimiento'],2)}x · {g['dias']} días")
    elegir('MAYOR CRECIMIENTO', lambda f: f['crecimiento'], compiten,
           lambda g: f"{fmt(g['crecimiento'],2)}x su propio promedio")
    elegir('BREAKTHROUGH', lambda f: f['pct_record'],
           [f for f in compiten if f['pct_record'] >= UMBRAL_BREAK],
           lambda g: f"superó su mejor semana en {fmt((g['pct_record']-1)*100)}%")
    elegir('MAYOR PRODUCTIVIDAD', lambda f: f['cr_dia'], compiten,
           lambda g: f"{fmt(g['cr_dia'])} créditos por día activo")
    elegir('CONSTANCIA', lambda f: f['ids'],
           [f for f in compiten if f['dias'] >= TOPE_DIAS and f['sem_act'] >= 4],
           lambda g: f"{g['dias']} días activos y las 4 semanas del mes")
    return titulos


def resumen(compiten, todos, lunes):
    fact_total = sum(f['cr'] for f in todos)
    return {
        'semana': lunes.isoformat(),
        'rango': rango_txt(lunes),
        'compiten': len(compiten),
        'con_facturacion': sum(1 for f in todos if f['cr'] > 0),
        'superaron_promedio': sum(1 for f in compiten if f['crecimiento'] > 1),
        'rompieron_record': sum(1 for f in compiten if f['pct_record'] > 1),
        'ids_promedio': round(sum(f['ids'] for f in compiten) / len(compiten), 1) if compiten else 0,
        'facturacion_grupo': round(fact_total, 2),
        'alertas': {
            'caida_fuerte': sum(1 for f in todos if f['cr'] > 0 and 0 < f['crecimiento'] < 0.6),
            'sin_dias_minimos': sum(1 for f in todos if f['cr'] >= PISO_CREDITOS and f['dias'] < DIAS_MIN),
            'sin_dias_reportados': sum(1 for f in todos if f['cr'] > 0 and f['dias'] == 0),
            'sin_historial': sum(1 for f in todos if f['cr'] > 0 and f['sem_act'] < SEMANAS_MIN),
            'variacion_sospechosa': sorted(
                [f['modelo'] for f in todos if f['prom4'] > 0 and f['crecimiento'] > 3][:20]),
        },
    }


# ══════════════════════════════════════════════════════════════════════════════
# HTML
# ══════════════════════════════════════════════════════════════════════════════

CSS = """:root{--bg:#0B1120;--card:#141E35;--card2:#1A2747;--border:rgba(255,255,255,0.07);
--gold:#F5B800;--gold-light:#FDD85D;--blue:#3B82F6;--green:#22C55E;--red:#EF4444;
--orange:#F59E0B;--text:#F1F5F9;--muted:#94A3B8;--nav-h:60px}
*{box-sizing:border-box;margin:0;padding:0}
html{scroll-behavior:smooth;color-scheme:dark}
body{font-family:'Segoe UI',system-ui,-apple-system,sans-serif;background:var(--bg);color:var(--text);line-height:1.6;font-size:15px}
nav{position:sticky;top:0;z-index:100;background:rgba(11,17,32,.95);backdrop-filter:blur(12px);
border-bottom:1px solid var(--border);height:var(--nav-h);display:flex;align-items:center;justify-content:space-between;padding:0 24px;gap:14px}
.nav-brand{font-size:13px;color:var(--muted);font-weight:600;letter-spacing:.05em;display:flex;align-items:center;gap:10px;min-width:0}
.brand-mark{width:30px;height:30px;border-radius:8px;background:linear-gradient(135deg,var(--gold),#B8860B);color:#0B1120;
font-size:12px;font-weight:800;display:flex;align-items:center;justify-content:center;flex-shrink:0}
.nav-name{color:#F1F5F9;font-size:12px;letter-spacing:.08em;text-transform:uppercase;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.stamp{font-size:10px;color:#64748B;text-align:right;line-height:1.5;white-space:nowrap}
.stamp b{display:block;font-weight:600;color:#94A3B8}
@media(max-width:560px){.stamp{display:none}}
main{max-width:1100px;margin:0 auto;padding:0 20px 90px}
section{padding:52px 0 0;scroll-margin-top:70px}
.hero{padding:44px 0 0;display:grid;grid-template-columns:1fr auto;gap:26px;align-items:start}
@media(max-width:760px){.hero{grid-template-columns:1fr}}
.hero-badge{display:inline-flex;align-items:center;gap:8px;background:linear-gradient(135deg,rgba(245,184,0,.15),rgba(245,184,0,.05));
border:1px solid rgba(245,184,0,.3);color:var(--gold);font-size:12px;font-weight:700;padding:5px 14px;border-radius:20px;
letter-spacing:.08em;text-transform:uppercase;margin-bottom:16px}
.hero-badge::before{content:'\\2605';font-size:10px}
.hero h1{font-size:clamp(38px,7vw,60px);font-weight:800;line-height:1.05;
background:linear-gradient(135deg,#fff 30%,var(--gold-light));-webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;margin-bottom:6px}
.hero-sub{color:var(--muted);font-size:14px;margin-bottom:18px}
.hero-intro{font-size:16px;color:#CBD5E1;max-width:560px;line-height:1.7;margin-bottom:22px}
.hero-tags{display:flex;flex-wrap:wrap;gap:8px}
.tag{font-size:12px;font-weight:600;padding:4px 12px;border-radius:12px;border:1px solid var(--border)}
.tag-a{background:rgba(59,130,246,.12);border-color:rgba(59,130,246,.3);color:#93C5FD}
.tag-b{background:rgba(34,197,94,.1);border-color:rgba(34,197,94,.25);color:#86EFAC}
.tag-c{background:rgba(245,184,0,.1);border-color:rgba(245,184,0,.28);color:var(--gold-light)}
.hero-aside{background:var(--card);border:1px solid var(--border);border-radius:16px;padding:20px 24px;min-width:230px}
.hero-aside-title{font-size:11px;font-weight:700;color:var(--muted);letter-spacing:.1em;text-transform:uppercase;margin-bottom:14px}
.aside-row{display:flex;justify-content:space-between;align-items:center;gap:14px;padding:8px 0;border-bottom:1px solid var(--border)}
.aside-row:last-child{border-bottom:none}
.aside-label{font-size:12px;color:var(--muted)}
.aside-val{font-size:13px;font-weight:600;text-align:right}
.section-header{margin-bottom:22px}
.section-header h2{font-size:22px;font-weight:700;margin-bottom:4px}
.section-header p{color:var(--muted);font-size:14px;max-width:640px}
.section-divider{width:40px;height:3px;border-radius:2px;background:linear-gradient(90deg,var(--gold),transparent);margin:10px 0 4px}
.kpi-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:14px}
.kpi-card{background:var(--card);border:1px solid var(--border);border-radius:16px;padding:20px;transition:transform .2s,border-color .2s}
.kpi-card:hover{transform:translateY(-2px);border-color:rgba(255,255,255,.15)}
.kpi-label{font-size:11px;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.08em;margin-bottom:10px}
.kpi-value{font-size:28px;font-weight:800;line-height:1;margin-bottom:4px}
.kpi-sub{font-size:12px;color:var(--muted)}
.kpi-badge{display:inline-block;font-size:11px;font-weight:700;padding:2px 8px;border-radius:8px;margin-top:8px}
.badge-up{background:rgba(34,197,94,.15);color:var(--green)}
.badge-warn{background:rgba(245,158,11,.15);color:var(--orange)}
.card{background:var(--card);border:1px solid var(--border);border-radius:16px;padding:24px}
.comp{display:flex;flex-direction:column;gap:18px}
.comp-row{display:grid;grid-template-columns:170px 1fr 74px;gap:16px;align-items:center}
@media(max-width:680px){.comp-row{grid-template-columns:1fr;gap:6px}}
.comp-name{font-size:13px;font-weight:700;letter-spacing:.04em;text-transform:uppercase}
.comp-desc{font-size:11px;color:var(--muted);font-weight:400;text-transform:none;letter-spacing:0;margin-top:2px}
.bar{height:12px;border-radius:8px;background:rgba(255,255,255,.06);overflow:hidden}
.bar span{display:block;height:100%;border-radius:8px}
.comp-pts{font-size:15px;font-weight:800;text-align:right;white-space:nowrap}
@media(max-width:680px){.comp-pts{text-align:left}}
.comp-pts small{font-size:11px;color:var(--muted);font-weight:600}
.total-row{display:flex;justify-content:space-between;align-items:baseline;margin-top:24px;padding-top:20px;border-top:1px solid var(--border)}
.total-row .lbl{font-size:13px;font-weight:700;letter-spacing:.06em;text-transform:uppercase;color:var(--muted)}
.total-row .val{font-size:34px;font-weight:800;color:var(--gold)}
.sim-grid{display:grid;grid-template-columns:1fr 1fr;gap:20px}
@media(max-width:760px){.sim-grid{grid-template-columns:1fr}}
label.fld{display:block;font-size:11px;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.08em;margin-bottom:8px}
input[type=number]{width:100%;background:#0F172A;border:1px solid #334155;border-radius:10px;color:var(--text);
font-size:18px;font-weight:700;padding:12px 14px;outline:none;font-family:inherit}
input[type=number]:focus{border-color:var(--gold);box-shadow:0 0 0 3px rgba(245,184,0,.12)}
input[type=range]{width:100%;accent-color:var(--gold)}
.sim-out{background:var(--card2);border:1px solid var(--border);border-radius:14px;padding:20px;display:flex;flex-direction:column;gap:14px}
.sim-big{font-size:44px;font-weight:800;line-height:1;color:var(--gold)}
.sim-pos{font-size:14px;color:var(--muted)}
.sim-pos b{color:var(--text);font-size:17px}
.sim-mini{display:grid;grid-template-columns:repeat(4,1fr);gap:8px}
.sim-mini div{background:rgba(255,255,255,.04);border-radius:10px;padding:8px;text-align:center}
.sim-mini .m-l{font-size:9px;color:var(--muted);text-transform:uppercase;letter-spacing:.06em}
.sim-mini .m-v{font-size:15px;font-weight:800}
.hint{font-size:12px;color:#64748B;margin-top:14px;line-height:1.6}
.horas{display:grid;grid-template-columns:repeat(7,1fr);gap:10px}
@media(max-width:680px){.horas{grid-template-columns:repeat(4,1fr)}}
.hd{background:#0F172A;border:1px solid #334155;border-radius:12px;padding:10px 8px;text-align:center}
.hd label{display:block;font-size:10px;color:var(--muted);font-weight:700;text-transform:uppercase;letter-spacing:.06em;margin-bottom:6px}
.hd input{width:100%;background:transparent;border:none;color:var(--text);font-size:19px;font-weight:800;text-align:center;outline:none;font-family:inherit}
.btn{background:var(--gold);color:#0B1120;border:none;border-radius:10px;padding:12px 22px;font-size:14px;font-weight:700;cursor:pointer;font-family:inherit}
.btn:hover{filter:brightness(1.08)}
.saved{font-size:12px;font-weight:700;color:var(--green);opacity:0;transition:opacity .3s}
.saved.on{opacity:1}
.tit-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(250px,1fr));gap:14px}
.tit{background:var(--card);border:1px solid var(--border);border-radius:16px;padding:20px}
.tit.main{border-color:rgba(245,184,0,.35);background:linear-gradient(160deg,rgba(245,184,0,.09),var(--card) 60%)}
.tit-l{font-size:11px;font-weight:700;letter-spacing:.09em;text-transform:uppercase;color:var(--muted);margin-bottom:8px}
.tit.main .tit-l{color:var(--gold)}
.tit-m{font-size:22px;font-weight:800;line-height:1.15;margin-bottom:2px}
.tit-e{font-size:12px;color:var(--muted);margin-bottom:8px}
.tit-d{font-size:12.5px;color:#CBD5E1}
.tbl-wrap{overflow-x:auto;border-radius:16px;border:1px solid var(--border)}
table{width:100%;border-collapse:collapse;min-width:640px;background:var(--card)}
th{background:#101A2E;font-size:11px;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.07em;padding:12px 14px;text-align:left;white-space:nowrap}
td{font-size:13px;padding:12px 14px;border-top:1px solid var(--border);white-space:nowrap}
tr.me td{background:rgba(245,184,0,.07)}
tr.me td:first-child{box-shadow:inset 3px 0 0 var(--gold)}
.num{text-align:right;font-variant-numeric:tabular-nums}
.pill{display:inline-block;font-size:11px;font-weight:700;padding:2px 9px;border-radius:8px;background:rgba(245,184,0,.14);color:var(--gold-light)}
footer{max-width:1100px;margin:0 auto;padding:26px 20px 60px;border-top:1px solid var(--border);color:#64748B;font-size:12px;line-height:1.7}"""

PIN_JS = """<script>
(function(){var K='%(key)s',P='ok_%(pin)s',A=false;
try{A=sessionStorage.getItem(K)===P;}catch(e){}
if(A)return;document.body.style.display='none';
var o=document.createElement('div');o.id='_pin_overlay';
o.style.cssText='position:fixed;top:0;left:0;width:100%%;height:100%%;background:#0F172A;display:flex;align-items:center;justify-content:center;z-index:99999;font-family:Arial,sans-serif';
o.innerHTML=['<div style="background:#1E293B;border:1px solid #334155;border-radius:16px;padding:40px 32px;width:320px;max-width:90vw;text-align:center;box-shadow:0 25px 50px rgba(0,0,0,.6)">',
'<div style="font-size:32px;margin-bottom:12px">\\uD83D\\uDD10</div>',
'<div style="color:#F1F5F9;font-size:18px;font-weight:700;margin-bottom:6px">Dashboard Privado</div>',
'<div style="color:#64748B;font-size:13px;margin-bottom:24px">Ingresa tu c\\u00F3digo de acceso</div>',
'<input id="_pin_inp" type="password" inputmode="numeric" pattern="[0-9]*" maxlength="6" placeholder="\\u2022 \\u2022 \\u2022 \\u2022 \\u2022 \\u2022" style="width:100%%;box-sizing:border-box;background:#0F172A;border:1px solid #334155;border-radius:8px;color:#F1F5F9;font-size:22px;letter-spacing:6px;padding:12px;text-align:center;outline:none;margin-bottom:12px">',
'<div id="_pin_err" style="display:none;color:#EF4444;font-size:12px;margin-bottom:10px">C\\u00F3digo incorrecto. Intenta de nuevo.</div>',
'<button id="_pin_btn" style="width:100%%;background:#3B82F6;color:#fff;border:none;border-radius:8px;padding:13px;font-size:15px;font-weight:600;cursor:pointer">Ingresar</button>',
'<div style="color:#475569;font-size:11px;margin-top:20px">Grupo Empresarial J&D \\u00B7 Acceso restringido</div>','</div>'].join('');
document.body.parentNode.insertBefore(o,document.body);
function c(){var v=(document.getElementById('_pin_inp').value||'').trim();
if(v==='%(pin)s'){try{sessionStorage.setItem(K,P);}catch(e){}o.remove();document.body.style.display='';}
else{document.getElementById('_pin_err').style.display='block';document.getElementById('_pin_inp').value='';}}
document.getElementById('_pin_btn').addEventListener('click',c);
document.getElementById('_pin_inp').addEventListener('keydown',function(e){if(e.key==='Enter')c();});})();
</script>"""


def _pagina(titulo, nav_name, stamp, cuerpo, pin=None, extra_js=''):
    pin_js = (PIN_JS % {'pin': pin[0], 'key': pin[1]}) if pin else ''
    return f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1.0"/>
<title>{esc(titulo)}</title>
<style>{CSS}</style>
</head>
<body>
{pin_js}
<nav>
  <div class="nav-brand">
    <div class="brand-mark">J&amp;D</div>
    <span class="nav-name">{esc(nav_name)}</span>
  </div>
  <div class="stamp">{stamp}</div>
</nav>
<main>
{cuerpo}
</main>
<footer>
<b style="color:#94A3B8">Grupo Empresarial J&amp;D &middot; Programa J&amp;D Elite</b><br>
La facturaci&oacute;n y los d&iacute;as activos se toman del reporte oficial del grupo, con corte los lunes a las 12:00 del mediod&iacute;a.
El c&aacute;lculo completo se publica cada semana para que cualquiera pueda verificarlo.
</footer>
{extra_js}
</body>
</html>"""


def html_grupo(compiten, titulos, res, lunes):
    tit_html = []
    for i, t in enumerate(titulos):
        if not t['modelo']:
            tit_html.append(f"""<div class="tit"><div class="tit-l">{esc(t['titulo'])}</div>
<div class="tit-m" style="color:#64748B;font-size:17px">No se otorga esta semana</div>
<div class="tit-d" style="color:#64748B">Nadie cumpli&oacute; la condici&oacute;n</div></div>""")
            continue
        tit_html.append(f"""<div class="tit{' main' if i == 0 else ''}">
<div class="tit-l">{esc(t['titulo'])}</div>
<div class="tit-m">{esc(t['modelo'])}</div>
<div class="tit-e">{esc(t['estudio'])}</div>
<div class="tit-d">{esc(t['dato'])}</div></div>""")

    filas = []
    for f in compiten[:15]:
        marca = ' class="me"' if f['puesto'] == 1 else ''
        p = f'<span class="pill">1</span>' if f['puesto'] == 1 else str(f['puesto'])
        filas.append(
            f"<tr{marca}><td>{p}</td><td style='font-weight:600'>{esc(f['modelo'])}</td>"
            f"<td style='color:#94A3B8'>{esc(f['estudio'])}</td>"
            f"<td class='num'>{fmt(f['crecimiento'],2)}x</td><td class='num'>{f['dias']}</td>"
            f"<td class='num'>{fmt(f['p_resultado'],1)} &middot; {fmt(f['p_crecimiento'],1)} &middot; "
            f"{fmt(f['p_constancia'],1)} &middot; {fmt(f['p_evolucion'],1)}</td>"
            f"<td class='num' style='font-weight:800;color:{'#F5B800' if f['puesto']==1 else '#F1F5F9'}'>{fmt(f['ids'],1)}</td></tr>")

    g = titulos[0]
    cuerpo = f"""
<div class="hero">
  <div>
    <div class="hero-badge">J&amp;D Elite &middot; {esc(res['rango'])}</div>
    <h1>{esc(g['modelo']) if g['modelo'] else 'Semana sin t&iacute;tulo'}</h1>
    <div class="hero-sub">{esc(g['estudio']) if g['modelo'] else 'Ning&uacute;n modelo cumpli&oacute; las condiciones'}</div>
    <p class="hero-intro">{esc(g['dato']) if g['modelo'] else ''}</p>
    <div class="hero-tags">
      <span class="tag tag-c">{res['compiten']} modelos compitieron</span>
      <span class="tag tag-b">{res['superaron_promedio']} superaron su promedio</span>
      <span class="tag tag-a">{res['rompieron_record']} rompieron su r&eacute;cord</span>
    </div>
  </div>
  <aside class="hero-aside">
    <div class="hero-aside-title">La semana en cifras</div>
    <div class="aside-row"><span class="aside-label">Facturaci&oacute;n del grupo</span><span class="aside-val">{fmt(res['facturacion_grupo'])}</span></div>
    <div class="aside-row"><span class="aside-label">Modelos con facturaci&oacute;n</span><span class="aside-val">{res['con_facturacion']}</span></div>
    <div class="aside-row"><span class="aside-label">Compitieron</span><span class="aside-val">{res['compiten']}</span></div>
    <div class="aside-row"><span class="aside-label">&Iacute;ndice promedio</span><span class="aside-val">{fmt(res['ids_promedio'],1)}</span></div>
  </aside>
</div>

<section>
  <div class="section-header"><h2>Los t&iacute;tulos de la semana</h2><div class="section-divider"></div>
  <p>Ning&uacute;n modelo puede llevarse dos t&iacute;tulos la misma semana.</p></div>
  <div class="tit-grid">{''.join(tit_html)}</div>
</section>

<section>
  <div class="section-header"><h2>Top 15</h2><div class="section-divider"></div>
  <p>Se publica el puesto y el puntaje. Las cifras individuales de facturaci&oacute;n solo las ve cada modelo con su monitor.</p></div>
  <div class="tbl-wrap"><table>
    <thead><tr><th>#</th><th>Modelo</th><th>Estudio</th><th class="num">Crecim.</th><th class="num">D&iacute;as</th>
    <th class="num">RES &middot; CRE &middot; CON &middot; EVO</th><th class="num">&Iacute;ndice</th></tr></thead>
    <tbody>{''.join(filas)}</tbody>
  </table></div>
</section>

<section>
  <div class="section-header"><h2>Control de datos</h2><div class="section-divider"></div>
  <p>Revisar antes de anunciar. Estas alertas no bloquean el c&aacute;lculo, se&ntilde;alan qu&eacute; verificar.</p></div>
  <div class="kpi-grid">
    <div class="kpi-card"><div class="kpi-label">Cayeron m&aacute;s de 40%</div><div class="kpi-value" style="color:var(--orange)">{res['alertas']['caida_fuerte']}</div><div class="kpi-sub">seguimiento del monitor</div></div>
    <div class="kpi-card"><div class="kpi-label">Sin d&iacute;as m&iacute;nimos</div><div class="kpi-value">{res['alertas']['sin_dias_minimos']}</div><div class="kpi-sub">facturaron pero no compiten</div></div>
    <div class="kpi-card"><div class="kpi-label">Sin d&iacute;as reportados</div><div class="kpi-value" style="color:var(--red)">{res['alertas']['sin_dias_reportados']}</div><div class="kpi-sub">estudio que carga consolidado</div></div>
    <div class="kpi-card"><div class="kpi-label">Sin historial suficiente</div><div class="kpi-value">{res['alertas']['sin_historial']}</div><div class="kpi-sub">modelos nuevos</div></div>
  </div>
</section>"""
    stamp = (f"<b>Semana {esc(res['rango'])}</b><span>Generado el {date.today().strftime('%d/%m/%Y')}</span>")
    return _pagina('J&D Elite — Modelo de la Semana', 'Podio del grupo', stamp, cuerpo)


def html_modelo(f, compiten, res, lunes, pin):
    otros = [{'f': g['cr'], 'ids': g['ids']} for g in compiten if g['modelo'] != f['modelo']]
    top = [[g['modelo'], g['estudio'], g['crecimiento'], g['dias'], g['ids'], g['puesto']]
           for g in compiten[:10]]
    n = len(compiten)
    falta = round(P_RESULTADO - f['p_resultado'] + P_CRECIMIENTO - f['p_crecimiento']
                  + (P_DIAS + P_SEMANAS) - f['p_constancia'] + P_EVOLUCION - f['p_evolucion'], 1)

    def barra(pts, maxp, color):
        w = 0 if maxp == 0 else max(2, round(100 * pts / maxp))
        return f'<div class="bar"><span style="width:{w}%;background:{color}"></span></div>'

    tags = []
    if f['puesto'] == 1:
        tags.append('<span class="tag tag-c">Modelo de la Semana</span>')
    if f['pct_record'] >= 1:
        tags.append('<span class="tag tag-b">R&eacute;cord personal superado</span>')
    tags.append(f'<span class="tag tag-a">{f["dias"]} d&iacute;as activos</span>')

    filas = []
    for m, e, c, d, ids, pu in top:
        marca = ' class="me"' if m == f['modelo'] else ''
        filas.append(f"<tr{marca}><td>{pu}</td><td style='font-weight:600'>{esc(m)}</td>"
                     f"<td style='color:#94A3B8'>{esc(e)}</td><td class='num'>{fmt(c,2)}x</td>"
                     f"<td class='num'>{d}</td><td class='num' style='font-weight:800'>{fmt(ids,1)}</td></tr>")

    cuerpo = f"""
<div class="hero">
  <div>
    <div class="hero-badge">J&amp;D Elite &middot; {esc(res['rango'])}</div>
    <h1>{fmt(f['ids'],1)} puntos</h1>
    <div class="hero-sub">Puesto {f['puesto']} de {n} modelos que compitieron esta semana</div>
    <p class="hero-intro">Tu facturaci&oacute;n y tus d&iacute;as se cargan solos desde el reporte del grupo &mdash; aqu&iacute; no se digitan.
    Lo que s&iacute; puedes hacer: registrar tus horas y probar en el simulador qu&eacute; necesitas la pr&oacute;xima semana.</p>
    <div class="hero-tags">{''.join(tags)}</div>
  </div>
  <aside class="hero-aside">
    <div class="hero-aside-title">Mi ficha</div>
    <div class="aside-row"><span class="aside-label">Estudio</span><span class="aside-val">{esc(f['estudio'])}</span></div>
    <div class="aside-row"><span class="aside-label">Mi promedio (4 sem)</span><span class="aside-val">{fmt(f['prom4'])}</span></div>
    <div class="aside-row"><span class="aside-label">Mi mejor semana</span><span class="aside-val">{fmt(max(f['mejor12'], f['cr']))}</span></div>
    <div class="aside-row"><span class="aside-label">Cr&eacute;ditos nuevos</span><span class="aside-val">{fmt(f['nuevos'])}</span></div>
  </aside>
</div>

<section>
  <div class="section-header"><h2>Mi semana</h2><div class="section-divider"></div>
  <p>Estos cuatro datos vienen del reporte oficial del grupo. No se pueden editar aqu&iacute;: as&iacute; el ranking es igual para todas.</p></div>
  <div class="kpi-grid">
    <div class="kpi-card"><div class="kpi-label">Mi puesto</div>
      <div class="kpi-value" style="color:var(--gold)">{f['puesto']}<span style="font-size:16px;color:var(--muted);font-weight:600"> / {n}</span></div>
      <div class="kpi-sub">de las que compitieron</div></div>
    <div class="kpi-card"><div class="kpi-label">Facturaci&oacute;n</div>
      <div class="kpi-value">{fmt(f['cr'])}</div><div class="kpi-sub">cr&eacute;ditos esta semana</div></div>
    <div class="kpi-card"><div class="kpi-label">D&iacute;as activos</div>
      <div class="kpi-value">{f['dias']}<span style="font-size:16px;color:var(--muted);font-weight:600"> / 7</span></div>
      <div class="kpi-sub">{'semana completa' if f['dias'] >= 7 else 'de la semana'}</div></div>
    <div class="kpi-card"><div class="kpi-label">Crecimiento</div>
      <div class="kpi-value">{fmt(f['crecimiento'],2)}x</div>
      <div class="kpi-sub">frente a tu promedio de {fmt(f['prom4'])}</div></div>
  </div>
</section>

<section>
  <div class="section-header"><h2>Mis 100 puntos</h2><div class="section-divider"></div>
  <p>Tres de los cuatro bloques dependen solo de ti: {P_CRECIMIENTO + P_DIAS + P_SEMANAS + P_EVOLUCION} puntos que no tienen nada que ver con lo que facturen las dem&aacute;s.</p></div>
  <div class="card">
    <div class="comp">
      <div class="comp-row"><div><div class="comp-name">Resultado</div><div class="comp-desc">Tu posici&oacute;n por facturaci&oacute;n entre las {n} que compiten</div></div>
        {barra(f['p_resultado'], P_RESULTADO, 'linear-gradient(90deg,#3B82F6,#60A5FA)')}
        <div class="comp-pts">{fmt(f['p_resultado'],1)} <small>/ {P_RESULTADO}</small></div></div>
      <div class="comp-row"><div><div class="comp-name">Crecimiento</div><div class="comp-desc">Tu semana &divide; tu promedio. Duplicarlo da el m&aacute;ximo</div></div>
        {barra(f['p_crecimiento'], P_CRECIMIENTO, 'linear-gradient(90deg,#22C55E,#4ADE80)')}
        <div class="comp-pts">{fmt(f['p_crecimiento'],1)} <small>/ {P_CRECIMIENTO}</small></div></div>
      <div class="comp-row"><div><div class="comp-name">Constancia</div><div class="comp-desc">D&iacute;as activos de la semana + semanas activas del mes</div></div>
        {barra(f['p_constancia'], P_DIAS + P_SEMANAS, 'linear-gradient(90deg,#F59E0B,#FBBF24)')}
        <div class="comp-pts">{fmt(f['p_constancia'],1)} <small>/ {P_DIAS + P_SEMANAS}</small></div></div>
      <div class="comp-row"><div><div class="comp-name">Evoluci&oacute;n</div><div class="comp-desc">Qu&eacute; tan cerca quedaste de tu mejor semana</div></div>
        {barra(f['p_evolucion'], P_EVOLUCION, 'linear-gradient(90deg,#F5B800,#FDD85D)')}
        <div class="comp-pts">{fmt(f['p_evolucion'],1)} <small>/ {P_EVOLUCION}</small></div></div>
    </div>
    <div class="total-row"><span class="lbl">Mi &iacute;ndice de la semana</span><span class="val">{fmt(f['ids'],1)}</span></div>
    <p class="hint"><b style="color:#CBD5E1">Lo que te falt&oacute;:</b> {fmt(falta,1)} puntos para el 100.</p>
  </div>
</section>

<section>
  <div class="section-header"><h2>Simulador &middot; &iquest;qu&eacute; necesito la pr&oacute;xima semana?</h2><div class="section-divider"></div>
  <p>Escribe una cifra y mira qu&eacute; puntaje tendr&iacute;as y en qu&eacute; puesto quedar&iacute;as. Esto es solo tuyo: no cambia tu resultado real ni lo ve nadie m&aacute;s.</p></div>
  <div class="card">
    <div class="sim-grid">
      <div>
        <label class="fld" for="simF">Si esta semana facturo&hellip;</label>
        <input type="number" id="simF" value="{int(f['cr'])}" min="0" step="500">
        <div style="margin-top:20px">
          <label class="fld" for="simD">&hellip;trabajando <span id="simDTxt" style="color:var(--gold)">{f['dias']}</span> d&iacute;as</label>
          <input type="range" id="simD" min="0" max="7" step="1" value="{f['dias']}">
        </div>
        <p class="hint">Tu promedio de las &uacute;ltimas 4 semanas es <b style="color:#CBD5E1">{fmt(f['prom4'])}</b> y tu r&eacute;cord es
        <b style="color:#CBD5E1">{fmt(max(f['mejor12'], f['cr']))}</b>. Con esos dos n&uacute;meros se calculan Crecimiento y Evoluci&oacute;n.</p>
      </div>
      <div class="sim-out">
        <div><div style="font-size:11px;font-weight:700;color:var(--muted);text-transform:uppercase;letter-spacing:.08em">Tu puntaje ser&iacute;a</div>
          <div class="sim-big" id="simIDS">{fmt(f['ids'],1)}</div></div>
        <div class="sim-pos">Quedar&iacute;as en el puesto <b id="simPos">{f['puesto']}</b> de {n}</div>
        <div class="sim-mini">
          <div><div class="m-l">Result.</div><div class="m-v" id="mRes">{fmt(f['p_resultado'],1)}</div></div>
          <div><div class="m-l">Crecim.</div><div class="m-v" id="mCre">{fmt(f['p_crecimiento'],1)}</div></div>
          <div><div class="m-l">Constan.</div><div class="m-v" id="mCon">{fmt(f['p_constancia'],1)}</div></div>
          <div><div class="m-l">Evol.</div><div class="m-v" id="mEvo">{fmt(f['p_evolucion'],1)}</div></div>
        </div>
        <div id="simMsg" style="font-size:12.5px;color:#93C5FD;line-height:1.6"></div>
      </div>
    </div>
  </div>
</section>

<section>
  <div class="section-header"><h2>Mis horas de conexi&oacute;n</h2><div class="section-divider"></div>
  <p>Este es el &uacute;nico dato que el grupo todav&iacute;a no tiene y que t&uacute; puedes aportar. Por ahora no suma puntos: se est&aacute; midiendo para incorporarlo m&aacute;s adelante.</p></div>
  <div class="card">
    <div class="horas">
      {''.join(f'<div class="hd"><label>{d}</label><input type="number" min="0" max="24" step="0.5" data-h value=""></div>' for d in ['Lun','Mar','Mi&eacute;','Jue','Vie','S&aacute;b','Dom'])}
    </div>
    <div style="display:flex;flex-wrap:wrap;align-items:center;gap:16px;margin-top:22px">
      <button class="btn" id="btnH">Guardar mis horas</button>
      <span class="saved" id="okH">&#10003; Guardado</span>
      <div style="margin-left:auto;display:flex;gap:26px">
        <div><div style="font-size:10px;color:var(--muted);text-transform:uppercase;letter-spacing:.08em;font-weight:700">Total semana</div><div style="font-size:22px;font-weight:800" id="hTot">&mdash;</div></div>
        <div><div style="font-size:10px;color:var(--muted);text-transform:uppercase;letter-spacing:.08em;font-weight:700">Cr&eacute;ditos por hora</div><div style="font-size:22px;font-weight:800;color:var(--gold)" id="hCph">&mdash;</div></div>
      </div>
    </div>
    <p class="hint">Cuando todas las modelos del grupo registren sus horas durante un trimestre completo, esta cifra se convierte en el quinto componente del &iacute;ndice.</p>
  </div>
</section>

<section>
  <div class="section-header"><h2>Top 10 de la semana</h2><div class="section-divider"></div>
  <p>Se publica el puesto y el puntaje de todas las que compiten.</p></div>
  <div class="tbl-wrap"><table>
    <thead><tr><th>#</th><th>Modelo</th><th>Estudio</th><th class="num">Crecim.</th><th class="num">D&iacute;as</th><th class="num">&Iacute;ndice</th></tr></thead>
    <tbody>{''.join(filas)}</tbody>
  </table></div>
</section>"""

    js = f"""<script>
(function(){{
var PROM={f['prom4']}, RECORD={max(f['mejor12'], f['cr'])}, SEM_ACT={f['sem_act']}, N={n};
var OTROS={json.dumps(otros)};
var PISO={PISO_CREDITOS}, DMIN={DIAS_MIN};
function nf(n,d){{return n.toLocaleString('es-CO',{{minimumFractionDigits:d||0,maximumFractionDigits:d||0}});}}
var sF=document.getElementById('simF'), sD=document.getElementById('simD');
function sim(){{
  var f=Math.max(0,parseFloat(sF.value)||0), d=parseInt(sD.value,10);
  document.getElementById('simDTxt').textContent=d;
  var ok=(f>=PISO&&d>=DMIN), men=0;
  for(var i=0;i<OTROS.length;i++){{if(OTROS[i].f<=f)men++;}}
  var res=ok?Math.min({P_RESULTADO},{P_RESULTADO}*men/(N-1)):0;
  var cre=ok?{P_CRECIMIENTO}*Math.min(PROM>0?f/PROM:0,{TOPE_CRECIMIENTO})/{TOPE_CRECIMIENTO}:0;
  var con=ok?{P_DIAS}*Math.min(d,{TOPE_DIAS})/{TOPE_DIAS}+{P_SEMANAS}*SEM_ACT/4:0;
  var evo=ok?{P_EVOLUCION}*Math.min(RECORD>0?f/RECORD:1,1):0;
  var ids=res+cre+con+evo, pos=1;
  for(var j=0;j<OTROS.length;j++){{if(OTROS[j].ids>ids)pos++;}}
  document.getElementById('mRes').textContent=nf(res,1);
  document.getElementById('mCre').textContent=nf(cre,1);
  document.getElementById('mCon').textContent=nf(con,1);
  document.getElementById('mEvo').textContent=nf(evo,1);
  document.getElementById('simIDS').textContent=ok?nf(ids,1):'\\u2014';
  document.getElementById('simPos').textContent=ok?pos:'\\u2014';
  var m='';
  if(!ok){{m=f<PISO?'Con menos de '+nf(PISO)+' cr\\u00e9ditos no entras al ranking de esa semana.':'Con menos de '+DMIN+' d\\u00edas activos no entras al ranking de esa semana.';}}
  else if(pos===1){{m='Con esa semana ser\\u00edas la Modelo de la Semana.';}}
  else if(pos<=3){{m='Con esa semana estar\\u00edas en el podio.';}}
  else if(pos<=10){{m='Con esa semana entrar\\u00edas al Top 10 que se publica el lunes.';}}
  else {{m='Te faltar\\u00eda un poco m\\u00e1s para entrar al Top 10.';}}
  document.getElementById('simMsg').textContent=m;
}}
sF.addEventListener('input',sim); sD.addEventListener('input',sim); sim();
var hs=Array.prototype.slice.call(document.querySelectorAll('[data-h]'));
var KEY='jd_elite_horas_{f['slug']}_{lunes.isoformat()}';
function horas(){{
  var t=0,any=false; hs.forEach(function(i){{var v=parseFloat(i.value); if(!isNaN(v)){{t+=v;any=true;}}}});
  document.getElementById('hTot').textContent=any?nf(t,1)+' h':'\\u2014';
  document.getElementById('hCph').textContent=(any&&t>0)?nf({f['cr']}/t,0):'\\u2014';
}}
hs.forEach(function(i){{i.addEventListener('input',horas);}});
try{{var g=localStorage.getItem(KEY); if(g){{var v=JSON.parse(g); if(Array.isArray(v))hs.forEach(function(i,k){{if(v[k]!=null)i.value=v[k];}});}}}}catch(e){{}}
horas();
document.getElementById('btnH').addEventListener('click',function(){{
  try{{localStorage.setItem(KEY,JSON.stringify(hs.map(function(i){{return i.value;}})));}}catch(e){{}}
  var o=document.getElementById('okH'); o.classList.add('on'); setTimeout(function(){{o.classList.remove('on');}},2200);
}});
}})();
</script>"""
    stamp = f"<b>Semana {esc(res['rango'])}</b><span>Generado el {date.today().strftime('%d/%m/%Y')}</span>"
    return _pagina(f"Mi J&D Elite — {f['modelo']}", f"{f['modelo']} · {f['estudio']}", stamp, cuerpo, pin, js)


# ══════════════════════════════════════════════════════════════════════════════
# SALIDAS
# ══════════════════════════════════════════════════════════════════════════════

def guardar(compiten, todos, titulos, res, lunes, pins, solo_datos=False):
    os.makedirs(os.path.join(OUT_DIR, 'data'), exist_ok=True)

    ruta_json = os.path.join(OUT_DIR, 'data', f"semana_{lunes.isoformat()}.json")
    with open(ruta_json, 'w', encoding='utf-8') as fh:
        json.dump({'resumen': res, 'titulos': titulos, 'ranking': compiten},
                  fh, ensure_ascii=False, indent=1)
    print(f"  · {os.path.relpath(ruta_json, SCRIPT_DIR)}")

    ruta_csv = os.path.join(OUT_DIR, 'historial.csv')
    nuevo = not os.path.exists(ruta_csv)
    cols = ['semana', 'rango', 'modelo_semana', 'estudio', 'ids', 'facturacion', 'crecimiento',
            'dias', 'mayor_crecimiento', 'breakthrough', 'mayor_productividad', 'constancia',
            'compiten', 'facturacion_grupo']
    prev = []
    if not nuevo:
        with open(ruta_csv, encoding='utf-8') as fh:
            prev = [r for r in csv.DictReader(fh) if r.get('semana') != lunes.isoformat()]
    g = titulos[0]
    fila = {
        'semana': lunes.isoformat(), 'rango': res['rango'],
        'modelo_semana': g['modelo'] or '', 'estudio': g.get('estudio', ''),
        'ids': g.get('ids', ''), 'facturacion': '', 'crecimiento': '', 'dias': '',
        'mayor_crecimiento': titulos[1]['modelo'] or '', 'breakthrough': titulos[2]['modelo'] or '',
        'mayor_productividad': titulos[3]['modelo'] or '', 'constancia': titulos[4]['modelo'] or '',
        'compiten': res['compiten'], 'facturacion_grupo': round(res['facturacion_grupo']),
    }
    if compiten:
        w = compiten[0]
        fila.update({'facturacion': round(w['cr']), 'crecimiento': round(w['crecimiento'], 2),
                     'dias': w['dias']})
    with open(ruta_csv, 'w', encoding='utf-8', newline='') as fh:
        wr = csv.DictWriter(fh, fieldnames=cols)
        wr.writeheader()
        for r in sorted(prev + [fila], key=lambda r: r['semana']):
            wr.writerow({k: r.get(k, '') for k in cols})
    print(f"  · {os.path.relpath(ruta_csv, SCRIPT_DIR)}")

    if solo_datos:
        return

    with open(os.path.join(OUT_DIR, 'index.html'), 'w', encoding='utf-8') as fh:
        fh.write(html_grupo(compiten, titulos, res, lunes))
    print(f"  · elite/index.html")

    hechos, sin_pin = 0, []
    for f in compiten:
        pin = pins.get(f['slug'])
        if not pin:
            sin_pin.append(f['modelo'])
            continue
        d = os.path.join(OUT_DIR, 'modelos', f['slug'])
        os.makedirs(d, exist_ok=True)
        with open(os.path.join(d, 'index.html'), 'w', encoding='utf-8') as fh:
            fh.write(html_modelo(f, compiten, res, lunes, pin))
        hechos += 1
    print(f"  · elite/modelos/… {hechos} dashboards individuales")
    if sin_pin:
        print(f"    (sin PIN registrado, no se generaron: {', '.join(sin_pin[:8])}"
              f"{' …' if len(sin_pin) > 8 else ''})")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    ap = argparse.ArgumentParser(description='Programa J&D ELITE — Modelo de la Semana')
    ap.add_argument('--semana', help='Lunes de la semana a evaluar (AAAA-MM-DD)')
    ap.add_argument('--solo-datos', action='store_true', help='No generar HTML')
    ap.add_argument('--dry-run', action='store_true', help='No escribir nada')
    ap.add_argument('--excel', help='Ruta alternativa al archivo maestro')
    args = ap.parse_args()

    path = args.excel or GE_XLSX
    print(f"Leyendo {os.path.basename(path)} …")
    datos = leer_diario(path)
    sem = agrupar_semanas(datos)
    completas = semanas_disponibles(sem)
    if not completas:
        sys.exit("No hay ninguna semana completa en el archivo.")

    if args.semana:
        try:
            lunes = date.fromisoformat(args.semana)
        except ValueError:
            sys.exit("Formato de fecha inválido. Usar AAAA-MM-DD.")
        lunes = lunes_de(lunes)
    else:
        lunes = completas[-1]

    compiten, todos = evaluar(sem, lunes)
    if not compiten:
        sys.exit(f"Ningún modelo cumple las condiciones en la semana del {rango_txt(lunes)}.")
    titulos = definir_titulos(compiten)
    res = resumen(compiten, todos, lunes)

    print(f"\nSEMANA {rango_txt(lunes)}")
    print(f"  compiten {res['compiten']} de {res['con_facturacion']} con facturación · "
          f"índice promedio {fmt(res['ids_promedio'],1)}")
    print(f"  facturación del grupo: {fmt(res['facturacion_grupo'])} créditos\n")
    for t in titulos:
        if t['modelo']:
            print(f"  {t['titulo']:<22} {t['modelo']} ({t['estudio']}) — {t['dato']}")
        else:
            print(f"  {t['titulo']:<22} — no se otorga —")
    print("\n  TOP 5")
    for f in compiten[:5]:
        print(f"   {f['puesto']:>2}. {f['modelo']:<24} {fmt(f['ids'],1):>5}  "
              f"({fmt(f['p_resultado'],1)} + {fmt(f['p_crecimiento'],1)} + "
              f"{fmt(f['p_constancia'],1)} + {fmt(f['p_evolucion'],1)})")
    al = res['alertas']
    print(f"\n  ALERTAS  caída>40%: {al['caida_fuerte']} · sin días mínimos: {al['sin_dias_minimos']}"
          f" · sin días reportados: {al['sin_dias_reportados']} · sin historial: {al['sin_historial']}")
    if al['variacion_sospechosa']:
        print(f"  Verificar (más de 3x su promedio): {', '.join(al['variacion_sospechosa'][:6])}")

    if args.dry_run:
        print("\n[dry-run] No se escribió ningún archivo.")
        return

    print("\nEscribiendo:")
    guardar(compiten, todos, titulos, res, lunes, cargar_pins(), args.solo_datos)
    print("\nListo. Revisar antes de publicar con PUBLICAR.bat")


if __name__ == '__main__':
    main()
