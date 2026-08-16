#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ACTUALIZAR.py — Actualizador dashboards Grupo J&D
==================================================
Fuente de verdad: Cómo vamos Grupo Empresarial.xlsx (Desktop LOCAL)
Periodos CPP:     CALENDARIO MAESTRO (hoja en el mismo Excel)

Qué hace:
  1. Lee hojas JULIO y AGOSTO del Excel (datos diarios por modelo/studio)
  2. Lee CALENDARIO MAESTRO (períodos por plataforma)
  3. Reconstruye ALIADOS y ESTUDIO desde cero para Julio y Agosto
  4. Reconstruye GRUPO_PERIODOS / PERIODOS desde el calendario real
  5. Recalcula TOP_EST (por estudio) y TOP20G (grupo completo) para ambos meses
  6. Inyecta 'Agosto' en var MESES y var DMES de los 6 HTML
  7. Actualiza timestamp

FÓRMULA VERIFICADA: col_F4F(d) = 19 + (31-d)*5  [5 plats/día, LOCAL]
NO commit/push hasta validación del usuario.
"""

import openpyxl, re, base64, json, os, copy
from datetime import datetime, date, timedelta

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CV_PATH = os.path.join(SCRIPT_DIR, '..', 'Centro de Gestión Estratégica Grupo J&D',
                       'COMO VAMOS GRUPO', 'Cómo vamos Grupo Empresarial.xlsx')
DASHBOARDS = {
    'grupo':  os.path.join(SCRIPT_DIR, 'grupo579780',                     'index.html'),
    'erika':  os.path.join(SCRIPT_DIR, 'erika868527',                     'index.html'),
    'fabio':  os.path.join(SCRIPT_DIR, 'fabio473013',                     'index.html'),
    'fornax': os.path.join(SCRIPT_DIR, 'estudios', 'fornax-studios345929', 'index.html'),
    'gold':   os.path.join(SCRIPT_DIR, 'estudios', 'goldonline078939',     'index.html'),
    'cyv':    os.path.join(SCRIPT_DIR, 'estudios', 'cyv-studios837357',    'index.html'),
}

MES     = 'Julio'
MES_AGO = 'Agosto'
PLATS   = ['F4F', 'SC', 'CB', 'CAM', 'STR']

# Meses ordenados con días
ALL_MESES   = ['Enero','Febrero','Marzo','Abril','Mayo','Junio','Julio','Agosto']
DMES_VALS   = {'Enero':31,'Febrero':28,'Marzo':31,'Abril':30,
               'Mayo':31,'Junio':30,'Julio':31,'Agosto':31}
MES_NUM     = {1:'Enero',2:'Febrero',3:'Marzo',4:'Abril',5:'Mayo',6:'Junio',
               7:'Julio',8:'Agosto',9:'Septiembre',10:'Octubre',11:'Noviembre',12:'Diciembre'}
MES_ABREV   = {'Ene':'Enero','Feb':'Febrero','Mar':'Marzo','Abr':'Abril','May':'Mayo',
               'Jun':'Junio','Jul':'Julio','Ago':'Agosto','Sep':'Septiembre',
               'Oct':'Octubre','Nov':'Noviembre','Dic':'Diciembre'}
MONTH_ESP   = {1:'Ene',2:'Feb',3:'Mar',4:'Abr',5:'May',6:'Jun',
               7:'Jul',8:'Ago',9:'Sep',10:'Oct',11:'Nov',12:'Dic'}

# Plataforma → columna en CALENDARIO MAESTRO (0-indexed, R4 header)
CAL_COL = {'F4F':0, 'Chaturbate':1, 'Stripchat':2, 'CamSoda':3, 'Streamate':4}
# Display name → internal key
CAL_PLAT_DISPLAY = {
    'F4F':        ('F4F','F4F'),
    'Chaturbate': ('Chaturbate','CB'),
    'Stripchat':  ('Stripchat','SC'),
    'CamSoda':    ('CamSoda','CAM'),
    'Streamate':  ('Streamate','STR'),
}

TODAY = date.today()


# ═══════════════════════════════════════════════════════════════════
# FÓRMULA DE COLUMNA
# ═══════════════════════════════════════════════════════════════════
def day_col(d):
    return 19 + (31 - d) * 5


# ═══════════════════════════════════════════════════════════════════
# BASE64 HELPERS
# ═══════════════════════════════════════════════════════════════════
def b64enc(obj):
    return base64.b64encode(
        json.dumps(obj, ensure_ascii=False).encode('utf-8')
    ).decode('ascii')

def b64dec(s):
    return json.loads(base64.b64decode(s).decode('utf-8'))

def get_var(html, varname):
    for q in ['"', "'"]:
        m = re.search(rf"var {varname}\s*=\s*_b64dec\({q}([^{q}]+){q}\)", html)
        if m:
            try:
                return b64dec(m.group(1)), q
            except Exception:
                pass
    return None, None

def set_var(html, varname, data, quote):
    new_b64 = b64enc(data)
    pattern = rf"(var {varname}\s*=\s*_b64dec\(){quote}[^{quote}]+{quote}(\))"
    result, n = re.subn(pattern, rf'\g<1>{quote}{new_b64}{quote}\g<2>', html, count=1)
    if n == 0:
        print(f"  ⚠  No se encontró var {varname} (quote={quote})")
    return result


# ═══════════════════════════════════════════════════════════════════
# INYECCIÓN DE MESES EN JS
# ═══════════════════════════════════════════════════════════════════
def patch_meses_dmes(html, meses_list):
    """
    Reemplaza var MESES=[...] y var DMES={...} en el HTML
    para incluir todos los meses en meses_list.
    """
    meses_js   = "var MESES=[" + ",".join(f"'{m}'" for m in meses_list) + "];"
    dmes_parts = ",".join(f"{m}:{DMES_VALS[m]}" for m in meses_list if m in DMES_VALS)
    dmes_js    = "var DMES={" + dmes_parts + "};"

    html = re.sub(r"var MESES=\[[^\]]*\];", meses_js, html, count=1)
    html = re.sub(r"var DMES=\{[^}]*\};",  dmes_js,  html, count=1)
    return html


# ═══════════════════════════════════════════════════════════════════
# TIMESTAMP
# ═══════════════════════════════════════════════════════════════════
def inject_timestamp(html, ts_str, cutoff_str=None):
    cutoff_line = f'<br>Datos al: {cutoff_str}' if cutoff_str else ''
    badge = f'<span class="nav-update">Actualizado: {ts_str}{cutoff_line}</span>'
    css   = '.nav-update{display:block;font-size:9px;color:var(--muted);margin-top:1px;line-height:1.6;text-align:right}'
    if '.nav-update' not in html:
        html = html.replace('</style>', css + '\n</style>', 1)

    # Limpieza robusta: eliminar nav-update spans con contenido anidado
    # (conteo de profundidad para manejar spans dentro de spans)
    while '<span class="nav-update">' in html:
        start = html.find('<span class="nav-update">')
        depth, pos = 0, start
        while pos < len(html):
            if html[pos:pos+5] == '<span':
                depth += 1; pos += 5
            elif html[pos:pos+7] == '</span>':
                depth -= 1
                if depth == 0:
                    html = html[:start] + html[pos+7:]; break
                pos += 7
            else:
                pos += 1
        else:
            break  # evitar loop infinito si el HTML está malformado

    # Eliminar residuos sueltos de "Datos al:" que quedaron fuera del span
    html = re.sub(r'<span[^>]*>Datos al:[^<]*</span>', '', html)

    # Insertar el nuevo badge una sola vez
    html = html.replace('</div></nav>', badge + '</div></nav>', 1)
    return html


# ═══════════════════════════════════════════════════════════════════
# LEER EXCEL — HOJA DE DATOS DIARIOS
# ═══════════════════════════════════════════════════════════════════
def read_cv_sheet(wb, sheet):
    """
    Lee una hoja del workbook en RAM (iter_rows) → evita timeouts.
    Devuelve: cv = {(model, studio): {day_int: {plat: float|None}}}, last_day
    """
    if sheet not in wb.sheetnames:
        print(f"  ⚠  Hoja '{sheet}' no encontrada.")
        return {}, 0
    ws = wb[sheet]
    data = tuple(ws.iter_rows(values_only=True))
    MAX_COL = ws.max_column
    print(f"  📋 [{sheet}] {len(data)}r × {MAX_COL}c")

    # Verify structure
    ok = str(data[2][18] or '') == '31' and 'lirt' in str(data[3][18] or '').lower()
    print(f"  📋 Col19: día={data[2][18]!r} plat={data[3][18]!r}  {'✅' if ok else '❌'}")

    IGNORAR_STUDIOS = {
        'CAMSODA','CAMSODA POR QUINCENA','CHATURBATE','FLIRT4FREE',
        'STREAMATE','STREAMATE POR QUINCENA','STRIPCHAT','STRIPCHAT POR QUINCENA',
        'TOTAL','TOTALES','CAMSODA POR QUINCENA','STREAMATE POR QUINCENA','STRIPCHAT POR QUINCENA',
    }
    cv = {}; last_day_excel = 0; filas_sin_nombre = []
    for ri in range(4, len(data)):
        row    = data[ri]
        model  = str(row[0] or '').strip()
        studio = str(row[1] or '').strip()
        if not studio:
            continue
        if studio.upper() in IGNORAR_STUDIOS or studio.upper() in {s.upper() for s in IGNORAR_STUDIOS}:
            continue
        # Detectar filas con studio pero sin nombre de modelo que tengan producción
        if not model:
            tiene_prod = any(
                isinstance(row[day_col(d) + pi - 1], (int, float)) and float(row[day_col(d) + pi - 1] or 0) > 0
                for d in range(1, 32) for pi in range(5)
                if day_col(d) + pi - 1 < len(row)
            )
            if tiene_prod:
                model = 'Total Estudio'  # fila agregada de estudio — asignar pseudónombre
            else:
                continue
        if model.upper() in {'TOTAL','TOTALES','MODELO'}:
            continue
        entry = {}
        for d in range(1, 32):
            c0 = day_col(d)
            if c0 > MAX_COL:
                break
            day_data = {}; has = False
            for pi, p in enumerate(PLATS):
                ci = c0 + pi - 1
                if ci >= len(row):
                    day_data[p] = None; continue
                v = row[ci]
                if isinstance(v, (int, float)) and float(v) > 0:
                    val = float(v)
                    if p == 'STR':          # Streamate: USD → créditos (÷ 0.05 = × 20)
                        val = val / 0.05
                    day_data[p] = val; has = True
                    last_day_excel = max(last_day_excel, d)
                else:
                    day_data[p] = None
            if has:
                entry[d] = day_data
        if entry:
            cv[(model, studio)] = entry
    print(f"  📅 Último día: {last_day_excel} | Modelos: {len(cv)}")
    if filas_sin_nombre:
        print(f"  ⚠  FILAS CON PRODUCCIÓN PERO SIN NOMBRE — completar en Excel:")
        for msg in filas_sin_nombre:
            print(msg)
    return cv, last_day_excel


# ═══════════════════════════════════════════════════════════════════
# LEER CALENDARIO MAESTRO
# ═══════════════════════════════════════════════════════════════════
def read_calendario_maestro(wb):
    """
    Lee el CALENDARIO MAESTRO y devuelve:
    {
      'F4F':        [{'label':'27 Jul – 09 Ago','start':'27/07/2026','end':'09/08/2026'}, ...],
      'Chaturbate': [...],
      'Stripchat':  [...],
      'CamSoda':    [...],
      'Streamate':  [...],
    }
    Solo incluye periodos cuya fecha de INICIO <= hoy.
    """
    if 'CALENDARIO MAESTRO' not in wb.sheetnames:
        print("  ⚠  Hoja 'CALENDARIO MAESTRO' no encontrada.")
        return {}

    ws   = wb['CALENDARIO MAESTRO']
    rows = list(ws.iter_rows(values_only=True))

    # Row 4 (index 3) = headers, Row 5+ (index 4+) = period ranges
    col_order = ['F4F', 'Chaturbate', 'Stripchat', 'CamSoda', 'Streamate']
    result    = {k: [] for k in col_order}

    def parse_range(cell_str):
        """'27/07/2026 - 09/08/2026' → (start_date, end_date, start_str, end_str)"""
        if not cell_str or not isinstance(cell_str, str):
            return None
        cell_str = cell_str.strip()
        m = re.match(r'(\d{2}/\d{2}/\d{4})\s*[-–]\s*(\d{2}/\d{2}/\d{4})', cell_str)
        if not m:
            return None
        s_str, e_str = m.group(1), m.group(2)
        try:
            sd = datetime.strptime(s_str, '%d/%m/%Y').date()
            ed = datetime.strptime(e_str, '%d/%m/%Y').date()
            return sd, ed, s_str, e_str
        except ValueError:
            return None

    def fmt_label(sd, ed):
        """27/07/2026 → '27 Jul – 09 Ago'"""
        return f"{sd.day:02d} {MONTH_ESP[sd.month]} – {ed.day:02d} {MONTH_ESP[ed.month]}"

    for row in rows[4:]:  # skip header rows
        for ci, plat in enumerate(col_order):
            if ci >= len(row):
                continue
            cell_val = row[ci]
            if cell_val is None:
                continue
            parsed = parse_range(str(cell_val))
            if parsed is None:
                continue
            sd, ed, s_str, e_str = parsed
            # Only include periods whose START is <= today
            if sd > TODAY:
                continue
            result[plat].append({
                'label': fmt_label(sd, ed),
                'start': s_str,
                'end':   e_str,
                'models': {},
                'total':  0,
            })

    for plat, perds in result.items():
        print(f"  📅 {plat}: {len(perds)} periodos")
    return result


# ═══════════════════════════════════════════════════════════════════
# CÁLCULO DE TOTALES POR PERÍODO
# ═══════════════════════════════════════════════════════════════════
def _sum_period_from_aliados(all_mods_by_month, s_str, e_str, plat_key):
    """
    Suma producción de plat_key para el rango s_str→e_str usando
    all_mods_by_month = {mes_name: {model: {day_int: {plat: val}}}}.
    Devuelve {model: total}.
    """
    try:
        sd = datetime.strptime(s_str, '%d/%m/%Y').date()
        ed = datetime.strptime(e_str, '%d/%m/%Y').date()
    except ValueError:
        return {}

    new_models = {}
    # Iterate month by month within the period
    cur_month  = sd.replace(day=1)
    while cur_month <= ed.replace(day=1):
        mes_name = MES_NUM.get(cur_month.month)
        if mes_name and mes_name in all_mods_by_month:
            # Day range for this month within the period
            d_start = sd.day if (cur_month.year == sd.year and cur_month.month == sd.month) else 1
            import calendar as _cal
            last_in_month = _cal.monthrange(cur_month.year, cur_month.month)[1]
            d_end   = ed.day if (cur_month.year == ed.year and cur_month.month == ed.month) else last_in_month

            for model, days in all_mods_by_month[mes_name].items():
                for d in range(d_start, d_end + 1):
                    v = (days.get(d) or {}).get(plat_key) or 0
                    if v:
                        new_models[model] = new_models.get(model, 0) + v

        # Next month
        if cur_month.month == 12:
            cur_month = cur_month.replace(year=cur_month.year + 1, month=1)
        else:
            cur_month = cur_month.replace(month=cur_month.month + 1)

    return new_models


def _build_all_mods_by_month(aliados, aliados_key=None):
    """
    Construye {mes_name: {model: {day_int: {plat:val}}}} desde ALIADOS.
    Si aliados_key no es None, restringe a ese studio.
    """
    all_mods = {}
    scope = aliados if aliados_key is None else {aliados_key: aliados.get(aliados_key, {})}
    for ak, ainfo in scope.items():
        for mes, md in (ainfo.get('data') or {}).items():
            if mes not in all_mods:
                all_mods[mes] = {}
            for model, days in (md.get('modelos') or {}).items():
                if model not in all_mods[mes]:
                    all_mods[mes][model] = {}
                for dk, dv in days.items():
                    if dv:
                        all_mods[mes][model][int(dk)] = dv
    return all_mods


def build_grupo_periodos(cal, aliados, meta=None):
    """
    Construye GRUPO_PERIODOS completo desde el calendario y los datos de ALIADOS.
    Preserva periods históricos con sus totales originales si existen en el html.
    """
    plat_key_map = {'F4F':'F4F','Chaturbate':'CB','Stripchat':'SC','CamSoda':'CAM','Streamate':'STR'}
    all_mods_by_month = _build_all_mods_by_month(aliados)

    gp = {}
    for plat_display, perds in cal.items():
        plat_key = plat_key_map.get(plat_display, plat_display)
        out = []
        for p in perds:
            new_models = _sum_period_from_aliados(all_mods_by_month, p['start'], p['end'], plat_key)
            p['models'] = new_models
            p['total']  = sum(new_models.values())
            out.append(p)
        gp[plat_display] = out

    # Auto-calcular streamate_active desde datos reales:
    # true si existe al menos un periodo STR con total > 0
    str_has_data = any(p.get('total', 0) > 0 for p in gp.get('Streamate', []))

    if meta:
        new_meta = dict(meta)   # nunca mutar el original
        new_meta['streamate_active'] = str_has_data
        gp['_meta'] = new_meta
    else:
        gp['_meta'] = {'studio': 'Grupo Empresarial J&D', 'entity_label': 'Estudio',
                       'streamate_active': str_has_data}
    return gp


def build_studio_periodos(cal, aliados_entry):
    """
    Construye PERIODOS para un dashboard de estudio.
    """
    plat_key_map = {'F4F':'F4F','Chaturbate':'CB','Stripchat':'SC','CamSoda':'CAM','Streamate':'STR'}
    all_mods_by_month = {}
    for mes, md in (aliados_entry.get('data') or {}).items():
        all_mods_by_month[mes] = {}
        for model, days in (md.get('modelos') or {}).items():
            all_mods_by_month[mes][model] = {int(dk): dv for dk, dv in days.items() if dv}

    periodos = {}
    for plat_display, perds in cal.items():
        plat_key = plat_key_map.get(plat_display, plat_display)
        out = []
        for p in perds:
            new_models = _sum_period_from_aliados(all_mods_by_month, p['start'], p['end'], plat_key)
            # filter zero-models
            new_models = {k: v for k, v in new_models.items() if v > 0}
            out.append({
                'label':  p['label'],
                'start':  p['start'],
                'end':    p['end'],
                'models': new_models,
                'total':  sum(new_models.values()),
            })
        periodos[plat_display] = out
    return periodos


# ═══════════════════════════════════════════════════════════════════
# REBUILD ALIADOS
# ═══════════════════════════════════════════════════════════════════
def rebuild_aliados_from_excel(aliados, cv, key_map, mes, last_day, detect_new=True):
    """Reconstruye datos diarios del ALIADOS para el mes. Crea el mes si no existe.
    detect_new=True  → detecta e incorpora modelos nuevos del Excel (GRUPO_MAP)
    detect_new=False → para aliados individuales (ERIKA_MAP / FABIO_MAP).
                       Dentro del False, se activa auto-detect cuando cv_studios == [ak]
                       (entradas tipo-estudio donde el aliado key == nombre del estudio)."""
    cleared = 0; updated = 0; nuevos = 0
    for ak, ainfo in aliados.items():
        cv_studios = key_map.get(ak)
        if cv_studios is None:
            continue
        data = ainfo.setdefault('data', {})
        if mes not in data:
            ref_mods = list((data.get(MES, {}).get('modelos') or {}).keys())
            data[mes] = {'dias': 0, 'modelos': {m: {} for m in ref_mods}}
            print(f"  ➕ Creado {mes} en ALIADOS[{ak}] ({len(ref_mods)} modelos)")

        md      = data[mes]
        modelos = md.get('modelos')
        if modelos is None:
            continue
        # Limpiar datos diarios de modelos existentes
        for model in list(modelos.keys()):
            modelos[model] = {}; cleared += 1
        # Reconstruir datos de modelos existentes
        for model in list(modelos.keys()):
            for studio in cv_studios:
                key = (model, studio)
                if key not in cv:
                    continue
                for d, day_vals in cv[key].items():
                    entry = {p: (day_vals.get(p) if day_vals.get(p, 0) else None) for p in PLATS}
                    modelos[model][str(d)] = entry; updated += 1
                break
        # ── DETECCIÓN AUTOMÁTICA DE MODELOS NUEVOS ──────────────────
        # Activa cuando: detect_new=True (GRUPO_MAP)
        # O cuando: detect_new=False pero cv_studios==[ak] (entrada tipo-estudio en ERIKA/FABIO)
        is_studio_entry = (cv_studios == [ak])
        if detect_new or is_studio_entry:
            for (model, studio), day_entries in cv.items():
                if studio not in cv_studios:
                    continue
                if model in modelos:
                    continue  # ya existe
                modelos[model] = {}
                for d, day_vals in day_entries.items():
                    entry = {p: (day_vals.get(p) if day_vals.get(p, 0) else None) for p in PLATS}
                    modelos[model][str(d)] = entry; updated += 1
                print(f"  🆕 [{mes}] Nuevo modelo → ALIADOS[{ak}]: '{model}'")
                nuevos += 1
        # Eliminar modelos sin datos que puedan ser artefactos de runs anteriores
        if is_studio_entry:
            stale = [m for m, days in modelos.items() if not days]
            for m in stale:
                del modelos[m]
                if stale: print(f"  🗑  [{mes}] Eliminado artefacto sin datos: ALIADOS[{ak}]['{m}']")
        # ────────────────────────────────────────────────────────────
        md['dias'] = last_day
    print(f"  🗑  [{mes}] Limpiados: {cleared} | Reconstruidos: {updated} | Nuevos: {nuevos}")
    return aliados


# ═══════════════════════════════════════════════════════════════════
# CORRECCIÓN ALIADOS INDIVIDUALES DE EJECUTIVOS
# ═══════════════════════════════════════════════════════════════════
def fix_exec_individual_partners(aliados, key_map):
    """
    Para entradas INDIVIDUALES del ejecutivo (donde la clave del aliado es el
    nombre del modelo, no del estudio, p.ej. 'Dulce Luna' → ['Fornax Studios']),
    asegura que el dict modelos SOLO contenga el modelo cuyo nombre coincide
    con la clave del aliado.

    Problema que resuelve:
      rebuild_aliados_from_excel reutiliza los modelos ya presentes en el dict.
      Si en una ejecución anterior se contaminó la entrada con modelos extra del
      mismo estudio (ej: todos los modelos de Fornax dentro de la clave 'Dulce Luna'),
      cada ejecución posterior los reconstruye con datos reales, perpetuando el error.

    La regla: la clave 'X' → solo puede tener el modelo llamado 'X'.
    Se aplica a TODOS los meses presentes en la entrada.
    No elimina datos del archivo fuente — solo filtra qué puede verse en cada dashboard.
    """
    fixed = 0
    for ak, cv_studios in key_map.items():
        # Entrada tipo-estudio: cv_studios == [ak] → no tocar
        if cv_studios == [ak]:
            continue
        if ak not in aliados:
            continue
        for mes, md in (aliados[ak].get('data') or {}).items():
            modelos = md.get('modelos')
            if not modelos:
                continue
            extra = [m for m in list(modelos.keys()) if m != ak]
            for m in extra:
                del modelos[m]
                fixed += 1
    if fixed:
        print(f"  🔒 fix_exec_individual_partners: {fixed} modelo(s) extra removido(s) "
              f"(datos de otros estudios no autorizados para este ejecutivo)")
    else:
        print(f"  🔒 fix_exec_individual_partners: sin contaminación detectada")
    return aliados


# ═══════════════════════════════════════════════════════════════════
# PROPAGACIÓN DESDE GRUPO — FUENTE ÚNICA DE VERDAD
# ═══════════════════════════════════════════════════════════════════
def propagate_from_grupo(aliados_target, aliados_grupo, key_map, mes):
    """Propaga datos del mes desde Grupo al dashboard de un ejecutivo.

    GARANTÍA: cualquier aliado/estudio que existe TANTO en key_map como en
    aliados_grupo recibe exactamente los mismos datos que tiene Grupo.
    No se recalcula nada desde el Excel — se copia directamente desde la
    estructura ya calculada de Grupo (fuente central).

    Esto resuelve el problema donde Grupo podía mostrar un valor diferente
    al de Erika/Fabio para el mismo estudio, porque cada dashboard hacía un
    rebuild independiente desde el Excel con detect_new=False.

    Flujo correcto:
      Excel → Grupo (rebuild completo, detect_new=True)
             ↓  propagate_from_grupo
      Erika (datos idénticos para estudios compartidos)
             ↓  propagate_from_grupo
      Fabio (datos idénticos para estudios compartidos)

    Para aliados individuales (Alice Steel, Dulce Luna, etc.) que existen en
    key_map pero NO en aliados_grupo, la función no hace nada — esos ya fueron
    rebuilt correctamente desde el Excel en el paso anterior."""
    propagated = 0
    skipped_ind = 0
    for ak in key_map:
        if ak not in aliados_target:
            continue   # aliado no existe en el dashboard del ejecutivo
        if ak not in aliados_grupo:
            skipped_ind += 1
            continue   # aliado individual no está en Grupo — no propagar
        grupo_data = aliados_grupo[ak].get('data', {}).get(mes)
        if grupo_data is None:
            continue   # Grupo no tiene datos para este mes todavía
        aliados_target[ak].setdefault('data', {})[mes] = copy.deepcopy(grupo_data)
        propagated += 1
    print(f"  🔄 [{mes}] Propagados desde Grupo: {propagated} aliados | "
          f"individuales (excluidos): {skipped_ind}")
    return aliados_target


# ═══════════════════════════════════════════════════════════════════
# REBUILD ESTUDIO
# ═══════════════════════════════════════════════════════════════════
def rebuild_estudio_from_excel(estudio, cv, cv_studio, last_day, mes=MES):
    """Actualiza ESTUDIO.data[mes] desde Excel. Crea el mes si no existe.
    Incorpora automáticamente modelos nuevos detectados en 'Cómo Vamos'."""
    data = estudio.setdefault('data', {})
    if mes not in data:
        ref_mods = list((data.get(MES, {}).get('modelos') or {}).keys())
        data[mes] = {'dias': 0, 'modelos': {m: {} for m in ref_mods}}
        print(f"  ➕ Creado {mes} en ESTUDIO [{cv_studio}]")

    md      = data[mes]
    modelos = md.get('modelos') or {}
    cl = 0; up = 0; nuevos = 0
    # Limpiar datos diarios de modelos existentes
    for model in list(modelos.keys()):
        modelos[model] = {}; cl += 1
    # Reconstruir datos de modelos existentes
    for model in list(modelos.keys()):
        key = (model, cv_studio)
        if key not in cv:
            continue
        for d, day_vals in cv[key].items():
            entry = {p: (day_vals.get(p) if day_vals.get(p, 0) else None) for p in PLATS}
            modelos[model][str(d)] = entry; up += 1
    # ── DETECCIÓN AUTOMÁTICA DE MODELOS NUEVOS ──────────────────────
    for (model, studio), day_entries in cv.items():
        if studio != cv_studio:
            continue
        if model in modelos:
            continue  # ya existe
        modelos[model] = {}
        for d, day_vals in day_entries.items():
            entry = {p: (day_vals.get(p) if day_vals.get(p, 0) else None) for p in PLATS}
            modelos[model][str(d)] = entry; up += 1
        print(f"  🆕 [{mes}] Nuevo modelo incorporado → ESTUDIO[{cv_studio}]: '{model}'")
        nuevos += 1
    # ────────────────────────────────────────────────────────────────
    md['modelos'] = modelos
    md['dias'] = last_day
    print(f"  ✓ ESTUDIO [{cv_studio}] {mes}: {cl} limpiados, {up} entradas | nuevos={nuevos} | dias={last_day}")
    return estudio


# ═══════════════════════════════════════════════════════════════════
# CORRECCIÓN ALIADOS INDIVIDUALES EN GRUPO
# ═══════════════════════════════════════════════════════════════════
# Para cada aliado individual de Fabio en el GRUPO dashboard,
# mantener SOLO el modelo cuyo nombre coincide con la clave.
# Esto evita que detect_new=True agregue todos los modelos de Fabio
# a cada clave individual.
GRUPO_INDIVIDUAL_PARTNERS = {
    'Alice Steel':   'Fabio Robledo',
    'Eli Cortes':    'Fabio Robledo',
    'Evelyn Lovers': 'Fabio Robledo',
    'Jack Miller':   'Fabio Robledo',
    'Maximus Clark': 'Fabio Robledo',
    'Amanda Bond':   'Fabio Robledo',
    'Yessie Jacobs': 'Fabio Robledo',
    'Ana Black':     'Fabio Robledo',
}

def fix_grupo_individual_partners(aliados, cv_jul, last_day_jul, cv_ago, last_day_ago):
    """
    Reconstruye las claves individuales de Fabio en el GRUPO ALIADOS:
    cada clave (Alice Steel, Eli Cortes, ...) recibe SOLO su propio modelo
    con datos reales del Excel (cv), descartando todos los demás modelos.
    """
    for partner, cv_studio in GRUPO_INDIVIDUAL_PARTNERS.items():
        if partner not in aliados:
            continue
        if 'data' not in aliados[partner]:
            aliados[partner]['data'] = {}
        # Procesar Julio y Agosto
        for mes, cv, last_day in [('Julio', cv_jul, last_day_jul), ('Agosto', cv_ago, last_day_ago)]:
            partner_data = cv.get((partner, cv_studio), {})
            day_entries = {}
            for d in range(1, last_day + 1):
                if d in partner_data and any(v for v in partner_data[d].values() if v):
                    day_entries[str(d)] = partner_data[d]
            # Reemplazar completamente los modelos de este mes con solo el propio
            aliados[partner]['data'][mes] = {'modelos': {partner: day_entries} if day_entries else {}}
    print(f"  ✅ fix_grupo_individual_partners: {len(GRUPO_INDIVIDUAL_PARTNERS)} claves corregidas")
    return aliados


# ═══════════════════════════════════════════════════════════════════
# TOP_EST y TOP20G
# ═══════════════════════════════════════════════════════════════════
def recalc_top_est(top_est, aliados_entry, mes=MES):
    md       = (aliados_entry.get('data') or {}).get(mes, {})
    all_mods = md.get('modelos') or {}
    totals = {}
    for model, days in all_mods.items():
        t = sum(sum((dv.get(p) or 0) for p in PLATS) for dv in days.values() if dv)
        if t > 0:
            totals[model] = t
    ranked   = sorted(totals.items(), key=lambda x: -x[1])
    new_list = [{'modelo': m} for m, _ in ranked]
    if top_est is None:
        return {mes: new_list}
    top_est[mes] = new_list
    return top_est

PLAT_JS = {'F4F': 'f4f', 'SC': 'sc', 'CB': 'cb', 'CAM': 'cam', 'STR': 'str'}

def recalc_top_eje(top_eje, aliados, mes):
    """Recalcula TOP_EJE de un ejecutivo desde su ALIADOS para el mes dado."""
    totals = {}
    for ak, ainfo in aliados.items():
        md = (ainfo.get('data') or {}).get(mes, {})
        for model, days in (md.get('modelos') or {}).items():
            pt = {p: 0.0 for p in PLATS}
            for dv in days.values():
                if dv:
                    for p in PLATS:
                        pt[p] += (dv.get(p) or 0)
            total = sum(pt.values())
            if total > 0:
                # Si el modelo aparece en varios ALIADOS-keys, tomar el mayor total
                if model not in totals or totals[model]['total'] < total:
                    entry = {'modelo': model, 'studio': ak, 'total': total}
                    for p in PLATS:
                        entry[PLAT_JS[p]] = pt[p]
                    totals[model] = entry
    ranked = sorted(totals.values(), key=lambda x: -x['total'])
    print(f"    TOP_EJE [{mes}]: {len(ranked)} | #1={ranked[0]['modelo'] if ranked else '—'}")
    if top_eje is None:
        return {mes: ranked}
    top_eje[mes] = ranked
    return top_eje

ALIADOS_DISPLAY = {
    'Fornax Studios':    'Fornax Studios',
    'agatha_studios_':   'Agatha Studio',
    'goldonline':        'Gold Online',
    'kama_studio':       'Kama Studio',
    'the_Room_studios':  'The Room Studio',
    'CyV Studios':       'CyV Studios',
    'Studio Levi':       'Studios Levi',
    'The Online Agency': 'The Online Agency',
    'Elite Cam House':   'Elite Cam House',
    'Studio RWB':        'Studios RWB',
    'PrestigeCam':       'Prestige Cam',
    'Atelier_glamour':   'Atelier Glamour',
    'Dejavu Studio':     'Dejavu Studio',
    'Dynasty_studio_':   'Dynasty Studio',
    'piscis_studio':     'Piscis Studio',
    # Aliados Fabio Robledo
    'Amadeus Studio':    'Amadeus Studio',
    'Black Card':        'Black Card',
    'Studio JGM':        'Studio JGM',
    'Iridium Studio':    'Iridium Studio',
    'Alice Steel':       'Alice Steel',
    'Eli Cortes':        'Eli Cortes',
    'Evelyn Lovers':     'Evelyn Lovers',
    'Jack Miller':       'Jack Miller',
    'Maximus Clark':     'Maximus Clark',
    'Amanda Bond':       'Amanda Bond',
    'Yessie Jacobs':     'Yessie Jacobs',
    'Ana Black':         'Ana Black',
}

def recalc_top20g(top20g, grupo_aliados, mes=MES):
    all_totals = {}
    for ak, ainfo in grupo_aliados.items():
        studio_name = ALIADOS_DISPLAY.get(ak, ak)
        md = (ainfo.get('data') or {}).get(mes, {})
        for model, days in (md.get('modelos') or {}).items():
            t = sum(sum((dv.get(p) or 0) for p in PLATS) for dv in days.values() if dv)
            if t > 0:
                key = (model, studio_name)
                all_totals[key] = all_totals.get(key, 0) + t
    ranked   = sorted(all_totals.items(), key=lambda x: -x[1])[:20]
    new_list = [{'modelo': m, 'studio': s} for (m, s), _ in ranked]
    print(f"    TOP20G [{mes}]: {len(new_list)} | #1={new_list[0]['modelo'] if new_list else '—'}")
    if top20g is None:
        return {mes: new_list}
    top20g[mes] = new_list
    return top20g


# ═══════════════════════════════════════════════════════════════════
# MAPEOS ALIADOS KEY → CV STUDIOS
# ═══════════════════════════════════════════════════════════════════
GRUPO_MAP = {
    'Fornax Studios':    ['Fornax Studios'],
    'agatha_studios_':   ['Agatha Studio'],
    'goldonline':        ['Gold Online'],
    'kama_studio':       ['Kama Studio', 'GrupoJ&D'],
    'the_Room_studios':  ['The Room Studio', 'The Room Studios'],
    'CyV Studios':       ['CyV Studios'],
    'Studio Levi':       ['Studios Levi'],
    'The Online Agency': ['The Online Agency'],
    'Elite Cam House':   ['Elite Cam House'],
    'Studio RWB':        ['Studios RWB'],
    'PrestigeCam':       ['Prestige Cam'],
    'Atelier_glamour':   ['Atelier Glamour'],
    'Dejavu Studio':     ['Dejavu Studio'],
    'Dynasty_studio_':   [],
    # Estudios nuevos bajo Fabio Robledo
    'Amadeus Studio':    ['Amadeus Studio'],
    'Black Card':        ['Black Card'],
    'Studio JGM':        ['Studio JGM'],
    'Iridium Studio':    [],
    # Nota: aliados individuales ('Alice Steel', 'Eli Cortes', etc.) son corregidos
    # por fix_grupo_individual_partners() — NO agregar aquí con detect_new=True
    'piscis_studio':     ['Piscis Studio'],
}
ERIKA_MAP = {
    'CyV Studios':       ['CyV Studios'],
    'Studio Levi':       ['Studios Levi'],
    'The Online Agency': ['The Online Agency'],
    'Elite Cam House':   ['Elite Cam House'],
    'Studio RWB':        ['Studios RWB'],
    'PrestigeCam':       ['Prestige Cam'],
    'Dulce Luna':        ['Fornax Studios'],
    'Liam Terrier':      ['Fornax Studios'],
    'Zac Levis':         ['Gold Online'],
    'Joel Souza':        ['Erika Noguera'],
    'William Gardener':  ['Erika Noguera'],
}
FABIO_MAP = {
    'Atelier_glamour':   ['Atelier Glamour'],
    'Dejavu Studio':     ['Dejavu Studio'],
    'Dynasty_studio_':   [],
    'piscis_studio':     ['Piscis Studio'],
    'Alice Steel':       ['Fabio Robledo'],
    'Eli Cortes':        ['Fabio Robledo'],
    'Evelyn Lovers':     ['Fabio Robledo'],
    'Jack Miller':       ['Fabio Robledo'],
    'Maximus Clark':     ['Fabio Robledo'],
    'Amanda Bond':       ['Fabio Robledo'],
    'Yessie Jacobs':     ['Fabio Robledo'],
    'Ana Black':         ['Fabio Robledo'],
    'Amadeus Studio':    ['Amadeus Studio'],
    'Black Card':        ['Black Card'],
    'Iridium Studio':    [],
    'Studio JGM':        ['Studio JGM'],
}


# ═══════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════
def main():
    now         = datetime.now()
    ts_str      = now.strftime('%d/%m/%Y — %H:%M')
    cutoff_str  = '15/08/2026'  # Corte explícito
    meses_activos = ALL_MESES  # Enero → Agosto

    print('=' * 65)
    print(f'  ACTUALIZADOR J&D — {ts_str}')
    print(f'  MESES: {" | ".join(meses_activos)}')
    print(f'  Fórmula: day_col(d) = 19+(31-d)*5  [5 plats/día]')
    print('=' * 65)

    # ── Cargar Excel ──────────────────────────────────────────────
    print(f'\n📂 Cargando: {os.path.basename(CV_PATH)}')
    wb = openpyxl.load_workbook(CV_PATH, data_only=True)

    cv_jul, last_day_jul = read_cv_sheet(wb, 'JULIO')
    cv_ago, last_day_ago = read_cv_sheet(wb, 'AGOSTO')
    last_day_ago = min(last_day_ago, 15)  # Corte explícito: 15 de agosto

    print(f'\n📅 JULIO: último día={last_day_jul} | AGOSTO: último día={last_day_ago} (corte=15)')

    if last_day_jul == 0:
        print("⚠  Sin datos en JULIO — verificar Excel."); return

    # ── Leer calendario de periodos ───────────────────────────────
    print('\n📆 Leyendo CALENDARIO MAESTRO...')
    cal = read_calendario_maestro(wb)

    # ── Helpers ───────────────────────────────────────────────────
    def load_dash(path, al_varname):
        with open(path, 'r', encoding='utf-8') as f:
            html = f.read()
        aliados, al_q = get_var(html, al_varname)
        if aliados is None:
            print(f"  ⚠  No se encontró {al_varname} en {os.path.basename(os.path.dirname(path))}")
        return html, aliados, al_q

    def save_dash(path, html):
        with open(path, 'w', encoding='utf-8') as f:
            f.write(html)
        print(f"  💾 {os.path.relpath(path, SCRIPT_DIR)}")

    # ════════════════════════════════════════════════════════════
    # GRUPO EMPRESARIAL
    # ════════════════════════════════════════════════════════════
    print(f'\n▶  GRUPO EMPRESARIAL')
    html, aliados, al_q = load_dash(DASHBOARDS['grupo'], 'ALIADOS')
    if aliados is not None:
        # Rebuild datos diarios
        aliados = rebuild_aliados_from_excel(aliados, cv_jul, GRUPO_MAP, MES,     last_day_jul)
        aliados = rebuild_aliados_from_excel(aliados, cv_ago, GRUPO_MAP, MES_AGO, last_day_ago)
        # Corregir aliados individuales de Fabio: 1 modelo por clave, sin contaminación cruzada
        aliados = fix_grupo_individual_partners(aliados, cv_jul, last_day_jul, cv_ago, last_day_ago)

        # Rebuild GRUPO_PERIODOS desde el calendario completo
        gp, gp_q = get_var(html, 'GRUPO_PERIODOS')
        meta = (gp or {}).get('_meta')
        gp_new = build_grupo_periodos(cal, aliados, meta)
        for plat, perds in gp_new.items():
            if plat == '_meta':
                continue
            last = (perds or [{}])[-1]
            print(f"  ✓ GP {plat}: {len(perds)} periodos | último={last.get('label')} → {last.get('total',0):.0f}")

        # TOP20G
        top20g, t20_q = get_var(html, 'TOP20')
        if top20g is not None:
            for m in meses_activos:
                top20g = recalc_top20g(top20g, aliados, m)

        # Patch HTML
        html = set_var(html, 'ALIADOS',        aliados, al_q)
        html = set_var(html, 'GRUPO_PERIODOS', gp_new, gp_q)
        if top20g is not None:
            html = set_var(html, 'TOP20', top20g, t20_q)
        html = patch_meses_dmes(html, meses_activos)
        html = inject_timestamp(html, ts_str, cutoff_str)
        save_dash(DASHBOARDS['grupo'], html)

    # ════════════════════════════════════════════════════════════
    # ERIKA NOGUERA
    # ════════════════════════════════════════════════════════════
    print(f'\n▶  ERIKA NOGUERA')
    html, aliados_e, al_q = load_dash(DASHBOARDS['erika'], 'ALIADOS')
    if aliados_e is not None:
        # Paso 1 — rebuild entradas individuales desde Excel (Dulce Luna, Zac Levis, etc.)
        #           detect_new=False: no contaminar con modelos de otros estudios
        aliados_e = rebuild_aliados_from_excel(aliados_e, cv_jul, ERIKA_MAP, MES,     last_day_jul, detect_new=False)
        aliados_e = rebuild_aliados_from_excel(aliados_e, cv_ago, ERIKA_MAP, MES_AGO, last_day_ago, detect_new=False)
        # Paso 1b — purgar modelos extra de entradas individuales (Dulce Luna, Liam Terrier, etc.)
        #            Garantiza que cada entrada individual SOLO contiene su propio modelo,
        #            no todos los modelos del mismo estudio cv (ej: todos los de Fornax Studios)
        aliados_e = fix_exec_individual_partners(aliados_e, ERIKA_MAP)
        # Paso 2 — propagar estudios aliados DESDE GRUPO (fuente única de verdad)
        #           Garantiza que Erika muestre exactamente los mismos valores que Grupo
        #           para cualquier estudio que comparten (PrestigeCam, Studio Levi, etc.)
        aliados_e = propagate_from_grupo(aliados_e, aliados, ERIKA_MAP, MES)
        aliados_e = propagate_from_grupo(aliados_e, aliados, ERIKA_MAP, MES_AGO)

        ep, ep_q = get_var(html, 'EXEC_PERIODOS')
        gp, gp_q = get_var(html, 'GRUPO_PERIODOS')
        if ep is not None:
            ep_meta = ep.get('_meta')
            ep_new  = build_grupo_periodos(cal, aliados_e, ep_meta)
            html = set_var(html, 'EXEC_PERIODOS', ep_new, ep_q)
        if gp is not None:
            # Load grupo aliados for full GRUPO_PERIODOS
            with open(DASHBOARDS['grupo'], 'r', encoding='utf-8') as f:
                g_html = f.read()
            g_aliados, _ = get_var(g_html, 'ALIADOS')
            gp_meta = gp.get('_meta')
            gp_new  = build_grupo_periodos(cal, g_aliados or aliados_e, gp_meta)
            html = set_var(html, 'GRUPO_PERIODOS', gp_new, gp_q)

        # TOP_EJE — ranking propio Erika desde su ALIADOS
        top_eje_e, te_q_e = get_var(html, 'TOP_EJE')
        for m in meses_activos:
            top_eje_e = recalc_top_eje(top_eje_e, aliados_e, m)
        if top_eje_e is not None:
            html = set_var(html, 'TOP_EJE', top_eje_e, te_q_e)

        # TOP20 — copia del top20 del grupo (ya actualizado)
        top20_e, t20_q_e = get_var(html, 'TOP20')
        if top20_e is not None:
            with open(DASHBOARDS['grupo'], 'r', encoding='utf-8') as _gf:
                _g_html = _gf.read()
            top20g_src, _ = get_var(_g_html, 'TOP20')
            if top20g_src is not None:
                html = set_var(html, 'TOP20', top20g_src, t20_q_e)

        html = set_var(html, 'ALIADOS', aliados_e, al_q)
        html = patch_meses_dmes(html, meses_activos)
        html = inject_timestamp(html, ts_str, cutoff_str)
        save_dash(DASHBOARDS['erika'], html)

    # ════════════════════════════════════════════════════════════
    # FABIO ROBLEDO
    # ════════════════════════════════════════════════════════════
    print(f'\n▶  FABIO ROBLEDO')
    html, aliados_f, al_q = load_dash(DASHBOARDS['fabio'], 'ALIADOS')
    if aliados_f is not None:
        # Paso 1 — rebuild entradas individuales desde Excel (Alice Steel, Eli Cortes, etc.)
        aliados_f = rebuild_aliados_from_excel(aliados_f, cv_jul, FABIO_MAP, MES,     last_day_jul, detect_new=False)
        aliados_f = rebuild_aliados_from_excel(aliados_f, cv_ago, FABIO_MAP, MES_AGO, last_day_ago, detect_new=False)
        # Paso 1b — purgar modelos extra de entradas individuales (Alice Steel, Eli Cortes, etc.)
        aliados_f = fix_exec_individual_partners(aliados_f, FABIO_MAP)
        # Paso 2 — propagar estudios aliados DESDE GRUPO (fuente única de verdad)
        #           Garantiza que Fabio muestre exactamente los mismos valores que Grupo
        #           para cualquier estudio que comparten (Amadeus Studio, Black Card, etc.)
        aliados_f = propagate_from_grupo(aliados_f, aliados, FABIO_MAP, MES)
        aliados_f = propagate_from_grupo(aliados_f, aliados, FABIO_MAP, MES_AGO)

        ep, ep_q = get_var(html, 'EXEC_PERIODOS')
        if ep is not None:
            ep_meta = ep.get('_meta')
            ep_new  = build_grupo_periodos(cal, aliados_f, ep_meta)
            html = set_var(html, 'EXEC_PERIODOS', ep_new, ep_q)

        # TOP_EJE — ranking propio Fabio desde su ALIADOS
        top_eje_f, te_q_f = get_var(html, 'TOP_EJE')
        for m in meses_activos:
            top_eje_f = recalc_top_eje(top_eje_f, aliados_f, m)
        if top_eje_f is not None:
            html = set_var(html, 'TOP_EJE', top_eje_f, te_q_f)

        # TOP20 — copia del top20 del grupo (ya actualizado)
        top20_f, t20_q_f = get_var(html, 'TOP20')
        if top20_f is not None:
            with open(DASHBOARDS['grupo'], 'r', encoding='utf-8') as _gf:
                _g_html = _gf.read()
            top20g_src, _ = get_var(_g_html, 'TOP20')
            if top20g_src is not None:
                html = set_var(html, 'TOP20', top20g_src, t20_q_f)

        html = set_var(html, 'ALIADOS', aliados_f, al_q)
        html = patch_meses_dmes(html, meses_activos)
        html = inject_timestamp(html, ts_str, cutoff_str)
        save_dash(DASHBOARDS['fabio'], html)

    # ── Leer grupo actualizado (para TOP20G en estudios) ──────────
    with open(DASHBOARDS['grupo'], 'r', encoding='utf-8') as f:
        g_html = f.read()
    g_aliados, _ = get_var(g_html, 'ALIADOS')

    def update_studio(dash_key, cv_studio_name, aliados_map_key):
        """Actualiza un dashboard de estudio."""
        print(f'\n▶  {aliados_map_key.upper()} (estudio)')
        with open(DASHBOARDS[dash_key], 'r', encoding='utf-8') as f:
            html = f.read()

        # ESTUDIO
        estudio, es_q = get_var(html, 'ESTUDIO')
        if estudio is not None:
            estudio = rebuild_estudio_from_excel(estudio, cv_jul, cv_studio_name, last_day_jul, MES)
            estudio = rebuild_estudio_from_excel(estudio, cv_ago, cv_studio_name, last_day_ago, MES_AGO)
            html = set_var(html, 'ESTUDIO', estudio, es_q)

        # PERIODOS (desde calendario)
        aliados_entry = (g_aliados or {}).get(aliados_map_key, {})
        per_new = build_studio_periodos(cal, aliados_entry)
        for plat, perds in per_new.items():
            last = (perds or [{}])[-1]
            if last.get('total', 0):
                print(f"    {plat}: {len(perds)} periodos | último → {last['total']:.0f}")
        pq = get_var(html, 'PERIODOS')[1]
        html = set_var(html, 'PERIODOS', per_new, pq)

        # TOP_EST
        top_est, te_q = get_var(html, 'TOP_EST')
        if top_est is not None:
            for m in meses_activos:
                top_est = recalc_top_est(top_est, aliados_entry, m)
            html = set_var(html, 'TOP_EST', top_est, te_q)
            print(f"  ✓ TOP_EST: " +
                  " | ".join(f"{m}={len((top_est or {}).get(m,[]))}" for m in meses_activos))

        # TOP20G
        if g_aliados is not None:
            top20g, t20_q = get_var(html, 'TOP20G')
            if top20g is not None:
                for m in meses_activos:
                    top20g = recalc_top20g(top20g, g_aliados, m)
                html = set_var(html, 'TOP20G', top20g, t20_q)

        html = patch_meses_dmes(html, meses_activos)
        html = inject_timestamp(html, ts_str, cutoff_str)
        with open(DASHBOARDS[dash_key], 'w', encoding='utf-8') as f:
            f.write(html)
        print(f"  💾 {dash_key}")

    update_studio('fornax', 'Fornax Studios', 'Fornax Studios')
    update_studio('gold',   'Gold Online',    'goldonline')
    update_studio('cyv',    'CyV Studios',    'CyV Studios')

    # ── Auditoría de modelos: Excel vs Dashboard ──────────────────
    print(f'\n{"─" * 65}')
    print(f'  📋 AUDITORÍA DE MODELOS — comparación Excel vs Dashboard')
    print(f'{"─" * 65}')

    def _audit_studio(label, cv_combined, cv_studios, dash_modelos_ago):
        """Compara modelos en Excel (cualquier mes) vs dashboard."""
        excel_models = {m for (m, s), _ in cv_combined.items() if s in cv_studios}
        dash_models  = set(dash_modelos_ago.keys())
        nuevos_det   = excel_models - dash_models
        solo_dash    = dash_models - excel_models
        ok = len(nuevos_det) == 0
        estado = "✅" if ok else f"⚠ {len(nuevos_det)} nuevo(s)"
        print(f"  {estado}  {label}: Excel={len(excel_models)} | Dashboard={len(dash_models)}", end="")
        if nuevos_det:
            print(f"\n        🆕 Incorporados: {sorted(nuevos_det)}")
        else:
            print()
        if solo_dash:
            print(f"        ℹ️  Solo en dashboard (sin prod en Excel): {sorted(solo_dash)[:5]}")

    # cv combinado (julio + agosto)
    cv_combined_all = {}
    for k, v in cv_jul.items(): cv_combined_all[k] = v
    for k, v in cv_ago.items(): cv_combined_all[k] = v

    # GRUPO — por ALIADOS key
    with open(DASHBOARDS['grupo'], 'r', encoding='utf-8') as f: _gh = f.read()
    import re as _re, base64 as _b64, json as _json
    def _gv(h, v):
        for q in ['"', "'"]:
            m = _re.search(rf"var {v}\s*=\s*_b64dec\({q}([^{q}]+){q}\)", h)
            if m:
                try: return _json.loads(_b64.b64decode(m.group(1)).decode())
                except: pass
        return None
    g_al = _gv(_gh, 'ALIADOS') or {}
    for ak, cv_studs in GRUPO_MAP.items():
        if not cv_studs: continue
        dm = (g_al.get(ak, {}).get('data') or {}).get(MES_AGO, {}).get('modelos') or {}
        _audit_studio(ak, cv_combined_all, set(cv_studs), dm)

    # Estudios individuales
    for studio_label, cv_studio_name in [('Fornax Studios','Fornax Studios'),
                                          ('Gold Online','Gold Online'),
                                          ('CyV Studios','CyV Studios')]:
        dash_key = {'Fornax Studios':'fornax','Gold Online':'gold','CyV Studios':'cyv'}[studio_label]
        with open(DASHBOARDS[dash_key], 'r', encoding='utf-8') as f: _sh = f.read()
        s_est = _gv(_sh, 'ESTUDIO') or {}
        dm2 = (s_est.get('data') or {}).get(MES_AGO, {}).get('modelos') or {}
        _audit_studio(studio_label, cv_combined_all, {cv_studio_name}, dm2)

    print(f'{"─" * 65}')

    # ── Regenerar Dashboard de Monitores ─────────────────────────
    print(f'\n{"─" * 65}')
    print(f'  📊 Regenerando Dashboard de Monitores...')
    import subprocess, sys
    monitores_py = os.path.join(SCRIPT_DIR, 'MONITORES.py')
    if os.path.exists(monitores_py):
        result = subprocess.run(
            [sys.executable, monitores_py],
            capture_output=True, text=True, encoding='utf-8'
        )
        if result.returncode == 0:
            # Mostrar solo las líneas clave del output
            for line in result.stdout.splitlines():
                if any(x in line for x in ['✅','⚠','❌','Dashboard','Meses','Tamaño']):
                    print(f'  {line.strip()}')
        else:
            print(f'  ⚠  Error en MONITORES.py:')
            for line in (result.stderr or result.stdout).splitlines()[:8]:
                print(f'     {line}')
    else:
        print(f'  ⚠  MONITORES.py no encontrado en {SCRIPT_DIR}')

    # ── Resumen ───────────────────────────────────────────────────
    print(f'\n{"=" * 65}')
    print(f'  ✅ Rebuild completado — {ts_str}')
    print(f'  📅 JULIO días 1–{last_day_jul} | AGOSTO días 1–{last_day_ago}')
    print(f'  📆 Periodos CPP desde CALENDARIO MAESTRO')
    print(f'  🗓  MESES activos: {" | ".join(meses_activos)}')
    print(f'  📊 Dashboard Monitores: monitores/index.html')
    print(f'  ⚠  NO commit/push — esperar validación')
    print(f'{"=" * 65}\n')


if __name__ == '__main__':
    main()
